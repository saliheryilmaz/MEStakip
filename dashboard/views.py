from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse, HttpResponseRedirect
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, F, Case, When, DecimalField, IntegerField
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import datetime, timedelta, date
import os
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from django.urls import reverse
from django.views.decorators.http import require_POST, require_http_methods
from decimal import Decimal
from .models import Siparis, UserProfile, Notification, Transaction, TransactionCategory, Event, MalzemeHareketi, MalzemeDosya, CikmaLastik, JokerSatisDosya, JokerSatisHareketi, LastikModelBilgisi, Quotation, GarantiBelgesi
from .forms import SiparisForm, TransactionForm, MalzemeExcelUploadForm
# pandas removed - using openpyxl instead
from collections import defaultdict
from django.views.decorators.csrf import csrf_exempt
from django.utils.dateparse import parse_date
from .utils import parse_decimal_value, format_tire_size, normalize_turkish_text, create_turkish_search_variants
from functools import wraps
from django.conf import settings
from .groq_client import groq_chat_completion, GroqError

def misafir_forbidden(view_func):
    """Misafir kullanıcılarının erişimini engelleyen decorator"""
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        try:
            user_profile = UserProfile.objects.get(user=request.user)
            if user_profile.role == 'misafir':
                messages.error(request, 'Bu sayfaya erişim yetkiniz bulunmamaktadır.')
                return redirect('dashboard:cikma_lastikler')
        except UserProfile.DoesNotExist:
            pass
        return view_func(request, *args, **kwargs)
    return wrapper

@login_required
@misafir_forbidden
def index(request):
    """Dashboard ana sayfası"""
    # Sadece kontrol edilen siparişlerden lastik satış analizi verileri (sadece kullanıcının siparişleri)
    kontrol_siparisler = Siparis.objects.filter(durum='kontrol', user=request.user)
    
    # Lastik satış analizi verileri - mevsim ve araç tipi bazında
    tire_sales_data = {
        'yaz': {
            'binek': kontrol_siparisler.filter(mevsim='yaz', grup='binek').aggregate(total=Sum('adet'))['total'] or 0,
            'ticari': kontrol_siparisler.filter(mevsim='yaz', grup='ticari').aggregate(total=Sum('adet'))['total'] or 0
        },
        'kis': {
            'binek': kontrol_siparisler.filter(mevsim='kis', grup='binek').aggregate(total=Sum('adet'))['total'] or 0,
            'ticari': kontrol_siparisler.filter(mevsim='kis', grup='ticari').aggregate(total=Sum('adet'))['total'] or 0
        },
        'dort_mevsim': {
            'binek': kontrol_siparisler.filter(mevsim='dort-mevsim', grup='binek').aggregate(total=Sum('adet'))['total'] or 0,
            'ticari': kontrol_siparisler.filter(mevsim='dort-mevsim', grup='ticari').aggregate(total=Sum('adet'))['total'] or 0
        }
    }
    
    # Lastik marka dağılımı verileri - sadece kontrol edilen siparişler
    brand_distribution = (
        kontrol_siparisler
        .values('marka')
        .annotate(total_adet=Sum('adet'))
        .order_by('-total_adet')
    )
    
    # Chart için veri hazırlama
    brand_labels = []
    brand_data = []
    brand_colors = [
        '#3b82f6', '#ef4444', '#10b981', '#f59e0b', 
        '#8b5cf6', '#ec4899', '#06b6d4', '#84cc16'
    ]
    
    for i, brand in enumerate(brand_distribution):
        brand_labels.append(brand['marka'])
        brand_data.append(brand['total_adet'])
    
    # Real brand chart data from user's orders
    brand_chart_data = {
        'labels': brand_labels,
        'data': brand_data,
        'colors': brand_colors[:len(brand_labels)]
    }
    
    # Son eklenen işlemler (kullanıcının siparişlerinden son 5 kayıt - iptal edilenler hariç)
    son_islemler = Siparis.objects.filter(user=request.user).exclude(durum='iptal').order_by('-olusturma_tarihi')[:5]
    
    # Son finance işlemleri (kullanıcının işlemlerinden son 5 kayıt)
    son_finance_islemleri = Transaction.objects.filter(created_by=request.user).select_related(
        'kategori1', 'kategori1__parent', 'kategori2', 'kategori2__parent', 'kategori3', 'kategori3__parent'
    ).order_by('-created_at')[:5]
    
    # Gerçek istatistikler - sadece kullanıcının siparişleri (iptal edilenler hariç)
    aktif_siparisler = Siparis.objects.filter(user=request.user).exclude(durum='iptal')
    toplam_siparis = aktif_siparisler.count()
    toplam_ciro = aktif_siparisler.aggregate(total=Sum('toplam_fiyat'))['total'] or 0
    kontrol_edilen_siparis = kontrol_siparisler.count()
    toplam_adet = aktif_siparisler.aggregate(total=Sum('adet'))['total'] or 0
    
    # Aylık gelir/gider verileri (son 12 ay) - Transaction üzerinden (kullanıcıya göre)
    from datetime import datetime, timedelta
    monthly_income = []
    monthly_expense = []
    monthly_labels = []
    toplam_ifade = (F('nakit') + F('kredi_karti') + F('sanal_pos') + F('cari') + F('mehmet_havale'))
    
    for i in range(11, -1, -1):
        start_date = timezone.now() - timedelta(days=30*i)
        end_date = start_date + timedelta(days=30)
        
        gelir_toplam = Transaction.objects.filter(
            created_by=request.user,
            tarih__gte=start_date.date(),
            tarih__lt=end_date.date(),
            hareket_tipi='gelir'
        ).aggregate(total=Sum(toplam_ifade))['total'] or 0
        
        gider_toplam = Transaction.objects.filter(
            created_by=request.user,
            tarih__gte=start_date.date(),
            tarih__lt=end_date.date(),
            hareket_tipi='gider'
        ).aggregate(total=Sum(toplam_ifade))['total'] or 0
        
        monthly_income.append(float(gelir_toplam))
        monthly_expense.append(float(gider_toplam))
        monthly_labels.append(start_date.strftime('%b'))
    
    # En çok alım yaptığımız cariler (fiyat bazında) - Yeni Sipariş Ekle'den oluşturulan veriler
    # Sadece kontrol edilmiş siparişlerden en çok alım yapan cariler (sadece kullanıcının siparişleri)
    top_customers = (
        Siparis.objects
        .filter(durum='kontrol', user=request.user)  # Sadece kontrol edilmiş siparişler ve kullanıcının siparişleri
        .values('cari_firma')
        .annotate(
            total_purchase=Sum('toplam_fiyat'),
            toplam_adet=Sum('adet')
        )
        .order_by('-total_purchase')[:8]
    )
    
    customer_labels = []
    customer_data = []
    customer_details = []
    
    for customer in top_customers:
        # Firma adını kısalt (çok uzunsa)
        firma_name = customer['cari_firma']
        original_name = firma_name
        if len(firma_name) > 20:
            firma_name = firma_name[:17] + '...'
        
        customer_labels.append(firma_name)
        customer_data.append(float(customer['total_purchase']))
        customer_details.append({
            'name': original_name,
            'total': float(customer['total_purchase']),
            'count': customer['toplam_adet']
        })
    
    # Ödemeler analizi - Sipariş ödeme türlerine göre dağılım
    payment_stats = (
        Siparis.objects
        .filter(durum='kontrol', user=request.user)  # Sadece kontrol edilmiş siparişler
        .values('odeme')
        .annotate(
            count=Count('id'),
            total_amount=Sum('toplam_fiyat')
        )
        .order_by('-total_amount')
    )
    
    payment_data = []
    payment_colors = {
        'kredi-karti': '#3b82f6',  # Mavi
        'havale': '#10b981',       # Yeşil
        'cari': '#f59e0b',         # Sarı
    }
    
    payment_labels = {
        'kredi-karti': 'Kredi Kartı',
        'havale': 'Havale',
        'cari': 'Cari',
    }
    
    for payment in payment_stats:
        payment_type = payment['odeme']
        payment_data.append({
            'category': payment_labels.get(payment_type, payment_type.title()),
            'value': float(payment['total_amount']),
            'count': payment['count'],
            'color': payment_colors.get(payment_type, '#6b7280')
        })
    
    context = {
        'page_title': 'Dashboard',
        'stats': {
            'total_users': toplam_siparis,
            'revenue': toplam_ciro,
            'orders': kontrol_edilen_siparis,
            'avg_response': toplam_adet
        },
        'tire_sales_data': json.dumps(tire_sales_data),
        'brand_chart_data': json.dumps(brand_chart_data),
        'son_islemler': son_islemler,
        'son_finance_islemleri': son_finance_islemleri,
        'monthly_income': json.dumps(monthly_income),
        'monthly_expense': json.dumps(monthly_expense),
        'monthly_labels': json.dumps(monthly_labels),
        'top_customers_data': json.dumps({
            'labels': customer_labels,
            'data': customer_data,
            'details': customer_details
        }),
        'payment_data': json.dumps(payment_data)
    }
    return render(request, 'dashboard/index.html', context)


def _month_starts_last_12():
    today = timezone.now().date().replace(day=1)
    months = []
    y = today.year
    m = today.month
    for i in range(11, -1, -1):
        mm = m - i
        yy = y
        while mm <= 0:
            mm += 12
            yy -= 1
        months.append(date(yy, mm, 1))
    return months


def _tx_total_expr():
    return (
        F('nakit')
        + F('kredi_karti')
        + F('sanal_pos')
        + F('cari')
        + F('mehmet_havale')
        + F('banka_havale')
        + F('pafgo')
        + F('canta_cikis')
    )


def _expense_category_name(tx):
    if tx.kategori1:
        if tx.kategori1.parent:
            return tx.kategori1.parent.name
        return tx.kategori1.name
    return "Kategorisiz"


def _parse_period_from_text(text):
    """
    Very small Turkish period parser:
    - bugun
    - son N gun (7/30)
    - bu ay
    - YYYY-MM
    """
    t = (text or "").lower()
    today = timezone.now().date()

    if "bugün" in t or "bugun" in t:
        start = today
        end = today + timedelta(days=1)
        label = "Bugün"
        return start, end, label

    import re
    m = re.search(r"son\s+(\d+)\s*g(ü|u)n", t)
    if m:
        n = int(m.group(1))
        start = today - timedelta(days=n)
        end = today + timedelta(days=1)
        return start, end, f"Son {n} gün"

    if "bu ay" in t:
        start = today.replace(day=1)
        end = (start + timedelta(days=32)).replace(day=1)
        return start, end, "Bu ay"

    m2 = re.search(r"(\d{4})-(\d{2})", t)
    if m2:
        yy = int(m2.group(1))
        mm = int(m2.group(2))
        start = date(yy, mm, 1)
        end = (start + timedelta(days=32)).replace(day=1)
        return start, end, f"{yy}-{mm:02d}"

    # default: last 30 days
    start = today - timedelta(days=30)
    end = today + timedelta(days=1)
    return start, end, "Son 30 gün"


def _sum_tx(qs):
    total_expr = _tx_total_expr()
    return qs.aggregate(total=Sum(total_expr))["total"] or Decimal("0")


def _top_expense_categories(qs, limit=5):
    totals = defaultdict(Decimal)
    for tx in qs.select_related("kategori1", "kategori1__parent"):
        cat = _expense_category_name(tx)
        totals[cat] += (
            (tx.nakit or 0)
            + (tx.kredi_karti or 0)
            + (tx.sanal_pos or 0)
            + (tx.cari or 0)
            + (tx.mehmet_havale or 0)
            + (tx.banka_havale or 0)
            + (tx.pafgo or 0)
            + (tx.canta_cikis or 0)
        )
    return sorted(totals.items(), key=lambda x: x[1], reverse=True)[:limit]


def _format_try(amount):
    try:
        amt = Decimal(amount)
    except Exception:
        amt = Decimal("0")
    return f"{amt:,.0f}".replace(",", ".")


def _parse_period_for_used_tires(text):
    """
    Çıkma lastikler için tarih alanı soruya göre değişebilir:
    - 'sat'/'satıldı' geçiyorsa satış tarihi
    - aksi halde çıkış tarihi
    """
    t = (text or "").lower()
    start, end, label = _parse_period_from_text(text)
    date_field = "satis_tarihi" if ("sat" in t or "satıldı" in t or "satildi" in t) else "cikis_tarihi"
    return start, end, label, date_field


def _format_count(n):
    try:
        return f"{int(n):,}".replace(",", ".")
    except Exception:
        return "0"


def _format_money(amount):
    return f"{_format_try(amount)} ₺"


def _used_tires_summary_for_user(*, user, q):
    start, end, label, date_field = _parse_period_for_used_tires(q)
    qs = CikmaLastik.objects.filter(user=user).filter(**{f"{date_field}__gte": start, f"{date_field}__lt": end})

    # totals
    total_rows = qs.count()
    total_adet = qs.aggregate(total=Sum("adet"))["total"] or 0

    by_status = (
        qs.values("durum")
        .annotate(rows=Count("id"), adet=Sum("adet"))
        .order_by("durum")
    )
    status_map = {d["durum"]: d for d in by_status}

    sold_qs = qs.filter(durum="satildi")
    sold_rows = sold_qs.count()
    sold_adet = sold_qs.aggregate(total=Sum("adet"))["total"] or 0
    sold_revenue = Decimal("0")
    for r in sold_qs.only("adet", "satis_fiyati"):
        if r.satis_fiyati:
            sold_revenue += (r.satis_fiyati * Decimal(r.adet or 0))

    lines = [
        f"{label} çıkma lastik özeti ({'satış tarihi' if date_field=='satis_tarihi' else 'çıkış tarihi'}):",
        f"- Kayıt: {_format_count(total_rows)} | Adet: {_format_count(total_adet)}",
        f"- Satılan: {_format_count(sold_rows)} kayıt / {_format_count(sold_adet)} adet | Ciro: {_format_money(sold_revenue)}",
        f"- Depoda (cikti+depolandi): {_format_count((status_map.get('cikti',{}).get('adet') or 0) + (status_map.get('depolandi',{}).get('adet') or 0))} adet",
    ]
    return "\n".join(lines)


def _orders_summary_for_user(*, user, q):
    start, end, label = _parse_period_from_text(q)
    qs = Siparis.objects.filter(user=user, olusturma_tarihi__gte=start, olusturma_tarihi__lt=end)

    total_rows = qs.count()
    total_ciro = qs.aggregate(total=Sum("toplam_fiyat"))["total"] or 0
    by_status = qs.values("durum").annotate(rows=Count("id")).order_by("-rows")

    top_status = list(by_status[:5])
    status_lines = []
    for row in top_status:
        status_lines.append(f"  - {row['durum']}: {_format_count(row['rows'])}")

    text = (
        f"{label} sipariş özeti (oluşturma tarihine göre):\n"
        f"- Sipariş: {_format_count(total_rows)}\n"
        f"- Toplam ciro: {_format_money(total_ciro)}\n"
        f"- Durum dağılımı (ilk 5):\n" + ("\n".join(status_lines) if status_lines else "  - Kayıt yok")
    )
    return text


@login_required
@misafir_forbidden
def manager_assistant(request):
    return render(request, "dashboard/assistant.html", {})


@login_required
@misafir_forbidden
@require_POST
def manager_assistant_query(request):
    """
    Yönetici Asistanı — Tool-calling mimarisi ile gerçek veritabanı sorguları.
    Groq, hangi tool'u çağıracağını belirler → Django ORM üzerinden veri çekilir
    → Groq sonucu yorumlar → kullanıcıya döner.
    """
    import time
    from .ai_tools import TOOL_DEFINITIONS, TOOL_REGISTRY, USER_REQUIRED_TOOLS

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}

    q = (payload.get("q") or "").strip()
    if not q:
        return JsonResponse({"error": "Soru boş olamaz."}, status=400)

    # Chat history (session tabanlı, son 10 tur)
    history = request.session.get("ai_chat_history", [])
    if not isinstance(history, list):
        history = []

    user = request.user
    today = timezone.now().date()

    # API key
    api_key = getattr(settings, "GROQ_API_KEY", None) or os.environ.get("GROQ_API_KEY", "")
    if not api_key:
        try:
            from dotenv import load_dotenv
            load_dotenv(getattr(settings, "BASE_DIR", None) / ".env")
        except Exception:
            pass
        api_key = os.environ.get("GROQ_API_KEY", "")

    _bad_keys = {"buraya-yeni-key-yapistir", "your-api-key-here", "sk-xxx", ""}
    if not api_key or api_key.strip() in _bad_keys or api_key.strip().lower().startswith("buraya"):
        api_key = ""

    model = os.environ.get("GROQ_MODEL", "groq/compound-mini")

    # ── SYSTEM PROMPT ──────────────────────────────────────────────────────────
    system_prompt = f"""Sen MEStakip'in Yönetici Asistanısın. Bugünün tarihi: {today} (Europe/Istanbul).

ROLİN:
- Siparişler, finans, stok, çıkma lastikler, teklifler ve malzeme verilerini analiz edersin
- Kullanıcının sorularını anlayıp doğru tool'ları çağırırsın
- Tool'dan gelen gerçek veriye dayanarak cevap verirsin
- Veritabanında olmayan bilgiyi KESİNLİKLE uydurmaz, tahmin etmezsin
- Bir bilgi mevcut değilse bunu açıkça söylersin

CEVAP KURALLARI:
- Türkçe yaz, kısa ve net ol
- Sayıları Türk formatında yaz: 1.234,56 ₺
- Birden fazla veri varsa tablo veya liste kullan (Markdown)
- Gerçek veri ile öneriyi birbirinden ayır: öneri yapıyorsan "💡 Öneri:" önekini kullan
- Tool sonucu boş gelirse "Bu dönemde kayıt bulunamadı" de
- Konuşma geçmişine göre bağlamı koru ("peki geçen ay?" gibi follow-up sorularını anla)

ZAMAN ANLAMA:
- "Bu ay" = {today.replace(day=1)} – {today}
- "Geçen ay" = bir önceki ayın tamamı
- "Son 30 gün" = {today - timedelta(days=30)} – {today}
- "Son 3 ay" = {(today - timedelta(days=90)).replace(day=1)} – {today}
- "Bu yıl" = {today.replace(month=1, day=1)} – {today}

Tool çağırmadan kesin sayı söyleme."""

    # ── MESAJ YAPISI ───────────────────────────────────────────────────────────
    messages = [{"role": "system", "content": system_prompt}]

    # Son 8 tur geçmişi ekle (token tasarrufu için)
    for turn in history[-8:]:
        messages.append(turn)

    messages.append({"role": "user", "content": q})

    used_groq = False
    text = ""

    if api_key:
        try:
            t0 = time.time()

            # ── TUR 1: Tool selection ──────────────────────────────────────────
            response = groq_chat_completion(
                api_key=api_key,
                model=model,
                messages=messages,
                tools=TOOL_DEFINITIONS,
                tool_choice="auto",
                temperature=0.1,
                max_tokens=1000,
            )

            choice = response.get("choices", [{}])[0]
            msg = choice.get("message", {})
            finish_reason = choice.get("finish_reason", "")

            # ── Tool çağrısı var mı? ───────────────────────────────────────────
            if finish_reason == "tool_calls" and msg.get("tool_calls"):
                tool_calls = msg["tool_calls"]
                messages.append(msg)  # asistan mesajını geçmişe ekle

                tool_results = []
                for tc in tool_calls:
                    tool_name = tc["function"]["name"]
                    try:
                        tool_args = json.loads(tc["function"].get("arguments", "{}") or "{}")
                    except Exception:
                        tool_args = {}

                    fn = TOOL_REGISTRY.get(tool_name)
                    if fn is None:
                        result = {"hata": f"Bilinmeyen tool: {tool_name}"}
                    elif tool_name in USER_REQUIRED_TOOLS:
                        result = fn(user=user, **tool_args)
                    else:
                        result = fn(**tool_args)

                    tool_results.append({
                        "tool_call_id": tc["id"],
                        "tool_name": tool_name,
                        "result": result,
                    })
                    messages.append({
                        "role": "tool",
                        "tool_call_id": tc["id"],
                        "content": json.dumps(result, ensure_ascii=False, default=str),
                    })

                # ── TUR 2: Yorumlama ──────────────────────────────────────────
                final_response = groq_chat_completion(
                    api_key=api_key,
                    model=model,
                    messages=messages,
                    temperature=0.2,
                    max_tokens=1000,
                )
                text = (final_response or "").strip()

            else:
                # Tool çağrısı yok, doğrudan metin yanıtı
                text = (msg.get("content") or "").strip()

            if text:
                used_groq = True

                # Geçmişi güncelle (sadece user + assistant turları)
                history.append({"role": "user", "content": q})
                history.append({"role": "assistant", "content": text})
                # Son 20 mesajı tut (10 tur)
                request.session["ai_chat_history"] = history[-20:]
                request.session.modified = True

        except Exception as _e:
            text = ""

    # ── FALLBACK: Groq yoksa basit kural tabanlı yanıt ────────────────────────
    if not text:
        from .ai_tools import get_proactive_summary, get_orders_summary, get_financial_summary
        t_lower = q.lower()

        if any(w in t_lower for w in ["özet", "genel", "durum", "nasıl"]):
            data = get_proactive_summary(user=user)
            lines = data.get("bilgiler", []) + data.get("pozitifler", []) + data.get("uyarilar", [])
            finans = data.get("finans_bu_ay", {})
            text = (
                f"📊 Genel Durum ({today}):\n\n"
                + "\n".join(lines) + "\n\n"
                f"💰 Bu ay: Gelir {finans.get('gelir_tl', 0):,.0f} ₺ | "
                f"Gider {finans.get('gider_tl', 0):,.0f} ₺ | "
                f"Net {finans.get('net_tl', 0):+,.0f} ₺"
            )
        elif any(w in t_lower for w in ["sipariş", "siparis", "ciro", "yolda", "marka"]):
            data = get_orders_summary(user=user)
            text = (
                f"📦 Sipariş Özeti ({data.get('donem', '')}):\n"
                f"• Toplam: {data.get('toplam_siparis', 0)} sipariş\n"
                f"• Ciro: {data.get('toplam_ciro_tl', 0):,.0f} ₺"
            )
        elif any(w in t_lower for w in ["gelir", "gider", "finans", "net", "para", "kasa"]):
            data = get_financial_summary(user=user)
            text = (
                f"💰 Finans ({data.get('donem', '')}):\n"
                f"• Gelir: {data.get('gelir_tl', 0):,.0f} ₺\n"
                f"• Gider: {data.get('gider_tl', 0):,.0f} ₺\n"
                f"• Net: {data.get('net_tl', 0):+,.0f} ₺"
            )
        else:
            text = (
                "Şu anda AI servisine bağlanamıyorum. "
                "Hızlı sorulardan birini deneyebilirsin."
            )

    return JsonResponse({"text": text, "used_groq": used_groq})


@login_required
@misafir_forbidden
def clear_chat_history(request):
    """Sohbet geçmişini temizler."""
    if request.method == "POST":
        request.session.pop("ai_chat_history", None)
        request.session.modified = True
        return JsonResponse({"ok": True})
    return JsonResponse({"error": "POST required"}, status=405)


@login_required
@misafir_forbidden
def expenses_3d(request):
    months = _month_starts_last_12()
    start = months[0]
    end = (months[-1] + timedelta(days=32)).replace(day=1)

    # Grafikleri besleyen veri (income-expense-report ile aynı filtreleri kullan)
    qs = get_filtered_transactions(
        user=request.user,
        hareket_tipi='gider'
    ).filter(
        tarih__gte=start,
        tarih__lt=end
    ).order_by('tarih')

    month_labels = [d.strftime("%Y-%m") for d in months]
    month_index = {d.strftime("%Y-%m"): i for i, d in enumerate(months)}

    totals_by_cat = defaultdict(Decimal)
    totals_by_month_cat = defaultdict(lambda: defaultdict(Decimal))
    total_12m = Decimal("0")

    for tx in qs:
        label = tx.tarih.replace(day=1).strftime("%Y-%m")
        if label not in month_index:
            continue
        cat = _expense_category_name(tx)
        total = (tx.nakit or 0) + (tx.kredi_karti or 0) + (tx.sanal_pos or 0) + (tx.cari or 0) + (tx.mehmet_havale or 0) + (tx.banka_havale or 0) + (tx.pafgo or 0) + (tx.canta_cikis or 0)
        total_12m += total
        totals_by_cat[cat] += total
        totals_by_month_cat[label][cat] += total

    top_n = 10
    top_categories_sorted = sorted(totals_by_cat.items(), key=lambda x: x[1], reverse=True)[:top_n]
    categories = [name for name, _ in top_categories_sorted]

    z = []
    for cat in categories:
        row = []
        for ml in month_labels:
            row.append(float(totals_by_month_cat.get(ml, {}).get(cat, 0)))
        z.append(row)

    context = {
        "months_json": json.dumps(month_labels),
        "categories_json": json.dumps(categories),
        "z_json": json.dumps(z),
        "top_n": top_n,
        "total_12m": total_12m,
        "top_categories": [{"name": n, "total": t} for n, t in top_categories_sorted[:5]],
    }
    return render(request, "dashboard/expenses_3d.html", context)


@login_required
@misafir_forbidden
@require_POST
def expenses_3d_insights(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}

    months = _month_starts_last_12()
    start = months[0]
    end = (months[-1] + timedelta(days=32)).replace(day=1)
    month_labels_all = [d.strftime("%Y-%m") for d in months]

    # ── Gider verisi (income-expense-report ile aynı filtreleri kullan) ──
    # get_filtered_transactions fonksiyonu bazı gelir ve virmanları hariç tutuyor,
    # burada sadece 'gider' olanları alıyoruz.
    expense_qs = get_filtered_transactions(
        user=request.user,
        hareket_tipi="gider"
    ).filter(
        tarih__gte=start,
        tarih__lt=end
    )

    by_month = defaultdict(Decimal)
    by_cat = defaultdict(Decimal)

    for tx in expense_qs:
        ml = tx.tarih.replace(day=1).strftime("%Y-%m")
        cat = _expense_category_name(tx)
        amt = (
            (tx.nakit or 0) + (tx.kredi_karti or 0) + (tx.sanal_pos or 0)
            + (tx.cari or 0) + (tx.mehmet_havale or 0) + (tx.banka_havale or 0)
            + (tx.pafgo or 0) + (tx.canta_cikis or 0)
        )
        by_month[ml] += amt
        by_cat[cat] += amt

    # Aylık gider serisi (tüm 12 ay, boş aylar 0)
    monthly_expense = [
        {"ay": ml, "gider_tl": float(by_month.get(ml, Decimal("0")))}
        for ml in month_labels_all
    ]
    toplam_gider_12m = sum(by_month.values()) or Decimal("0")

    # Kategori dağılımı (yüzde dahil, en fazla 8)
    top_cats_raw = sorted(by_cat.items(), key=lambda x: x[1], reverse=True)[:8]
    toplam_cat = sum(v for _, v in top_cats_raw) or Decimal("1")
    kategori_dagilimi = [
        {
            "kategori": cat,
            "toplam_tl": float(amt),
            "yuzde": round(float(amt / toplam_cat * 100), 1),
        }
        for cat, amt in top_cats_raw
    ]

    # ── API key ──
    api_key = getattr(settings, "GROQ_API_KEY", None) or os.environ.get("GROQ_API_KEY", "")
    if not api_key:
        try:
            from dotenv import load_dotenv
            load_dotenv(getattr(settings, "BASE_DIR", None) / ".env")
        except Exception:
            pass
        api_key = os.environ.get("GROQ_API_KEY", "")
    _bad = {"buraya-yeni-key-yapistir", "your-api-key-here", "sk-xxx", ""}
    if not api_key or api_key.strip() in _bad or api_key.strip().lower().startswith("buraya"):
        return JsonResponse({"error": "GROQ_API_KEY tanımlı değil veya geçersiz."}, status=400)

    # ── Groq prompt ──
    system_prompt = (
        "Sen MEStakip gider analiz asistanısın. Türkçe, kısa ve net yaz.\n"
        "Verilen sayılar gerçek veritabanı kayıtlarından geliyor — yalnızca bu rakamlara dayan, uydurma yapma.\n\n"
        "ÇIKTI FORMATI (bu sırayla, başka bir şey ekleme):\n"
        "1) Özet: Son 12 aylık toplam gideri ve aylık ortalamayı tek cümlede yaz.\n"
        "2) Trend: Aylık gider verilerindeki artış/düşüş/dalgalı trendi açıkla.\n"
        "3) En Yüksek Kategoriler: İlk 3 kategoriyi yüzdesiyle listele.\n"
        "4) Risk/Anomali: Anormal yüksek gider ayı veya kategori varsa belirt; yoksa 'Yok' yaz.\n"
        "5) Öneri: Veriye dayalı en fazla 2 somut madde.\n"
        "Toplam çıktı 900 karakteri geçmesin."
    )

    user_content = json.dumps(
        {
            "para_birimi": "TRY",
            "donem": "son_12_ay",
            "toplam_gider_tl": float(toplam_gider_12m),
            "aylik_ortalama_gider_tl": round(float(toplam_gider_12m) / 12, 2),
            "aylik_gider_serisi": monthly_expense,
            "gider_kategori_dagilimi": kategori_dagilimi,
        },
        ensure_ascii=False,
    )

    try:
        text = groq_chat_completion(
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_content},
            ],
            api_key=api_key,
            model=os.environ.get("GROQ_MODEL", "openai/gpt-oss-20b"),
            temperature=0.1,
            max_tokens=500,
        )
    except GroqError as e:
        return JsonResponse({"error": str(e)}, status=502)

    text = (text or "").strip()
    if len(text) > 1100:
        text = text[:1100].rstrip() + "…"
    return JsonResponse({"text": text})

@misafir_forbidden
def analytics(request):
    """Analytics sayfası"""
    context = {
        'page_title': 'Analytics',
    }
    return render(request, 'dashboard/analytics.html', context)

@login_required
@misafir_forbidden
def users(request):
    """Users sayfası - Rol bazlı yetkilendirme"""
    # Kullanıcının profilini al
    try:
        user_profile = UserProfile.objects.get(user=request.user)
    except UserProfile.DoesNotExist:
        # Eğer profil yoksa, admin kullanıcısı için admin rolü ver
        if request.user.username == 'admin':
            user_profile = UserProfile.objects.create(user=request.user, role='admin')
        else:
            user_profile = UserProfile.objects.create(user=request.user, role='yonetici')
    
    # Admin kontrolü (Django bayrakları ile uyumlu)
    is_admin = user_profile.is_admin() or request.user.is_superuser or request.user.is_staff
    
    # Kullanıcı listesi - sadece admin görebilir
    users_list = []
    if is_admin:
        users_list = User.objects.all().select_related('userprofile')
    
    # Kullanıcı oluşturma
    if request.method == 'POST' and is_admin:
        username = request.POST.get('username')
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        password = request.POST.get('password')
        role = request.POST.get('role', 'yonetici')
        
        if username and password:
            try:
                # Kullanıcı oluştur
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name
                )
                
                # Signal'den önce profil oluştur
                UserProfile.objects.filter(user=user).delete()  # Eğer signal oluşturduysa sil
                UserProfile.objects.create(user=user, role=role)
                
                messages.success(request, f'Kullanıcı {username} başarıyla oluşturuldu!')
                return redirect('dashboard:users')
            except Exception as e:
                messages.error(request, f'Kullanıcı oluşturulurken hata: {str(e)}')
        else:
            messages.error(request, 'Kullanıcı adı ve şifre gereklidir!')
    
    context = {
        'page_title': 'Users',
        'is_admin': is_admin,
        'users_list': users_list,
        'user_profile': user_profile,
    }
    return render(request, 'dashboard/users.html', context)

@login_required
@misafir_forbidden
def products(request):
    """Products sayfası: tarih aralığı filtresi ile işlemleri göster"""
    from datetime import date
    
    # Tarih aralığı parametrelerini al
    baslangic_tarih = request.GET.get('baslangic_tarih')
    bitis_tarih = request.GET.get('bitis_tarih')
    
    # Eski tek tarih parametresi (geriye uyumluluk için)
    secilen_tarih = request.GET.get('tarih')
    if not secilen_tarih and not baslangic_tarih and not bitis_tarih:
        secilen_tarih = date.today().strftime('%Y-%m-%d')
    
    # İşlemleri filtrele (kategorileri ve parent kategorileri de yükle)
    # Kredi kartı giderlerini products sayfasından tamamen hariç tut
    qs = Transaction.objects.filter(created_by=request.user).select_related(
        'kategori1', 'kategori1__parent', 'kategori2', 'kategori2__parent', 'kategori3', 'kategori3__parent'
    ).exclude(Q(hareket_tipi='gider') & Q(kredi_karti__gt=0))
    
    if baslangic_tarih:
        qs = qs.filter(tarih__gte=baslangic_tarih)
    if bitis_tarih:
        qs = qs.filter(tarih__lte=bitis_tarih)
    if secilen_tarih and not baslangic_tarih and not bitis_tarih:
        qs = qs.filter(tarih=secilen_tarih)
    
    qs = qs.order_by('-created_at')
    
    # Özet bilgileri (Sanal Pos ve Banka Havale hariç, Kredi Kartı dahil)
    # Merkez Satış ve Virman kasalarını hariç tut
    # Kredi kartı giderleri zaten sorguda filtrelendi
    toplam_ifade = (F('nakit') + F('kredi_karti') + F('cari') + F('mehmet_havale') + F('canta_cikis'))
    
    gun_ozeti = qs.exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(
        gelir=Sum(Case(When(hareket_tipi='gelir', then=toplam_ifade), default=0, output_field=DecimalField(max_digits=12, decimal_places=2))),
        gider=Sum(Case(When(hareket_tipi='gider', then=toplam_ifade), default=0, output_field=DecimalField(max_digits=12, decimal_places=2)))
    )
    
    gun_ozeti['gelir'] = gun_ozeti['gelir'] or 0
    gun_ozeti['gider'] = gun_ozeti['gider'] or 0
    gun_ozeti['net'] = gun_ozeti['gelir'] - gun_ozeti['gider']
    gun_ozeti['islem_sayisi'] = qs.exclude(kasa_adi__in=['merkez-satis', 'virman']).count()
    
    # Ödeme yöntemlerine göre toplamlar (gelir - gider)
    # Merkez Satış ve Virman kasalarını hariç tut
    # Gelir toplamları
    gelir_nakit = qs.filter(hareket_tipi='gelir').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('nakit', default=0))['total'] or 0
    gelir_kredi_karti = qs.filter(hareket_tipi='gelir').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('kredi_karti', default=0))['total'] or 0
    gelir_cari = qs.filter(hareket_tipi='gelir').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('cari', default=0))['total'] or 0
    gelir_sanal_pos = qs.filter(hareket_tipi='gelir').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('sanal_pos', default=0))['total'] or 0
    gelir_mehmet_havale = qs.filter(hareket_tipi='gelir').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('mehmet_havale', default=0))['total'] or 0
    gelir_banka_havale = qs.filter(hareket_tipi='gelir').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('banka_havale', default=0))['total'] or 0
    gelir_canta_cikis = qs.filter(hareket_tipi='gelir').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('canta_cikis', default=0))['total'] or 0
    
    # Gider toplamları
    gider_nakit = qs.filter(hareket_tipi='gider').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('nakit', default=0))['total'] or 0
    gider_kredi_karti = qs.filter(hareket_tipi='gider').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('kredi_karti', default=0))['total'] or 0
    gider_cari = qs.filter(hareket_tipi='gider').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('cari', default=0))['total'] or 0
    gider_sanal_pos = qs.filter(hareket_tipi='gider').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('sanal_pos', default=0))['total'] or 0
    gider_mehmet_havale = qs.filter(hareket_tipi='gider').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('mehmet_havale', default=0))['total'] or 0
    gider_banka_havale = 0  # Banka Havale gider gösterilmiyor
    gider_canta_cikis = qs.filter(hareket_tipi='gider').exclude(kasa_adi__in=['merkez-satis', 'virman']).aggregate(total=Sum('canta_cikis', default=0))['total'] or 0
    
    # Excel verileri için tarih filtrelemesi (Servis toplamları için)
    excel_hareketler = MalzemeHareketi.objects.filter(kullanici=request.user)
    
    # Tarih filtrelerini uygula
    if baslangic_tarih:
        excel_hareketler = excel_hareketler.filter(tarih__gte=baslangic_tarih)
    if bitis_tarih:
        excel_hareketler = excel_hareketler.filter(tarih__lte=bitis_tarih)
    if secilen_tarih and not baslangic_tarih and not bitis_tarih:
        excel_hareketler = excel_hareketler.filter(tarih=secilen_tarih)
    
    # KATEGORİ'de "hizmet" yazan Excel verilerini Servis toplamlarına ekle
    # Türkçe karakter sorunu için hem büyük hem küçük harf kontrol et
    excel_hizmet_hareketler = excel_hareketler.filter(
        Q(kategori__icontains='hizmet') | Q(kategori__icontains='HİZMET')
    )
    
    # Excel hizmet tutarlarını ödeme şekline göre dağıt
    # Nakit hesaplamasından vade içeren kayıtları (GÜN) hariç tut
    # NAKİT (Türkçe İ) - Excel'de "NAKİT" yazılırsa eşleşsin
    excel_nakit = excel_hizmet_hareketler.filter(
        Q(odeme_sekli__icontains='nakit') |
        Q(odeme_sekli__icontains='NAKİT') |
        Q(odeme_sekli__iexact='nakit') |
        Q(odeme_sekli__iexact='NAKİT')
    ).exclude(
        Q(odeme_sekli__icontains='gün') |
        Q(odeme_sekli__iregex=r'\d+\s*gün') |
        Q(odeme_sekli__iregex=r'\d+\s*GÜN')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    # Kredi kartı eşleştirmesi (kart, pos, kuveyttürk pos vb.) - SANAL POS hariç
    excel_kart = excel_hizmet_hareketler.filter(
        Q(odeme_sekli__icontains='kart') | 
        Q(odeme_sekli__icontains='pos') |
        Q(odeme_sekli__icontains='kuveyttürk') |
        Q(odeme_sekli__icontains='kuveyt') |
        Q(odeme_sekli__iexact='kredi kartı') |
        Q(odeme_sekli__iexact='kredi karti')
    ).exclude(
        Q(odeme_sekli__icontains='sanal pos') |
        Q(odeme_sekli__icontains='sanal') |
        Q(odeme_sekli__iexact='SANAL POS') |
        Q(odeme_sekli__icontains='sanalpos')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    # Cari eşleştirmesi (cari 1 gün, cari 5 gün vb.) ve GÜN ödeme şekli
    excel_cari = excel_hizmet_hareketler.filter(
        Q(odeme_sekli__icontains='cari') |
        Q(odeme_sekli__icontains='carı') |
        Q(odeme_sekli__iexact='cari') |
        Q(odeme_sekli__iregex=r'cari\s*\d+\s*gün') |
        Q(odeme_sekli__iregex=r'carı\s*\d+\s*gün') |
        Q(odeme_sekli__iexact='GÜN') |
        Q(odeme_sekli__icontains='gün') |
        Q(odeme_sekli__iregex=r'\d+\s*gün')  # "5 GÜN", "30 GÜN" gibi
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    # Sanal POS eşleştirmesi
    excel_sanal_pos = excel_hizmet_hareketler.filter(
        Q(odeme_sekli__icontains='sanal pos') |
        Q(odeme_sekli__icontains='sanal') |
        Q(odeme_sekli__iexact='sanal pos') |
        Q(odeme_sekli__iexact='SANAL POS') |
        Q(odeme_sekli__icontains='sanalpos')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    excel_havale = excel_hizmet_hareketler.filter(
        Q(odeme_sekli__icontains='havale') |
        Q(odeme_sekli__iexact='havale') |
        Q(odeme_sekli__icontains='garanti havale') |
        Q(odeme_sekli__icontains='GARANTİ') |
        Q(odeme_sekli__icontains='garantı') |
        Q(odeme_sekli__icontains='vakif') |
        Q(odeme_sekli__icontains='vakıf')
    ).exclude(
        Q(odeme_sekli__icontains='m.havale') |
        Q(odeme_sekli__icontains='m havale') |
        Q(odeme_sekli__icontains='mhavale') |
        Q(odeme_sekli__icontains='mehmet havale') |
        Q(odeme_sekli__iexact='M.HAVALE')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    # Diğer ödeme şekilleri (belirtilmemiş olanlar) nakit olarak kabul et
    # Ancak vade içeren kayıtları (GÜN) hariç tut
    excel_diger = excel_hizmet_hareketler.exclude(
        odeme_sekli__icontains='nakit'
    ).exclude(
        Q(odeme_sekli__icontains='kart') | 
        Q(odeme_sekli__icontains='pos') |
        Q(odeme_sekli__icontains='kuveyttürk')
    ).exclude(
        Q(odeme_sekli__icontains='cari') |
        Q(odeme_sekli__icontains='carı') |
        Q(odeme_sekli__iexact='cari') |
        Q(odeme_sekli__iregex=r'cari\s*\d+\s*gün') |
        Q(odeme_sekli__iregex=r'carı\s*\d+\s*gün') |
        Q(odeme_sekli__icontains='gün') |
        Q(odeme_sekli__iregex=r'\d+\s*gün') |
        Q(odeme_sekli__iregex=r'\d+\s*GÜN')
    ).exclude(
        odeme_sekli__icontains='sanal pos'
    ).exclude(
        odeme_sekli__icontains='havale'
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    excel_nakit_toplam = excel_nakit + excel_diger
    
    # Toplam Excel hizmet tutarı
    excel_servis_toplam = excel_nakit_toplam + excel_kart + excel_cari + excel_sanal_pos + excel_havale
    
    # Hizmet olmayan kategorilerdeki tutarları Merkez Satış'a ekle (LASTİK, AKÜ, JANT vb.)
    excel_merkez_hareketler = excel_hareketler.exclude(
        Q(kategori__icontains='hizmet') | Q(kategori__icontains='HİZMET')
    )
    
    # Merkez Satış nakit hesaplamasından vade içeren kayıtları (GÜN) hariç tut
    # NAKİT (Türkçe İ) için icontains ekle - Excel'de "NAKİT" yazılırsa eşleşsin
    excel_merkez_nakit = excel_merkez_hareketler.filter(
        Q(odeme_sekli__icontains='nakit') |
        Q(odeme_sekli__icontains='NAKİT') |
        Q(odeme_sekli__iexact='nakit') |
        Q(odeme_sekli__iexact='NAKİT')
    ).exclude(
        Q(odeme_sekli__icontains='gün') |
        Q(odeme_sekli__iregex=r'\d+\s*gün') |
        Q(odeme_sekli__iregex=r'\d+\s*GÜN')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    excel_merkez_kart = excel_merkez_hareketler.filter(
        Q(odeme_sekli__icontains='kart') | 
        Q(odeme_sekli__icontains='kuveyttürk') |
        Q(odeme_sekli__icontains='kuveyt')
    ).exclude(
        Q(odeme_sekli__icontains='sanal pos') |
        Q(odeme_sekli__icontains='sanal')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    excel_merkez_cari = excel_merkez_hareketler.filter(
        Q(odeme_sekli__icontains='cari') |
        Q(odeme_sekli__icontains='carı') |
        Q(odeme_sekli__iregex=r'cari\s*\d+\s*gün') |
        Q(odeme_sekli__iregex=r'carı\s*\d+\s*gün') |
        Q(odeme_sekli__iexact='GÜN') |
        Q(odeme_sekli__icontains='gün') |
        Q(odeme_sekli__iregex=r'\d+\s*gün')  # "5 GÜN", "30 GÜN" gibi
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    excel_merkez_sanal_pos = excel_merkez_hareketler.filter(
        Q(odeme_sekli__icontains='sanal pos') |
        Q(odeme_sekli__icontains='sanal') |
        Q(odeme_sekli__iexact='SANAL POS') |
        Q(odeme_sekli__icontains='sanalpos')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    # M.Havale: sadece m.havale, m havale, mhavale, mehmet havale, M.HAVALE kalıpları (merkez)
    excel_merkez_mhavale = excel_merkez_hareketler.filter(
        Q(odeme_sekli__icontains='m.havale') |
        Q(odeme_sekli__icontains='m havale') |
        Q(odeme_sekli__icontains='mhavale') |
        Q(odeme_sekli__icontains='mehmet havale') |
        Q(odeme_sekli__iexact='M.HAVALE')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    # Garanti/Banka Havale: garanti havale, GARANTİ HAVALE (Türkçe İ), banka havale, toplam havale, b.havale, b havale, vakif havale kalıpları (merkez)
    excel_merkez_garanti_havale = excel_merkez_hareketler.filter(
        Q(odeme_sekli__icontains='garanti havale') |
        Q(odeme_sekli__icontains='GARANTİ') |
        Q(odeme_sekli__icontains='garantı') |
        Q(odeme_sekli__icontains='banka havale') |
        Q(odeme_sekli__icontains='toplam havale') |
        Q(odeme_sekli__icontains='b.havale') |
        Q(odeme_sekli__icontains='b havale') |
        Q(odeme_sekli__icontains='garantihavale') |
        Q(odeme_sekli__icontains='bankahavale') |
        Q(odeme_sekli__icontains='vakif') |
        Q(odeme_sekli__icontains='vakıf')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    # Toplam (debug dict için)
    excel_merkez_havale = excel_merkez_mhavale + excel_merkez_garanti_havale
    
    # Hizmet olmayan kategorilerdeki diğer ödeme şekilleri (belirtilmemiş olanlar)
    # Ancak vade içeren kayıtları (GÜN) hariç tut
    # NAKİT (Türkçe İ) hariç - nakit/diger çift sayımını önle
    excel_merkez_diger = excel_merkez_hareketler.exclude(
        Q(odeme_sekli__icontains='nakit') |
        Q(odeme_sekli__icontains='NAKİT') |
        Q(odeme_sekli__iexact='nakit') |
        Q(odeme_sekli__iexact='NAKİT')
    ).exclude(
        Q(odeme_sekli__icontains='kart') | 
        Q(odeme_sekli__icontains='pos') |
        Q(odeme_sekli__icontains='kuveyttürk')
    ).exclude(
        Q(odeme_sekli__icontains='cari') |
        Q(odeme_sekli__icontains='carı') |
        Q(odeme_sekli__iexact='cari') |
        Q(odeme_sekli__iregex=r'cari\s*\d+\s*gün') |
        Q(odeme_sekli__iregex=r'carı\s*\d+\s*gün') |
        Q(odeme_sekli__icontains='gün') |
        Q(odeme_sekli__iregex=r'\d+\s*gün') |
        Q(odeme_sekli__iregex=r'\d+\s*GÜN')
    ).exclude(
        odeme_sekli__icontains='sanal pos'
    ).exclude(
        odeme_sekli__icontains='havale'
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    excel_merkez_nakit_toplam = excel_merkez_nakit + excel_merkez_diger
    
    # Tüm kategorilerdeki M.HAVALE ödeme şekilli tutarları M.Havale'ye ekle
    excel_mhavale = excel_hareketler.filter(
        Q(odeme_sekli__icontains='m.havale') |
        Q(odeme_sekli__icontains='m havale') |
        Q(odeme_sekli__icontains='mhavale') |
        Q(odeme_sekli__icontains='mehmet havale') |
        Q(odeme_sekli__iexact='M.HAVALE')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    # Excel hizmet tutarlarını ödeme yöntemlerine göre dağıt (SADECE HİZMET KATEGORİSİ)
    # Net tutarları göster (gelir - gider)
    gun_ozeti['nakit_toplam'] = (gelir_nakit - gider_nakit) + excel_nakit_toplam  # Sadece hizmet nakit tutarları (net)
    gun_ozeti['kredi_karti_toplam'] = (gelir_kredi_karti - gider_kredi_karti) + excel_kart  # Sadece hizmet kart tutarları (net)
    gun_ozeti['cari_toplam'] = (gelir_cari - gider_cari) + excel_cari  # Sadece hizmet cari tutarları (net)
    gun_ozeti['sanal_pos_toplam'] = (gelir_sanal_pos - gider_sanal_pos) + excel_sanal_pos  # Transaction + hizmet sanal pos tutarları (net)
    gun_ozeti['mehmet_havale_toplam'] = (gelir_mehmet_havale - gider_mehmet_havale) + excel_mhavale  # Transaction + tüm kategorilerdeki M.HAVALE tutarları (net)
    gun_ozeti['banka_havale_toplam'] = (gelir_banka_havale - gider_banka_havale) + excel_havale  # Transaction + Excel hizmet Garanti/Banka Havale
    gun_ozeti['canta_cikis_toplam'] = gelir_canta_cikis - gider_canta_cikis  # Net

    # Servis Toplamları genel toplamı
    gun_ozeti['servis_genel_toplam'] = (
        gun_ozeti['nakit_toplam'] +
        gun_ozeti['kredi_karti_toplam'] +
        gun_ozeti['cari_toplam'] +
        gun_ozeti['sanal_pos_toplam'] +
        gun_ozeti['mehmet_havale_toplam'] +
        gun_ozeti['banka_havale_toplam']
    )

    # Merkez Satış toplamını hesapla ve Toplam Gelir'e ekle
    merkez_satis_toplam = (excel_merkez_nakit_toplam + excel_merkez_kart + 
                          excel_merkez_cari + excel_merkez_sanal_pos + excel_merkez_havale)
    gun_ozeti['merkez_satis_toplam'] = merkez_satis_toplam  # Debug için
    gun_ozeti['gelir'] = gun_ozeti['gelir'] + merkez_satis_toplam
    gun_ozeti['net'] = gun_ozeti['gelir'] - gun_ozeti['gider']
    
    # Servis, Merkez Satış ve Çıkma Lastik kasaları için Nakit, Kredi Kartı ve M.Havale toplamları
    servis_merkez_qs = qs.filter(kasa_adi__in=['servis', 'merkez-satis', 'cikma-lastik'])
    
    # Nakit toplamları (Excel Merkez nakit + diğer tutarlarını ekle)
    servis_merkez_nakit_gelir = servis_merkez_qs.filter(hareket_tipi='gelir').aggregate(total=Sum('nakit', default=0))['total'] or 0
    servis_merkez_nakit_gelir += excel_merkez_nakit_toplam  # excel_merkez_nakit + excel_merkez_diger (diger: belirtilmemiş, NAKİT hariç)
    servis_merkez_nakit_gider = servis_merkez_qs.filter(hareket_tipi='gider').aggregate(total=Sum('nakit', default=0))['total'] or 0
    servis_merkez_nakit_net = servis_merkez_nakit_gelir - servis_merkez_nakit_gider
    
    # Kredi Kartı toplamları (Excel Merkez kart tutarlarını da ekle)
    servis_merkez_kredi_gelir = servis_merkez_qs.filter(hareket_tipi='gelir').aggregate(total=Sum('kredi_karti', default=0))['total'] or 0
    servis_merkez_kredi_gelir += excel_merkez_kart  # Excel'den gelen merkez kart tutarlarını ekle
    servis_merkez_kredi_gider = servis_merkez_qs.filter(hareket_tipi='gider').aggregate(total=Sum('kredi_karti', default=0))['total'] or 0
    servis_merkez_kredi_net = servis_merkez_kredi_gelir - servis_merkez_kredi_gider
    
    # M.Havale toplamları (Excel'den gelen M.HAVALE tutarlarını ekle - tüm kategoriler)
    servis_merkez_mhavale_gelir = servis_merkez_qs.filter(hareket_tipi='gelir').aggregate(total=Sum('mehmet_havale', default=0))['total'] or 0
    servis_merkez_mhavale_gelir += excel_mhavale  # Excel'den gelen tüm kategorilerdeki M.HAVALE tutarlarını ekle
    servis_merkez_mhavale_gider = servis_merkez_qs.filter(hareket_tipi='gider').aggregate(total=Sum('mehmet_havale', default=0))['total'] or 0
    servis_merkez_mhavale_net = servis_merkez_mhavale_gelir - servis_merkez_mhavale_gider
    
    # Cari toplamları (Excel Merkez cari tutarlarını da ekle)
    servis_merkez_cari_gelir = servis_merkez_qs.filter(hareket_tipi='gelir').aggregate(total=Sum('cari', default=0))['total'] or 0
    servis_merkez_cari_gelir += excel_merkez_cari  # Excel'den gelen merkez cari tutarlarını ekle
    servis_merkez_cari_gider = servis_merkez_qs.filter(hareket_tipi='gider').aggregate(total=Sum('cari', default=0))['total'] or 0
    servis_merkez_cari_net = servis_merkez_cari_gelir - servis_merkez_cari_gider
    
    # Sanal Pos toplamları (Excel Merkez sanal pos tutarlarını da ekle)
    servis_merkez_sanal_pos_gelir = servis_merkez_qs.filter(hareket_tipi='gelir').aggregate(total=Sum('sanal_pos', default=0))['total'] or 0
    servis_merkez_sanal_pos_gelir += excel_merkez_sanal_pos  # Excel'den gelen merkez sanal pos tutarlarını ekle
    servis_merkez_sanal_pos_gider = servis_merkez_qs.filter(hareket_tipi='gider').aggregate(total=Sum('sanal_pos', default=0))['total'] or 0
    servis_merkez_sanal_pos_net = servis_merkez_sanal_pos_gelir - servis_merkez_sanal_pos_gider
    
    # Banka Havale toplamları (Excel Merkez banka havale tutarlarını da ekle)
    servis_merkez_banka_havale_gelir = servis_merkez_qs.filter(hareket_tipi='gelir').aggregate(total=Sum('banka_havale', default=0))['total'] or 0
    servis_merkez_banka_havale_gelir += excel_merkez_garanti_havale  # Excel'den gelen merkez garanti/banka havale tutarlarını ekle
    servis_merkez_banka_havale_gider = 0  # Banka Havale gider gösterilmiyor
    servis_merkez_banka_havale_net = servis_merkez_banka_havale_gelir - servis_merkez_banka_havale_gider
    
    # Çanta Çıkış toplamları
    servis_merkez_canta_cikis_gelir = servis_merkez_qs.filter(hareket_tipi='gelir').aggregate(total=Sum('canta_cikis', default=0))['total'] or 0
    servis_merkez_canta_cikis_gider = servis_merkez_qs.filter(hareket_tipi='gider').aggregate(total=Sum('canta_cikis', default=0))['total'] or 0
    servis_merkez_canta_cikis_net = servis_merkez_canta_cikis_gelir - servis_merkez_canta_cikis_gider
    
    # Excel verileri için tarih filtrelemesi (Servis toplamları için)
    excel_hareketler = MalzemeHareketi.objects.filter(kullanici=request.user)
    
    # Tarih filtrelerini uygula
    if baslangic_tarih:
        excel_hareketler = excel_hareketler.filter(tarih__gte=baslangic_tarih)
    if bitis_tarih:
        excel_hareketler = excel_hareketler.filter(tarih__lte=bitis_tarih)
    if secilen_tarih and not baslangic_tarih and not bitis_tarih:
        excel_hareketler = excel_hareketler.filter(tarih=secilen_tarih)
    
    # KATEGORİ'de "hizmet" yazan Excel verilerini Servis toplamlarına ekle
    excel_servis_toplam = excel_hareketler.filter(
        Q(kategori__icontains='hizmet') | Q(kategori__icontains='HİZMET')
    ).aggregate(total=Sum('tutar', default=0))['total'] or 0
    
    # Toplam (Nakit + Kredi Kartı + Cari + Sanal Pos + M.Havale + Banka Havale + Çanta Çıkış + Excel Hizmet)
    # NOT: merkez_satis_toplam eklenmez - Excel merkez tutarları zaten servis_merkez_*_gelir değişkenlerine eklendi (çift sayım önlendi)
    servis_merkez_toplam_gelir = servis_merkez_nakit_gelir + servis_merkez_kredi_gelir + servis_merkez_cari_gelir + servis_merkez_sanal_pos_gelir + servis_merkez_mhavale_gelir + servis_merkez_banka_havale_gelir + servis_merkez_canta_cikis_gelir + excel_servis_toplam
    servis_merkez_toplam_gider = servis_merkez_nakit_gider + servis_merkez_kredi_gider + servis_merkez_cari_gider + servis_merkez_sanal_pos_gider + servis_merkez_mhavale_gider + servis_merkez_canta_cikis_gider  # Banka Havale gider dahil değil
    servis_merkez_toplam_net = servis_merkez_toplam_gelir - servis_merkez_toplam_gider
    
    gun_ozeti['servis_merkez_toplam'] = {
        'gelir': servis_merkez_toplam_gelir,
        'gider': servis_merkez_toplam_gider,
        'net': servis_merkez_toplam_net,
        'excel_hizmet': excel_servis_toplam,  # Debug için
        'excel_debug': {  # Debug bilgileri
            'toplam_excel_kayit': excel_hareketler.count(),
            'hizmet_kayit': excel_hizmet_hareketler.count(),
            'excel_nakit': excel_nakit,
            'excel_kart': excel_kart,
            'excel_cari': excel_cari,
            'excel_havale': excel_havale,
            'excel_merkez_nakit': excel_merkez_nakit,
            'excel_merkez_kart': excel_merkez_kart,
            'excel_merkez_cari': excel_merkez_cari,
            'excel_merkez_havale': excel_merkez_havale,
            'excel_merkez_mhavale': excel_merkez_mhavale,
            'excel_merkez_garanti_havale': excel_merkez_garanti_havale,
        }
    }

    # Excel verileri için tarih filtrelemesi
    hareketler = MalzemeHareketi.objects.filter(kullanici=request.user).order_by('-tarih')
    dosyalar = MalzemeDosya.objects.filter(kullanici=request.user).prefetch_related('satirlar').order_by('-yukleme_tarihi')
    
    # Tarih aralığı filtresi uygula
    if baslangic_tarih:
        hareketler = hareketler.filter(tarih__gte=baslangic_tarih)
        dosyalar = dosyalar.filter(yukleme_tarihi__date__gte=baslangic_tarih)
    if bitis_tarih:
        hareketler = hareketler.filter(tarih__lte=bitis_tarih)
        dosyalar = dosyalar.filter(yukleme_tarihi__date__lte=bitis_tarih)
    
    # Eski parametreler (geriye uyumluluk)
    start_s = request.GET.get('start-date')
    end_s = request.GET.get('end-date')
    if start_s and not baslangic_tarih:
        hareketler = hareketler.filter(tarih__gte=start_s)
        dosyalar = dosyalar.filter(yukleme_tarihi__date__gte=start_s)
    if end_s and not bitis_tarih:
        hareketler = hareketler.filter(tarih__lte=end_s)
        dosyalar = dosyalar.filter(yukleme_tarihi__date__lte=end_s)
    from collections import defaultdict
    import datetime
    gunluk = defaultdict(list)
    for h in hareketler:
        tarih_str = h.tarih.strftime('%d.%m.%Y')
        gunluk[tarih_str].append(h)
    gunluk_excel = defaultdict(list)
    for d in dosyalar:
        # Dosyayı hem yükleme tarihine hem de içindeki kayıtların tarihlerine göre gruplandır
        yukleme_tarihi_str = d.yukleme_tarihi.strftime('%d.%m.%Y')
        gunluk_excel[yukleme_tarihi_str].append(d)
        
        # Ayrıca dosya içindeki kayıtların tarihlerine göre de gruplandır
        for satir in d.satirlar.all():
            satir_tarih_str = satir.tarih.strftime('%d.%m.%Y')
            if satir_tarih_str != yukleme_tarihi_str:
                # Eğer kayıt tarihi farklıysa, o tarihe de ekle
                if d not in gunluk_excel[satir_tarih_str]:
                    gunluk_excel[satir_tarih_str].append(d)
    gunluk_sorted = dict(sorted(gunluk.items(), key=lambda x: datetime.datetime.strptime(x[0], '%d.%m.%Y'), reverse=True))
    gunluk_excel_sorted = dict(sorted(gunluk_excel.items(), key=lambda x: datetime.datetime.strptime(x[0], '%d.%m.%Y'), reverse=True))
    
    # Excel verilerini belirle
    secilen_gun_excel = []
    secilen_tarih_str = ''
    
    if secilen_tarih and not baslangic_tarih and not bitis_tarih:
        # Tek tarih seçilmişse
        from datetime import datetime as dt
        secilen_tarih_obj = dt.strptime(secilen_tarih, '%Y-%m-%d')
        secilen_tarih_str = secilen_tarih_obj.strftime('%d.%m.%Y')
        secilen_gun_excel = gunluk_excel_sorted.get(secilen_tarih_str, [])
    else:
        # Tarih aralığı seçilmişse, tüm filtrelenmiş dosyaları göster
        secilen_gun_excel = list(dosyalar)
    
    # Excel verilerine göre ödeme şekillerine göre toplamlar
    excel_odeme_toplamlari = {}
    
    # Tüm filtrelenmiş Excel satırlarını al (HİZMET hariç - LASTİK, AKÜ, JANT vb.)
    excel_satirlar = MalzemeHareketi.objects.filter(kullanici=request.user).exclude(
        Q(kategori__icontains='hizmet') | Q(kategori='HİZMET')
    )
    if baslangic_tarih:
        excel_satirlar = excel_satirlar.filter(tarih__gte=baslangic_tarih)
    if bitis_tarih:
        excel_satirlar = excel_satirlar.filter(tarih__lte=bitis_tarih)
    if secilen_tarih and not baslangic_tarih and not bitis_tarih:
        excel_satirlar = excel_satirlar.filter(tarih=secilen_tarih)
    
    # Ödeme şekillerine göre grupla ve topla
    excel_odeme_dict = {
        'Nakit': Decimal('0'),
        'Kredi_Karti': Decimal('0'),
        'Cari': Decimal('0'),
        'Sanal_Pos': Decimal('0'),
        'Mehmet_Havale': Decimal('0'),
        'Banka_Havale': Decimal('0')
    }
    
    for satir in excel_satirlar:
        odeme_sekli = satir.odeme_sekli or 'Belirtilmemiş'
        # Ödeme şeklini normalize et (Türkçe karakterleri dönüştür, küçük harfe çevir, boşlukları temizle)
        odeme_sekli_normalized = odeme_sekli.replace('İ', 'i').replace('I', 'ı').lower().strip().replace(' ', '').replace('.', '').replace('-', '')
        odeme_sekli_original = odeme_sekli.strip()
        
        # Ana ödeme şekillerini belirle (tam eşleşme ve içerme kontrolü)
        # Nakit kontrolü
        amount = parse_decimal_value(satir.tutar)
        if (odeme_sekli_original.replace('İ', 'i').replace('I', 'ı').lower() == 'nakit' or 
            'nakit' in odeme_sekli_normalized or 
            odeme_sekli_normalized == 'nakit'):
            excel_odeme_dict['Nakit'] += amount
        # Kredi Kartı kontrolü
        elif ('kart' in odeme_sekli_normalized or 
              'kredi' in odeme_sekli_normalized or 
              'kredit' in odeme_sekli_normalized or
              'kuveyttürk' in odeme_sekli_normalized or
              'kuveyt' in odeme_sekli_normalized or
              odeme_sekli_original.lower() in ['kredi kartı', 'kredi karti', 'kart', 'credit card']) and \
              'sanal' not in odeme_sekli_normalized:  # SANAL POS hariç
            excel_odeme_dict['Kredi_Karti'] += amount
        # Sanal Pos kontrolü
        elif ('sanal' in odeme_sekli_normalized or 
              'sanalpos' in odeme_sekli_normalized or
              'sanal pos' in odeme_sekli_original.lower() or
              odeme_sekli_original.upper() == 'SANAL POS'):
            excel_odeme_dict['Sanal_Pos'] += amount
        # Banka Havale kontrolü (Mehmet Havale'den önce kontrol edilmeli)
        # GARANTİ HAVALE / VAKIF HAVALE: Türkçe İ karakteri için normalize + orijinal kontrol
        elif ('banka' in odeme_sekli_normalized or 
              'bankahavale' in odeme_sekli_normalized or
              'bhavale' in odeme_sekli_normalized or
              'garanti' in odeme_sekli_normalized or
              'garantihavale' in odeme_sekli_normalized or
              'GARANTİ' in (odeme_sekli_original or '') or
              'garantı' in odeme_sekli_normalized or
              'vakif' in odeme_sekli_normalized or
              'vakıf' in odeme_sekli_normalized or
              'VAKIF' in (odeme_sekli_original or '') or
              'VAKIF HAVALE' in (odeme_sekli_original or '').upper() or
              odeme_sekli_original.replace('İ', 'i').replace('I', 'ı').lower() in ['b.havale', 'b havale', 'bhavale', 'banka havale', 'banka havalesi', 'garanti havale', 'toplam havale', 'vakif havale', 'vakıf havale']):
            excel_odeme_dict['Banka_Havale'] += amount
        # Mehmet Havale kontrolü
        elif ('mehmet' in odeme_sekli_normalized or 
              'mhavale' in odeme_sekli_normalized or
              odeme_sekli_original.lower() in ['m.havale', 'm havale', 'mhavale', 'mehmet havale'] or
              ('havale' in odeme_sekli_normalized and 'banka' not in odeme_sekli_normalized)):
            excel_odeme_dict['Mehmet_Havale'] += amount
        # Cari kontrolü (GÜN ödeme şekli de dahil)
        elif ('cari' in odeme_sekli_normalized or 
              'carı' in odeme_sekli_normalized or
              odeme_sekli_original.lower() == 'cari' or
              odeme_sekli_normalized == 'cari' or
              'gün' in odeme_sekli_normalized or
              odeme_sekli_original.upper() == 'GÜN' or
              # Regex ile "5 GÜN", "30 GÜN" gibi formatları yakala
              any(char.isdigit() for char in odeme_sekli_original) and 'gün' in odeme_sekli_original.lower()):  
            excel_odeme_dict['Cari'] += amount
        else:
            # Eğer eşleşme yoksa, varsayılan olarak hiçbir şeye eklenmez
            pass
    
    # Merkez Satış kasasındaki işlemleri de Excel toplamlarına ekle (Gelir - Gider)
    merkez_satis_gelir = qs.filter(Q(kasa_adi='merkez-satis') & Q(hareket_tipi='gelir'))
    merkez_satis_gider = qs.filter(Q(kasa_adi='merkez-satis') & Q(hareket_tipi='gider'))
    
    # Merkez Satış Gelir işlemlerini ödeme şekillerine göre topla
    for islem in merkez_satis_gelir:
        excel_odeme_dict['Nakit'] += parse_decimal_value(islem.nakit)
        excel_odeme_dict['Kredi_Karti'] += parse_decimal_value(islem.kredi_karti)
        excel_odeme_dict['Cari'] += parse_decimal_value(islem.cari)
        excel_odeme_dict['Sanal_Pos'] += parse_decimal_value(islem.sanal_pos)
        excel_odeme_dict['Mehmet_Havale'] += parse_decimal_value(islem.mehmet_havale)
        excel_odeme_dict['Banka_Havale'] += parse_decimal_value(islem.banka_havale)
    
    # Merkez Satış Gider işlemlerini ödeme şekillerine göre düş (Banka Havale hariç)
    for islem in merkez_satis_gider:
        excel_odeme_dict['Nakit'] -= parse_decimal_value(islem.nakit)
        excel_odeme_dict['Kredi_Karti'] -= parse_decimal_value(islem.kredi_karti)
        excel_odeme_dict['Cari'] -= parse_decimal_value(islem.cari)
        excel_odeme_dict['Sanal_Pos'] -= parse_decimal_value(islem.sanal_pos)
        excel_odeme_dict['Mehmet_Havale'] -= parse_decimal_value(islem.mehmet_havale)
        # Banka Havale gider düşülmüyor
    
    # Dictionary'yi context'e gönder - Merkez Satış Excel değerleri zaten döngüde hesaplandı
    # excel_odeme_dict zaten doğru değerleri içeriyor, tekrar eklemeye gerek yok
    
    excel_odeme_toplamlari = excel_odeme_dict
    
    # Merkez Ekstra İşlemler - Merkez Satış ve Virman kasalarını dahil et
    # Merkez Satış kasasında Gider işlemlerinde Kredi Kartı ve Banka Havale hariç
    # Pafgo işlemleri products sayfasında gösterilmeyecek
    merkez_ekstra_islemler = qs.filter(kasa_adi__in=['merkez-satis', 'virman']).exclude(
        Q(kasa_adi='merkez-satis') & Q(hareket_tipi='gider') & (Q(kredi_karti__gt=0) | Q(banka_havale__gt=0))
    ).exclude(pafgo__gt=0).select_related('kategori1', 'kategori1__parent', 'kategori2', 'kategori2__parent', 'kategori3', 'kategori3__parent')
    # Detaylı işlemlerden Merkez Satış ve Virman kasalarını hariç tut
    # Gider ise: Kredi Kartı, Sanal Pos ve Banka Havale hariç tut
    # Gelir ise: Sadece Sanal Pos ve Banka Havale hariç tut (Kredi Kartı göster)
    # Pafgo işlemleri sadece Gelir/Gider raporunda gösterilecek, products sayfasında gösterilmeyecek
    detayli_islemler = (
        qs.exclude(kasa_adi__in=['merkez-satis', 'virman'])
          .exclude(Q(hareket_tipi='gider') & Q(kredi_karti__gt=0))
          .exclude(sanal_pos__gt=0)
          .exclude(banka_havale__gt=0)
          .exclude(pafgo__gt=0)
    )
    # Merkez Ekstra İşlemler toplamını hesapla
    # Merkez Satış Gider işlemlerinde Kredi Kartı ve Banka Havale hariç
    merkez_ekstra_toplam = Decimal('0')
    for islem in merkez_ekstra_islemler:
        if islem.kasa_adi == 'merkez-satis' and islem.hareket_tipi == 'gider':
            # Merkez Satış Gider: Sadece Nakit, Cari ve Mehmet Havale
            merkez_ekstra_toplam += (
                parse_decimal_value(islem.nakit) + 
                parse_decimal_value(islem.cari) + 
                parse_decimal_value(islem.mehmet_havale)
            )
        elif islem.kasa_adi == 'merkez-satis' and islem.hareket_tipi == 'gelir':
            # Merkez Satış Gelir: Tüm ödeme yöntemleri (Sanal Pos ve Pafgo hariç, Banka Havale dahil)
            merkez_ekstra_toplam += (
                parse_decimal_value(islem.nakit) + 
                parse_decimal_value(islem.kredi_karti) + 
                parse_decimal_value(islem.cari) + 
                parse_decimal_value(islem.mehmet_havale) + 
                parse_decimal_value(islem.banka_havale)
            )
        else:
            # Diğer işlemler (Virman): Tüm ödeme yöntemleri (Sanal Pos ve Banka Havale hariç)
            merkez_ekstra_toplam += (
                parse_decimal_value(islem.nakit) + 
                parse_decimal_value(islem.kredi_karti) + 
                parse_decimal_value(islem.cari) + 
                parse_decimal_value(islem.mehmet_havale)
            )

    context = {
        'page_title': 'Products',
        'secilen_tarih': secilen_tarih or date.today().strftime('%Y-%m-%d'),
        'baslangic_tarih': baslangic_tarih,
        'bitis_tarih': bitis_tarih,
        'gun_ozeti': gun_ozeti,
        'gunun_islemleri': detayli_islemler,
        'merkez_ekstra_islemler': merkez_ekstra_islemler,
        'merkez_ekstra_toplam': merkez_ekstra_toplam,
        'malzeme_gunluk': gunluk_sorted,
        'gunluk_excel': gunluk_excel_sorted,
        'secilen_gun_excel': secilen_gun_excel,
        'start_selected': start_s or '',
        'end_selected': end_s or '',
        # Debug bilgileri
        'debug_secilen_tarih_str': secilen_tarih_str,
        'debug_gunluk_excel_keys': list(gunluk_excel_sorted.keys()),
        'debug_toplam_dosya': len(dosyalar),
        'excel_odeme_toplamlari': excel_odeme_toplamlari,
        'debug_merkez_satis_toplam': merkez_satis_toplam,
        'debug_excel_merkez_nakit': excel_merkez_nakit_toplam,
        'debug_excel_merkez_kart': excel_merkez_kart,
    }
    return render(request, 'dashboard/products.html', context)




@login_required
@require_POST
def transaction_sil(request, transaction_id):
    """Detaylı işlemler tablosundan tek bir Transaction kaydını sil."""
    transaction = get_object_or_404(Transaction, id=transaction_id, created_by=request.user)
    next_url = request.POST.get('next') or request.META.get('HTTP_REFERER') or reverse('dashboard:products')

    try:
        aciklama = transaction.aciklama or f'#{transaction.id}'
        transaction.delete()
        messages.success(request, f'"{aciklama}" işlemi silindi.')
    except Exception as exc:
        messages.error(request, f'İşlem silinirken bir hata oluştu: {exc}')

    return redirect(next_url)


@login_required
@misafir_forbidden
def orders(request):
    """Sipariş Envanteri Dashboard (Orders)"""
    # Hızlı arama (cari ile arama)
    query = request.GET.get('q', '').strip()

    siparisler = Siparis.objects.filter(user=request.user)
    if query:
        siparisler = siparisler.filter(cari_firma__icontains=query)

    # Metrikler
    toplam_adet = siparisler.aggregate(total=Sum('adet'))['total'] or 0
    stoktaki_adet = siparisler.filter(ambar='stok').aggregate(total=Sum('adet'))['total'] or 0
    satistaki_adet = siparisler.filter(ambar='satis').aggregate(total=Sum('adet'))['total'] or 0
    toplam_ciro = siparisler.aggregate(total=Sum('toplam_fiyat'))['total'] or 0
    yolda_siparisler = siparisler.filter(durum='yolda').count()
    islemde_siparisler = siparisler.filter(durum='islemde').count()

    # Son eklenen işlemler (son 10 kayıt)
    son_islemler = siparisler.order_by('-olusturma_tarihi')[:10]

    # Marka dağılımı (marka bazında adet toplamı ve kayıt sayısı)
    marka_dagilimi = (
        siparisler.values('marka')
        .annotate(
            kayit_sayisi=Count('id'),
            adet_toplam=Sum('adet'),
            tutar_toplam=Sum('toplam_fiyat'),
        )
        .order_by('-adet_toplam')
    )

    context = {
        'page_title': 'Sipariş Envanteri Dashboard',
        'q': query,
        'stats': {
            'toplam_adet': toplam_adet,
            'stoktaki_adet': stoktaki_adet,
            'satistaki_adet': satistaki_adet,
            'toplam_ciro': toplam_ciro,
            'yolda': yolda_siparisler,
            'islemde': islemde_siparisler,
        },
        'son_islemler': son_islemler,
        'marka_dagilimi': marka_dagilimi,
    }
    return render(request, 'dashboard/orders.html', context)

@login_required
@misafir_forbidden
def forms(request):
    """Kontrol Edilen Siparişler Sayfası"""
    # Filtreleme parametreleri
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    urun_arama = request.GET.get('urun_arama', '')
    grup = request.GET.get('grup', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Sadece kontrol edilen siparişleri getir (sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(durum='kontrol', user=request.user)
    
    # Filtreleme uygula
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        # Türkçe karakter varyantları oluştur ve hepsiyle ara
        search_variants = create_turkish_search_variants(marka)
        q_objects = Q()
        for variant in search_variants:
            q_objects |= Q(marka__icontains=variant)
        siparisler = siparisler.filter(q_objects)
    if urun_arama:
        # Ebat formatını otomatik düzenle (2055516 -> 205/55R16)
        formatted_urun_arama = format_tire_size(urun_arama)
        siparisler = siparisler.filter(
            Q(urun__icontains=urun_arama) | Q(marka__icontains=urun_arama) |
            Q(urun__icontains=formatted_urun_arama) | Q(marka__icontains=formatted_urun_arama)
        )
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    
    # Eğer hiçbir tarih filtresi yoksa, varsayılan olarak son 3 ayın verilerini getir
    if not tarih_filtre and not baslangic_tarihi and not bitis_tarihi:
        start_date = now - timedelta(days=90)
        siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        tarih_filtre = 'son-3-ay'  # Varsayılan filtreyi işaretle
    
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # İstatistikler hesapla
    toplam_kontrol = siparisler.count()
    toplam_tutar = siparisler.aggregate(total=Sum('toplam_fiyat'))['total'] or 0
    toplam_adet = siparisler.aggregate(total=Sum('adet'))['total'] or 0
    
    # Grup bazında kontrol istatistikleri (lastik adet toplamı)
    grup_istatistikleri = siparisler.values('grup').annotate(
        total_adet=Sum('adet'),
        total_amount=Sum('toplam_fiyat')
    ).order_by('-total_adet')
    
    
    # Sayfalama
    paginator = Paginator(siparisler, 50)  # Sayfa başına 50 kayıt
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': 'Kontrol Edilen Siparişler',
        'siparisler': page_obj,
        'filters': {
            'firma': firma,
            'marka': marka,
            'urun_arama': urun_arama,
            'grup': grup,
            'mevsim': mevsim,
            'ambar': ambar,
            'tarih': tarih_filtre,
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
        },
        'stats': {
            'toplam_kontrol': toplam_kontrol,
            'toplam_tutar': toplam_tutar,
            'toplam_adet': toplam_adet,
        },
        'grup_istatistikleri': grup_istatistikleri,
    }
    return render(request, 'dashboard/forms.html', context)

@login_required
@misafir_forbidden
def elements(request):
    """Sipariş Envanteri Listesi sayfası"""
    # Filtreleme parametreleri
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    urun_arama = request.GET.get('urun_arama', '')
    grup = request.GET.get('grup', '')
    durum = request.GET.get('durum', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Siparişleri getir (iptal edilenleri ve kontrol edilenleri hariç tut, sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(user=request.user).exclude(durum__in=['iptal', 'kontrol'])
    
    # Filtreleme uygula
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        # Türkçe karakter varyantları oluştur ve hepsiyle ara
        search_variants = create_turkish_search_variants(marka)
        q_objects = Q()
        for variant in search_variants:
            q_objects |= Q(marka__icontains=variant)
        siparisler = siparisler.filter(q_objects)
    if urun_arama:
        # Ebat formatını otomatik düzenle (2055516 -> 205/55R16)
        formatted_urun_arama = format_tire_size(urun_arama)
        siparisler = siparisler.filter(
            Q(urun__icontains=urun_arama) | Q(marka__icontains=urun_arama) |
            Q(urun__icontains=formatted_urun_arama) | Q(marka__icontains=formatted_urun_arama)
        )
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if durum:
        siparisler = siparisler.filter(durum=durum)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            # Tarih aralığını günün başı ve sonu olarak ayarla
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            # Geçersiz tarih formatı
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass

    # Duruma göre özel sıralama:
    # 1) Teslim Edildi (teslim) - her zaman en üstte
    # 2) Takılacak/Faturası İşlendi (takilacak-faturasi-islendi)
    # 3) Yolda/Fatura İşlendi (yolda-fatura-islendi)
    # 4) İşlem devam ediyor/Faturası İşlendi (islemde-faturasi-islendi)
    # 5) İşleme Devam Ediyor (islemde)
    # 6) Yolda (yolda) - en aşağıda
    # 7) Diğer durumlar
    durum_sira = Case(
        When(durum='teslim', then=0),
        When(durum='takilacak-faturasi-islendi', then=1),
        When(durum='yolda-fatura-islendi', then=2),
        When(durum='islemde-faturasi-islendi', then=3),
        When(durum='islemde', then=4),
        When(durum='yolda', then=6),
        When(durum='takildi-ft-islendi-islem-devam', then=5),
        default=7,
        output_field=IntegerField(),
    )
    siparisler = siparisler.annotate(durum_sira=durum_sira).order_by('durum_sira', '-olusturma_tarihi')
    
    # Sayfalama
    paginator = Paginator(siparisler, 50)  # Sayfa başına 50 kayıt
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': 'Sipariş Envanteri',
        'siparisler': page_obj,
        'filters': {
            'firma': firma,
            'marka': marka,
            'urun_arama': urun_arama,
            'grup': grup,
            'durum': durum,
            'mevsim': mevsim,
            'ambar': ambar,
            'tarih': tarih_filtre,
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
        }
    }
    return render(request, 'dashboard/elements.html', context)

def elements_buttons(request):
    """Elements buttons sayfası"""
    context = {
        'page_title': 'Elements - Buttons',
    }
    return render(request, 'dashboard/elements-buttons.html', context)

def elements_alerts(request):
    """Elements alerts sayfası"""
    context = {
        'page_title': 'Elements - Alerts',
    }
    return render(request, 'dashboard/elements-alerts.html', context)

def elements_badges(request):
    """Elements badges sayfası"""
    context = {
        'page_title': 'Elements - Badges',
    }
    return render(request, 'dashboard/elements-badges.html', context)

def elements_cards(request):
    """Elements cards sayfası"""
    context = {
        'page_title': 'Elements - Cards',
    }
    return render(request, 'dashboard/elements-cards.html', context)

def elements_modals(request):
    """Elements modals sayfası"""
    context = {
        'page_title': 'Elements - Modals',
    }
    return render(request, 'dashboard/elements-modals.html', context)

def elements_forms(request):
    """Elements forms sayfası"""
    context = {
        'page_title': 'Elements - Forms',
    }
    return render(request, 'dashboard/elements-forms.html', context)

def elements_tables(request):
    """Elements tables sayfası"""
    context = {
        'page_title': 'Elements - Tables',
    }
    return render(request, 'dashboard/elements-tables.html', context)

@login_required
def yeni_lastik(request):
    """Yeni lastik ekleme sayfası"""
    if request.method == 'POST':
        form = SiparisForm(request.POST)
        if form.is_valid():
            siparis = form.save(commit=False)
            # Kullanıcıyı otomatik ata
            siparis.user = request.user
            siparis.save()

            # --- Otomatik öğrenme: ürün adından model çıkar ve mevsimi kaydet ---
            if siparis.mevsim and siparis.urun:
                try:
                    urun_text = siparis.urun.strip()
                    # Ürün adından ebat bilgisini çıkar (örn: "205/55R16" gibi kısımları at)
                    import re
                    # Ebat pattern: sayı/sayıRsayı veya sayıxsayıRsayı
                    model_adi = re.sub(r'\b\d{2,3}[/x]\d{2,3}[Rr]\d{2}\b', '', urun_text).strip()
                    model_adi = re.sub(r'\s+', ' ', model_adi).strip()
                    if model_adi and len(model_adi) >= 3:
                        LastikModelBilgisi.objects.update_or_create(
                            model_adi=model_adi,
                            defaults={
                                'mevsim': siparis.mevsim,
                                'son_guncelleyen': request.user,
                            }
                        )
                        # Güncelleme sayısını artır
                        LastikModelBilgisi.objects.filter(model_adi=model_adi).update(
                            guncelleme_sayisi=F('guncelleme_sayisi') + 1
                        )
                except Exception:
                    pass  # Öğrenme hatası siparişi etkilemesin

            messages.success(request, f'Sipariş başarıyla kaydedildi! ID: {siparis.id}')
            return redirect('dashboard:elements')
        else:
            # Form hatalarını detaylı göster
            error_messages = []
            for field, errors in form.errors.items():
                for error in errors:
                    error_messages.append(f'{field}: {error}')
            messages.error(request, f'Form hataları: {", ".join(error_messages)}')
            print(f"Form errors: {form.errors}")  # Debug için
    else:
        form = SiparisForm()
    
    context = {
        'page_title': 'Yeni Lastik Ekle',
        'form': form,
    }
    return render(request, 'dashboard/yeni_lastik.html', context)

@login_required
def siparis_detay(request, siparis_id):
    """Sipariş detay sayfası"""
    siparis = get_object_or_404(Siparis, id=siparis_id, user=request.user)
    context = {
        'page_title': f'Sipariş Detayı - #{siparis.id}',
        'siparis': siparis,
    }
    return render(request, 'dashboard/siparis_detay.html', context)

@login_required
def siparis_duzenle(request, siparis_id):
    """Sipariş düzenleme sayfası"""
    siparis = get_object_or_404(Siparis, id=siparis_id, user=request.user)
    
    if request.method == 'POST':
        form = SiparisForm(request.POST, instance=siparis)
        if form.is_valid():
            form.save()
            messages.success(request, f'Sipariş #{siparis.id} başarıyla güncellendi!')
            return redirect('dashboard:elements')
        else:
            messages.error(request, 'Form hataları var. Lütfen kontrol edin.')
    else:
        form = SiparisForm(instance=siparis)
    
    context = {
        'page_title': f'Sipariş Düzenle - #{siparis.id}',
        'form': form,
        'siparis': siparis,
    }
    return render(request, 'dashboard/siparis_duzenle.html', context)

@login_required
def siparis_sil(request, siparis_id):
    """Sipariş silme"""
    siparis = get_object_or_404(Siparis, id=siparis_id, user=request.user)
    
    if request.method == 'POST':
        siparis_id = siparis.id
        siparis.delete()
        messages.success(request, f'Sipariş #{siparis_id} başarıyla silindi!')
        return redirect('dashboard:elements')
    
    context = {
        'page_title': f'Sipariş Sil - #{siparis.id}',
        'siparis': siparis,
    }
    return render(request, 'dashboard/siparis_sil.html', context)

@login_required
def siparis_whatsapp(request, siparis_id):
    """WhatsApp mesajı gönder"""
    siparis = get_object_or_404(Siparis, id=siparis_id, user=request.user)
    
    # Cari firma adını al, yoksa varsayılan olarak "MesTakip" kullan
    firma_adi = siparis.cari_firma if siparis.cari_firma and siparis.cari_firma.strip() else 'MesTakip'
    
    # WhatsApp mesajı oluştur (Güncellenen Son Tarih kaldırıldı)
    mesaj = f"""*MesTakip - {firma_adi}*

*Ürün:* {siparis.urun}
*Marka:* {siparis.marka}
*Adet:* {siparis.adet}
*Durum:* {siparis.get_durum_display()}"""
    
    # SMS durumunu "gönderildi" olarak güncelle
    siparis.sms_durum = 'gonderildi'
    siparis.save()
    
    # WhatsApp URL'si oluştur (telefon numarası placeholder)
    encoded_mesaj = mesaj.replace(' ', '%20').replace('\n', '%0A')
    whatsapp_url = f"https://wa.me/?text={encoded_mesaj}"
    
    return redirect(whatsapp_url)

@login_required
def siparis_kontrol_edildi(request, siparis_id):
    """Siparişi kontrol edildi olarak işaretle"""
    if request.method == 'POST':
        try:
            siparis = get_object_or_404(Siparis, id=siparis_id, user=request.user)
            
            # Siparişin durumunu kontrol olarak güncelle
            siparis.durum = 'kontrol'
            siparis.save()
            
            return JsonResponse({
                'success': True,
                'message': 'Sipariş başarıyla kontrol edildi olarak işaretlendi.'
            })
            
        except Siparis.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Sipariş bulunamadı.'
            }, status=404)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'error': 'Sadece POST istekleri kabul edilir.'
    }, status=405)

@login_required
def siparis_teslim_edildi(request, siparis_id):
    """Siparişi teslim edildi olarak işaretle"""
    if request.method == 'POST':
        try:
            siparis = get_object_or_404(Siparis, id=siparis_id, user=request.user)
            siparis.durum = 'teslim'
            siparis.save()
            return JsonResponse({
                'success': True,
                'message': 'Sipariş başarıyla teslim edildi olarak işaretlendi.'
            })
        except Siparis.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Sipariş bulunamadı.'}, status=404)
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
    return JsonResponse({'success': False, 'error': 'Sadece POST istekleri kabul edilir.'}, status=405)

@login_required
def reports(request):
    """İptal Edilen Siparişler Raporu"""
    # Filtreleme parametreleri
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    grup = request.GET.get('grup', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Sadece iptal edilen siparişleri getir (sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(durum='iptal', user=request.user)
    
    # Filtreleme uygula
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        # Türkçe karakter varyantları oluştur ve hepsiyle ara
        search_variants = create_turkish_search_variants(marka)
        q_objects = Q()
        for variant in search_variants:
            q_objects |= Q(marka__icontains=variant)
        siparisler = siparisler.filter(q_objects)
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # İstatistikler hesapla
    toplam_iptal = siparisler.count()
    toplam_tutar = siparisler.aggregate(total=Sum('toplam_fiyat'))['total'] or 0
    toplam_adet = siparisler.aggregate(total=Sum('adet'))['total'] or 0
    
    # Grup bazında iptal istatistikleri (lastik adet toplamı)
    grup_istatistikleri = siparisler.values('grup').annotate(
        total_adet=Sum('adet'),
        total_amount=Sum('toplam_fiyat')
    ).order_by('-total_adet')
    
    # Sayfalama
    paginator = Paginator(siparisler, 50)  # Sayfa başına 50 kayıt
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_title': 'İptal Edilen Siparişler Raporu',
        'siparisler': page_obj,
        'filters': {
            'firma': firma,
            'marka': marka,
            'grup': grup,
            'mevsim': mevsim,
            'ambar': ambar,
            'tarih': tarih_filtre,
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
        },
        'stats': {
            'toplam_iptal': toplam_iptal,
            'toplam_tutar': toplam_tutar,
            'toplam_adet': toplam_adet,
        },
        'grup_istatistikleri': grup_istatistikleri,
    }
    return render(request, 'dashboard/reports.html', context)

def messages_view(request):
    """Gelir/Gider Toplamları sayfası (kasa bazında ve genel toplam)."""
    qs = Transaction.objects.filter(created_by=request.user)

    start_date_str = request.GET.get('baslangic_tarih') or request.GET.get('start_date') or ''
    end_date_str = request.GET.get('bitis_tarih') or request.GET.get('end_date') or ''
    start_date = end_date = None

    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            qs = qs.filter(tarih__gte=start_date)
        except ValueError:
            start_date_str = ''

    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            qs = qs.filter(tarih__lte=end_date)
        except ValueError:
            end_date_str = ''

    if start_date and end_date and start_date > end_date:
        start_date, end_date = end_date, start_date
        start_date_str, end_date_str = end_date_str, start_date_str

    def build_entries(queryset, amount_getter):
        entries = []
        # Kategorileri de yükle
        queryset = queryset.select_related('kategori1', 'kategori1__parent')
        for tx in queryset.order_by('-tarih', '-id'):
            raw_amount = amount_getter(tx) or 0
            try:
                amount = float(raw_amount)
            except (TypeError, ValueError):
                amount = 0
            if amount == 0:
                continue
            if tx.hareket_tipi == 'gider':
                amount = -abs(amount)
            else:
                amount = abs(amount)

            # Ana kategori ve alt kategori bilgilerini belirle
            ana_kategori = ''
            alt_kategori = ''
            
            if tx.kategori1:
                if tx.kategori1.parent:
                    # kategori1'in parent'ı varsa, parent ana kategori, kategori1 alt kategori
                    ana_kategori = tx.kategori1.parent.name
                    alt_kategori = tx.kategori1.name
                else:
                    # kategori1'in parent'ı yoksa, kategori1 ana kategori
                    ana_kategori = tx.kategori1.name

            entries.append({
                'id': tx.id,
                'tarih': tx.tarih.strftime('%d.%m.%Y'),
                'tarih_sort': tx.tarih,  # Sıralama için datetime objesi
                'kasa_adi': tx.get_kasa_adi_display() if tx.kasa_adi else '-',
                'hareket': tx.get_hareket_tipi_display(),
                'ana_kategori': ana_kategori,
                'alt_kategori': alt_kategori,
                'aciklama': tx.aciklama or '-',
                'amount': amount,
            })
        return entries

    # Kasa bazında hesaplama için mehmet_havale'yi dahil etmiyoruz (ayrı hesaplanacak)
    kasa_toplam_ifade = (F('nakit') + F('kredi_karti') + F('sanal_pos') + F('cari'))

    # Kasa bazında gelir, gider ve net (gelir - gider) - mehmet_havale hariç
    kasa_ozet = (
        qs.values('kasa_adi')
        .annotate(
            gelir=Sum(Case(When(hareket_tipi='gelir', then=kasa_toplam_ifade), default=0, output_field=DecimalField(max_digits=14, decimal_places=2))),
            gider=Sum(Case(When(hareket_tipi='gider', then=kasa_toplam_ifade), default=0, output_field=DecimalField(max_digits=14, decimal_places=2)))
        )
        .annotate(net=F('gelir') - F('gider'))
        .order_by('kasa_adi')
    )

    # Kartlar için belirli kasalar
    canta_toplam = 0
    mehmet_havale_toplam = 0
    rows = []
    
    for row in kasa_ozet:
        kasa = row['kasa_adi']
        net = float(row['net'] or 0)
        # Virman ve Çıkma Lastik kasalarını listeden hariç tut
        if kasa not in ('virman', 'cikma-lastik'):
            rows.append({'kasa_adi': kasa, 'bakiye': net})
        if kasa == 'canta':
            canta_toplam = net
        elif kasa == 'mehmet-havale':
            mehmet_havale_toplam = net

    # Mehmet havale field'ından toplam hesapla (her zaman field bazında)
    mehmet_havale_field_toplam = qs.aggregate(
        mehmet_havale_gelir=Sum(Case(When(hareket_tipi='gelir', then='mehmet_havale'), default=0, output_field=DecimalField(max_digits=14, decimal_places=2))),
        mehmet_havale_gider=Sum(Case(When(hareket_tipi='gider', then='mehmet_havale'), default=0, output_field=DecimalField(max_digits=14, decimal_places=2)))
    )
    mehmet_havale_toplam = float((mehmet_havale_field_toplam['mehmet_havale_gelir'] or 0) - (mehmet_havale_field_toplam['mehmet_havale_gider'] or 0))
    
    # Excel'den sadece M.HAVALE ödeme şekli olanları ekle
    from dashboard.models import MalzemeHareketi
    
    # Excel kayıtlarını al
    excel_lastik_mhavale_qs = MalzemeHareketi.objects.filter(kullanici=request.user)
    
    # Tarih filtrelerini uygula
    if start_date:
        excel_lastik_mhavale_qs = excel_lastik_mhavale_qs.filter(tarih__gte=start_date)
    if end_date:
        excel_lastik_mhavale_qs = excel_lastik_mhavale_qs.filter(tarih__lte=end_date)
    
    # Sadece M.HAVALE ödeme şekli olanları filtrele
    excel_lastik_mhavale_qs = excel_lastik_mhavale_qs.filter(
        Q(odeme_sekli__icontains='M.HAVALE') |
        Q(odeme_sekli__icontains='M HAVALE') |
        Q(odeme_sekli__icontains='MHAVALE')
    )
    
    # Toplamı hesapla ve mehmet_havale_toplam'a ekle
    excel_lastik_mhavale_sum = excel_lastik_mhavale_qs.aggregate(
        total=Sum('tutar', default=0)
    )['total'] or 0
    mehmet_havale_toplam += float(excel_lastik_mhavale_sum)
    
    # Banka havale field'ından toplam hesapla
    banka_havale_field_toplam = qs.aggregate(
        banka_havale_gelir=Sum(Case(When(hareket_tipi='gelir', then='banka_havale'), default=0, output_field=DecimalField(max_digits=14, decimal_places=2))),
        banka_havale_gider=Sum(Case(When(hareket_tipi='gider', then='banka_havale'), default=0, output_field=DecimalField(max_digits=14, decimal_places=2)))
    )
    banka_havale_toplam = float((banka_havale_field_toplam['banka_havale_gelir'] or 0) - (banka_havale_field_toplam['banka_havale_gider'] or 0))
    
    # Ödeme yöntemlerine göre toplamlar (Detaylı İşlemler'den)
    # Gelir toplamları
    gelir_nakit = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('nakit', default=0))['total'] or 0
    gelir_kredi_karti = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('kredi_karti', default=0))['total'] or 0
    gelir_cari = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('cari', default=0))['total'] or 0
    gelir_sanal_pos = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('sanal_pos', default=0))['total'] or 0
    gelir_mehmet_havale = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('mehmet_havale', default=0))['total'] or 0
    gelir_banka_havale = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('banka_havale', default=0))['total'] or 0
    
    # Gider toplamları
    gider_nakit = qs.filter(hareket_tipi='gider').aggregate(total=Sum('nakit', default=0))['total'] or 0
    gider_kredi_karti = qs.filter(hareket_tipi='gider').aggregate(total=Sum('kredi_karti', default=0))['total'] or 0
    gider_cari = qs.filter(hareket_tipi='gider').aggregate(total=Sum('cari', default=0))['total'] or 0
    gider_sanal_pos = qs.filter(hareket_tipi='gider').aggregate(total=Sum('sanal_pos', default=0))['total'] or 0
    gider_mehmet_havale = qs.filter(hareket_tipi='gider').aggregate(total=Sum('mehmet_havale', default=0))['total'] or 0
    gider_banka_havale = 0  # Banka Havale gider gösterilmiyor
    
    # Çanta Çıkış toplamları
    gelir_canta_cikis = qs.filter(hareket_tipi='gelir').aggregate(total=Sum('canta_cikis', default=0))['total'] or 0
    gider_canta_cikis = qs.filter(hareket_tipi='gider').aggregate(total=Sum('canta_cikis', default=0))['total'] or 0
    canta_cikis_net = float(gelir_canta_cikis - gider_canta_cikis)
    
    # Net toplamlar (gelir - gider)
    nakit_net = float(gelir_nakit - gider_nakit)
    kredi_karti_net = float(gelir_kredi_karti - gider_kredi_karti)
    cari_net = float(gelir_cari - gider_cari)
    sanal_pos_net = float(gelir_sanal_pos - gider_sanal_pos)
    mehmet_havale_field_net = float(gelir_mehmet_havale - gider_mehmet_havale)
    banka_havale_field_net = float(gelir_banka_havale - gider_banka_havale)
    
    # Excel'den LASTİK + NAKİT ve HİZMET + NAKİT kayıtlarını al ve çanta toplamına ekle
    excel_lastik_nakit_qs = MalzemeHareketi.objects.filter(kullanici=request.user)
    
    # Tarih filtrelerini uygula
    if start_date:
        excel_lastik_nakit_qs = excel_lastik_nakit_qs.filter(tarih__gte=start_date)
    if end_date:
        excel_lastik_nakit_qs = excel_lastik_nakit_qs.filter(tarih__lte=end_date)
    
    # LASTİK, HİZMET, JANT veya AKÜ kategorisi ve NAKİT ödeme şekli olanları filtrele
    excel_lastik_nakit_qs = excel_lastik_nakit_qs.filter(
        Q(kategori__icontains='LASTİK') | Q(kategori__icontains='LASTIK') |
        Q(kategori__icontains='HİZMET') | Q(kategori__icontains='HIZMET') |
        Q(kategori__icontains='JANT') |
        Q(kategori__icontains='AKÜ') | Q(kategori__icontains='AKU')
    ).filter(
        Q(odeme_sekli__icontains='NAKİT') | Q(odeme_sekli__icontains='NAKIT')
    )
    
    # Toplamı hesapla ve nakit_net'e ekle
    excel_lastik_nakit_sum = excel_lastik_nakit_qs.aggregate(
        total=Sum('tutar', default=0)
    )['total'] or 0
    nakit_net += float(excel_lastik_nakit_sum)
    
    # Nakit → Çanta'ya ekle
    canta_toplam += nakit_net
    
    # Çanta Çıkış → Çanta'ya ekle
    canta_toplam += canta_cikis_net
    
    # Genel Toplam: Çanta + Mehmet Havale
    genel_toplam = canta_toplam + mehmet_havale_toplam
    
    # Kart modal verileri
    canta_entries = build_entries(qs.filter(Q(kasa_adi='canta') | ~Q(nakit=0)), lambda tx: tx.nakit)
    
    # Çanta Çıkış işlemlerini canta_entries'e ekle
    canta_cikis_entries = build_entries(qs.filter(~Q(canta_cikis=0)), lambda tx: tx.canta_cikis)
    canta_entries.extend(canta_cikis_entries)
    
    # Excel'den LASTİK + NAKİT ve HİZMET + NAKİT kayıtlarını canta_entries'e ekle
    for hareket in excel_lastik_nakit_qs.order_by('-tarih', '-id'):
        canta_entries.append({
            'id': f'excel_{hareket.id}',
            'tarih': hareket.tarih.strftime('%d.%m.%Y'),
            'tarih_sort': hareket.tarih,  # Sıralama için datetime objesi
            'kasa_adi': 'Merkez Satış İş Akışı - (Excel)',
            'hareket': 'Gelir',
            'ana_kategori': hareket.kategori or '-',
            'alt_kategori': '-',
            'aciklama': f'{hareket.urun} - {hareket.musteri}',
            'amount': float(hareket.tutar),
        })
    
    # Çanta işlemlerini tarihe göre sırala (yeniden eskiye)
    # Önce tarih_sort alanını ekle (eğer yoksa tarih string'inden parse et)
    for entry in canta_entries:
        if 'tarih_sort' not in entry:
            try:
                # dd.mm.yyyy formatından datetime'a çevir
                entry['tarih_sort'] = datetime.strptime(entry['tarih'], '%d.%m.%Y').date()
            except:
                entry['tarih_sort'] = datetime.now().date()
    
    # Tarihe göre sırala (yeniden eskiye)
    canta_entries.sort(key=lambda x: x['tarih_sort'], reverse=True)
    
    mehmet_havale_entries = build_entries(qs.filter(~Q(mehmet_havale=0)), lambda tx: tx.mehmet_havale)
    
    # Excel'den LASTİK + M.HAVALE kayıtlarını mehmet_havale_entries'e ekle
    for hareket in excel_lastik_mhavale_qs.order_by('-tarih', '-id'):
        mehmet_havale_entries.append({
            'id': f'excel_{hareket.id}',
            'tarih': hareket.tarih.strftime('%d.%m.%Y'),
            'tarih_sort': hareket.tarih,  # Sıralama için datetime objesi
            'kasa_adi': 'Merkez Satış İş Akışı - (Excel)',
            'hareket': 'Gelir',
            'ana_kategori': hareket.kategori or '-',
            'alt_kategori': '-',
            'aciklama': f'{hareket.urun} - {hareket.musteri}',
            'amount': float(hareket.tutar),
        })
    
    # Mehmet Havale işlemlerini tarihe göre sırala (yeniden eskiye)
    for entry in mehmet_havale_entries:
        if 'tarih_sort' not in entry:
            try:
                entry['tarih_sort'] = datetime.strptime(entry['tarih'], '%d.%m.%Y').date()
            except:
                entry['tarih_sort'] = datetime.now().date()
    mehmet_havale_entries.sort(key=lambda x: x['tarih_sort'], reverse=True)
    
    banka_havale_entries = build_entries(qs.filter(~Q(banka_havale=0)), lambda tx: tx.banka_havale)
    # Genel entries: Nakit, Mehmet Havale ve Çanta Çıkış olan işlemler (Çanta + Mehmet Havale)
    genel_entries = build_entries(qs.filter(Q(nakit__gt=0) | Q(mehmet_havale__gt=0) | Q(canta_cikis__gt=0)), lambda tx: (tx.nakit or 0) + (tx.mehmet_havale or 0) + (tx.canta_cikis or 0))
    
    # Excel'den LASTİK + NAKİT ve HİZMET + NAKİT kayıtlarını genel_entries'e de ekle
    for hareket in excel_lastik_nakit_qs.order_by('-tarih', '-id'):
        genel_entries.append({
            'id': f'excel_{hareket.id}',
            'tarih': hareket.tarih.strftime('%d.%m.%Y'),
            'tarih_sort': hareket.tarih,  # Sıralama için datetime objesi
            'kasa_adi': 'Merkez Satış İş Akışı - (Excel)',
            'hareket': 'Gelir',
            'ana_kategori': hareket.kategori or '-',
            'alt_kategori': '-',
            'aciklama': f'{hareket.urun} - {hareket.musteri}',
            'amount': float(hareket.tutar),
        })
    
    # Excel'den LASTİK + M.HAVALE kayıtlarını genel_entries'e de ekle
    for hareket in excel_lastik_mhavale_qs.order_by('-tarih', '-id'):
        genel_entries.append({
            'id': f'excel_{hareket.id}',
            'tarih': hareket.tarih.strftime('%d.%m.%Y'),
            'tarih_sort': hareket.tarih,  # Sıralama için datetime objesi
            'kasa_adi': 'Merkez Satış İş Akışı - (Excel)',
            'hareket': 'Gelir',
            'ana_kategori': hareket.kategori or '-',
            'alt_kategori': '-',
            'aciklama': f'{hareket.urun} - {hareket.musteri}',
            'amount': float(hareket.tutar),
        })
    
    # Genel işlemleri tarihe göre sırala (yeniden eskiye)
    for entry in genel_entries:
        if 'tarih_sort' not in entry:
            try:
                entry['tarih_sort'] = datetime.strptime(entry['tarih'], '%d.%m.%Y').date()
            except:
                entry['tarih_sort'] = datetime.now().date()
    genel_entries.sort(key=lambda x: x['tarih_sort'], reverse=True)

    if start_date and end_date:
        date_range_label = f"{start_date.strftime('%d.%m.%Y')} - {end_date.strftime('%d.%m.%Y')}"
    elif start_date:
        date_range_label = f"{start_date.strftime('%d.%m.%Y')} sonrası"
    elif end_date:
        date_range_label = f"{end_date.strftime('%d.%m.%Y')} öncesi"
    else:
        date_range_label = "Tüm zamanlar"

    context = {
        'page_title': 'Gelir/Gider Toplamları',
        'canta_toplam': canta_toplam,
        'mehmet_havale_toplam': mehmet_havale_toplam,
        'banka_havale_toplam': banka_havale_toplam,
        'genel_toplam': genel_toplam,
        'kasa_satirlari': rows,
        'canta_islemleri': canta_entries,
        'mehmet_havale_islemleri': mehmet_havale_entries,
        'banka_havale_islemleri': banka_havale_entries,
        'genel_islemleri': genel_entries,
        # Ödeme yöntemleri toplamları (debug/ekstra bilgi için)
        'odeme_toplamlari': {
            'nakit': nakit_net,
            'kredi_karti': kredi_karti_net,
            'sanal_pos': sanal_pos_net,
            'cari': cari_net,
            'mehmet_havale': mehmet_havale_field_net,
            'banka_havale': banka_havale_field_net,
            'canta_cikis': canta_cikis_net,
        },
        'filters': {
            'start_date': start_date_str,
            'end_date': end_date_str,
        },
        'date_range_label': date_range_label,
    }
    return render(request, 'dashboard/messages.html', context)

@login_required
def finance_overview(request):
    """JSON: Son 12 ay için aylık gelir/gider toplamları (canlı)."""
    try:
        labels = []
        income = []
        expense = []
        toplam_ifade = (F('nakit') + F('kredi_karti') + F('sanal_pos') + F('cari') + F('mehmet_havale'))

        now = timezone.now()
        for i in range(11, -1, -1):
            start_date = (now - timedelta(days=30*i)).date().replace(day=1)
            # Bir sonraki ayın 1'i
            if start_date.month == 12:
                next_month = start_date.replace(year=start_date.year + 1, month=1, day=1)
            else:
                next_month = start_date.replace(month=start_date.month + 1, day=1)

            gelir = Transaction.objects.filter(
                created_by=request.user,
                tarih__gte=start_date,
                tarih__lt=next_month,
                hareket_tipi='gelir'
            ).aggregate(total=Sum(toplam_ifade))['total'] or 0

            gider = Transaction.objects.filter(
                created_by=request.user,
                tarih__gte=start_date,
                tarih__lt=next_month,
                hareket_tipi='gider'
            ).aggregate(total=Sum(toplam_ifade))['total'] or 0

            labels.append(start_date.strftime('%b'))
            income.append(float(gelir))
            expense.append(float(gider))

        return JsonResponse({
            'success': True,
            'labels': labels,
            'income': income,
            'expense': expense,
        })
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

def calendar(request):
    """Calendar sayfası"""
    context = {
        'page_title': 'Takvim',
        'events': [],
    }
    return render(request, 'dashboard/calendar.html', context)

@login_required
def quotations(request):
    """Teklifler sayfası"""
    from datetime import date
    from .models import Quotation
    
    # Tüm teklifleri getir
    quotations_list = Quotation.objects.filter(olusturan=request.user).select_related('olusturan')
    
    context = {
        'page_title': 'Teklifler',
        'today': date.today(),
        'quotations': quotations_list,
    }
    return render(request, 'dashboard/quotations.html', context)

@login_required
def quotation_view(request, quotation_id):
    """Teklif görüntüleme sayfası"""
    from .models import Quotation
    quotation = get_object_or_404(Quotation, id=quotation_id, olusturan=request.user)
    
    context = {
        'page_title': f'Teklif - {quotation.teklif_no}',
        'quotation': quotation,
    }
    return render(request, 'dashboard/quotation_view.html', context)

@login_required
def quotation_pdf(request, quotation_id):
    """Teklifi PDF olarak indir - HTML render"""
    from .models import Quotation
    from django.conf import settings
    import os
    
    quotation = get_object_or_404(Quotation, id=quotation_id, olusturan=request.user)
    
    # Logo yolunu bul
    logo_url = request.build_absolute_uri(settings.STATIC_URL + 'images/Meslas-Otomotiv.png')
    
    context = {
        'quotation': quotation,
        'logo_url': logo_url,
    }
    
    # HTML template'i render et
    return render(request, 'dashboard/quotation_pdf_print.html', context)
    """Teklifi PDF olarak indir"""
    from .models import Quotation
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    from django.conf import settings
    import os
    from io import BytesIO
    
    quotation = get_object_or_404(Quotation, id=quotation_id, olusturan=request.user)
    
    # PDF buffer oluştur
    buffer = BytesIO()
    
    # PDF dökümanı oluştur
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=1.5*cm,
        leftMargin=1.5*cm,
        topMargin=1.5*cm,
        bottomMargin=1.5*cm
    )
    
    # Story (içerik) listesi
    story = []
    
    # Stiller
    styles = getSampleStyleSheet()
    
    # Başlık stili
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#dc3545'),
        spaceAfter=5,
        alignment=TA_LEFT
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.grey,
        spaceAfter=20
    )
    
    section_style = ParagraphStyle(
        'SectionTitle',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor('#333333'),
        spaceAfter=10,
        spaceBefore=10,
        leftIndent=10,
        borderWidth=0,
        borderColor=colors.HexColor('#dc3545'),
        borderPadding=5,
        backColor=colors.HexColor('#f8f9fa')
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontSize=9,
        leading=12
    )
    
    # Header - Logo ve Başlık
    header_data = []
    
    # Logo yolunu bul
    logo_path = None
    if settings.STATIC_ROOT and os.path.exists(os.path.join(settings.STATIC_ROOT, 'images', 'Meslas-Otomotiv.png')):
        logo_path = os.path.join(settings.STATIC_ROOT, 'images', 'Meslas-Otomotiv.png')
    elif hasattr(settings, 'STATICFILES_DIRS') and settings.STATICFILES_DIRS:
        for static_dir in settings.STATICFILES_DIRS:
            test_path = os.path.join(static_dir, 'images', 'Meslas-Otomotiv.png')
            if os.path.exists(test_path):
                logo_path = test_path
                break
    if not logo_path:
        test_path = os.path.join(settings.BASE_DIR, 'static', 'images', 'Meslas-Otomotiv.png')
        if os.path.exists(test_path):
            logo_path = test_path
    
    # Başlık tablosu
    title_cell = [
        Paragraph("E-TEKLİF", title_style),
        Paragraph("Created by MesTakip", subtitle_style)
    ]
    
    if logo_path and os.path.exists(logo_path):
        logo = Image(logo_path, width=4*cm, height=2*cm)
        header_data = [[title_cell, logo]]
    else:
        header_data = [[title_cell, ""]]
    
    header_table = Table(header_data, colWidths=[12*cm, 6*cm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 0.5*cm))
    
    # Firma ve Müşteri Bilgileri
    info_data = [
        [Paragraph("<b>FİRMA BİLGİLERİ</b>", section_style), Paragraph("<b>MÜŞTERİ BİLGİLERİ</b>", section_style)],
        [
            Paragraph(f"<b>Ünvan:</b> MESLAS OTOMOTİV - ERHAN ERYILMAZ<br/>"
                     f"<b>Adres:</b> Abdurrahmangazi Mah. Atayolu Cad. No:187/A SANCAKTEPE / İSTANBUL<br/>"
                     f"<b>Telefon:</b> 0216 311 60 34<br/>"
                     f"<b>Web/E-Posta:</b> meslas.com / info@meslas.com<br/>"
                     f"<b>Vergi Dairesi/No:</b> SULTANBEYLİ V.D. / 4485882148", normal_style),
            Paragraph(f"<b>Cari:</b> {quotation.cari}<br/>"
                     f"<b>İlgili Kişi:</b> {quotation.ilgili_kisi or '-'}<br/>"
                     f"<b>Email:</b> {quotation.email or '-'}<br/>"
                     f"<b>Ödeme Şekli:</b> {quotation.odeme_sekli or '-'}", normal_style)
        ]
    ]
    
    info_table = Table(info_data, colWidths=[9*cm, 9*cm])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f8f9fa')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    story.append(info_table)
    story.append(Spacer(1, 0.3*cm))
    
    # Tarih Bilgileri
    story.append(Paragraph("<b>TARİH BİLGİLERİ</b>", section_style))
    date_data = [
        ["Teklif No", quotation.teklif_no],
        ["Teklif Tarihi", quotation.teklif_tarihi.strftime('%d.%m.%Y')],
        ["Son Geçerlilik Tarihi", quotation.gecerlilik_tarihi.strftime('%d.%m.%Y') if quotation.gecerlilik_tarihi else '-'],
        ["Düzenleyen", quotation.olusturan.get_full_name() or quotation.olusturan.username]
    ]
    
    date_table = Table(date_data, colWidths=[4*cm, 14*cm])
    date_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    story.append(date_table)
    story.append(Spacer(1, 0.3*cm))
    
    # Ürünler Tablosu
    story.append(Paragraph("<b>ÜRÜN/HİZMET DETAYLARI</b>", section_style))
    
    products_data = [["MARKA", "ÜRÜN AÇIKLAMASI", "MEVSİM", "MİKTAR", "BİRİM FİYAT", "TOPLAM"]]
    
    for item in quotation.urunler.all():
        mevsim_display = ""
        if item.mevsim == 'kis':
            mevsim_display = "Kış"
        elif item.mevsim == 'yaz':
            mevsim_display = "Yaz"
        elif item.mevsim == 'dort-mevsim':
            mevsim_display = "Dört Mevsim"
        else:
            mevsim_display = "-"
        
        products_data.append([
            item.marka or '-',
            item.urun_adi,
            mevsim_display,
            str(int(item.miktar)),
            f"{item.birim_fiyat:.2f} TL",
            f"{item.toplam:.2f} TL"
        ])
    
    products_table = Table(products_data, colWidths=[2*cm, 7*cm, 2*cm, 2*cm, 2.5*cm, 2.5*cm])
    products_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc3545')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),
        ('ALIGN', (2, 1), (2, -1), 'CENTER'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('ALIGN', (4, 1), (5, -1), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
    ]))
    story.append(products_table)
    story.append(Spacer(1, 0.3*cm))
    
    # Açıklama ve Toplamlar
    bottom_data = []
    
    if quotation.aciklama:
        aciklama_cell = Paragraph(f"<b>GENEL AÇIKLAMALAR</b><br/>{quotation.aciklama}", normal_style)
    else:
        aciklama_cell = ""
    
    totals_data = [
        ["TOPLAM MİKTAR", f"{quotation.toplam_adet}"],
        ["TOPLAM TUTAR", f"{quotation.ara_toplam:.2f} TL"],
        ["KDV %20", f"{quotation.kdv_tutari:.2f} TL"],
    ]
    
    totals_table = Table(totals_data, colWidths=[5*cm, 3.5*cm])
    totals_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#dee2e6')),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
    ]))
    
    grand_total_data = [[Paragraph(f"<b>KDV DAHİL GENEL TOPLAM: {quotation.genel_toplam:.2f} TL</b>", 
                                   ParagraphStyle('GrandTotal', parent=normal_style, fontSize=11, textColor=colors.white))]]
    grand_total_table = Table(grand_total_data, colWidths=[8.5*cm])
    grand_total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#dc3545')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('LEFTPADDING', (0, 0), (-1, -1), 8),
        ('RIGHTPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    
    totals_cell = [totals_table, Spacer(1, 0.2*cm), grand_total_table]
    
    bottom_data = [[aciklama_cell, totals_cell]]
    
    bottom_table = Table(bottom_data, colWidths=[9*cm, 9*cm])
    bottom_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    story.append(bottom_table)
    story.append(Spacer(1, 0.3*cm))
    
    # Footer Note
    footer_text = ("<b>Önemli Not:</b> Bu ileti hukuku korunmuş, gizli veya ifşa edilmemesi gereken bilgiler içerebilir. "
                  "Sayın mesajın gönderildiği kişi değilseniz, bu iletiye çoğaltmak ve dağıtmak yasaktır. "
                  "Bu mesajı yanlışlıkla alan kişi, bu durumu derhal gönderene telefon ya da e-posta ile bildirmeli ve bilgisayarından silmelidir.")
    
    footer_style = ParagraphStyle(
        'Footer',
        parent=normal_style,
        fontSize=7,
        textColor=colors.HexColor('#856404'),
        backColor=colors.HexColor('#fff3cd'),
        borderWidth=0,
        borderColor=colors.HexColor('#ffc107'),
        borderPadding=8,
        leftIndent=8,
        rightIndent=8,
        spaceBefore=5,
        spaceAfter=5
    )
    
    story.append(Paragraph(footer_text, footer_style))
    
    # PDF oluştur
    doc.build(story)
    
    # Response döndür
    pdf = buffer.getvalue()
    buffer.close()
    
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Teklif_{quotation.teklif_no}.pdf"'
    
    return response

@login_required
def quotation_edit(request, quotation_id):
    """Teklif düzenleme sayfası"""
    from .models import Quotation
    quotation = get_object_or_404(Quotation, id=quotation_id, olusturan=request.user)
    
    context = {
        'page_title': f'Teklif Düzenle - {quotation.teklif_no}',
        'quotation': quotation,
    }
    return render(request, 'dashboard/quotation_edit.html', context)

@login_required
@require_POST
def update_quotation(request, quotation_id):
    """Teklif güncelleme API endpoint'i"""
    try:
        data = json.loads(request.body)
        from .models import Quotation, QuotationItem
        from datetime import datetime
        
        quotation = get_object_or_404(Quotation, id=quotation_id, olusturan=request.user)
        
        # Teklif bilgilerini güncelle
        quotation.cari = data.get('cari', '')
        quotation.ilgili_kisi = data.get('ilgili_kisi', '')
        quotation.email = data.get('email', '')
        quotation.odeme_sekli = data.get('odeme_sekli', '')
        quotation.aciklama = data.get('aciklama', '')
        quotation.teklif_tarihi = datetime.strptime(data.get('teklif_tarihi'), '%Y-%m-%d').date()
        quotation.gecerlilik_tarihi = datetime.strptime(data.get('gecerlilik_tarihi'), '%Y-%m-%d').date() if data.get('gecerlilik_tarihi') else None
        quotation.durum = data.get('durum', 'acik')
        quotation.ara_toplam = Decimal(str(data.get('ara_toplam', 0)))
        quotation.kdv_tutari = Decimal(str(data.get('kdv_tutari', 0)))
        quotation.genel_toplam = Decimal(str(data.get('genel_toplam', 0)))
        quotation.rezerve = data.get('rezerve', False)
        quotation.proforma = data.get('proforma', False)
        quotation.ic_not = data.get('ic_not', '')
        quotation.save()
        
        # Mevcut ürünleri sil ve yenilerini ekle
        quotation.urunler.all().delete()
        
        for idx, item in enumerate(data.get('items', [])):
            QuotationItem.objects.create(
                teklif=quotation,
                marka=item.get('marka', ''),
                urun_adi=item.get('urun_adi', ''),
                mevsim=item.get('mevsim', ''),
                miktar=Decimal(str(item.get('miktar', 1))),
                birim_fiyat=Decimal(str(item.get('birim_fiyat', 0))),
                toplam=Decimal(str(item.get('toplam', 0))),
                sira=idx
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Teklif başarıyla güncellendi!'
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)

@login_required
def create_event(request):
    """Etkinlik oluşturma API endpoint'i"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            # Gerekli alanları kontrol et
            required_fields = ['title', 'date', 'time']
            for field in required_fields:
                if not data.get(field):
                    return JsonResponse({
                        'success': False,
                        'error': f'{field} alanı gereklidir'
                    }, status=400)
            
            # Tarih ve saat verilerini parse et
            from datetime import datetime
            date_obj = datetime.strptime(data['date'], '%Y-%m-%d').date()
            time_obj = datetime.strptime(data['time'], '%H:%M').time()
            
            # Event oluştur ve veritabanına kaydet
            event = Event.objects.create(
                title=data.get('title', ''),
                description=data.get('description', ''),
                type=data.get('type', 'event'),
                date=date_obj,
                time=time_obj,
                attendees=data.get('attendees', ''),
                recurring=data.get('recurring', False),
                recurrence=data.get('recurrence', 'none'),
                reminders=json.dumps(data.get('reminders', ['15'])),
                created_by=request.user
            )
            
            # Bildirim oluştur
            try:
                create_event_notifications(event)
            except Exception as e:
                print(f"Bildirim oluşturma hatası: {e}")
            
            return JsonResponse({
                'success': True,
                'message': 'Etkinlik başarıyla oluşturuldu!',
                'event': {
                    'id': event.id,
                    'title': event.title,
                    'date': event.date.strftime('%Y-%m-%d'),
                    'time': event.time.strftime('%H:%M'),
                    'description': event.description,
                    'location': event.location,
                    'type': event.type,
                    'priority': event.priority
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Geçersiz JSON verisi'
            }, status=400)
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)
    
    return JsonResponse({
        'success': False,
        'error': 'Sadece POST istekleri kabul edilir'
    }, status=405)

@login_required
@require_POST
def save_quotation(request):
    """Teklif kaydetme API endpoint'i"""
    try:
        data = json.loads(request.body)
        from .models import Quotation, QuotationItem
        from datetime import datetime
        
        # Teklif numarası oluştur
        last_quotation = Quotation.objects.order_by('-id').first()
        if last_quotation and last_quotation.teklif_no.isdigit():
            teklif_no = str(int(last_quotation.teklif_no) + 1)
        else:
            teklif_no = "1000"
        
        # Teklif oluştur
        quotation = Quotation.objects.create(
            teklif_no=teklif_no,
            cari=data.get('cari', ''),
            ilgili_kisi=data.get('ilgili_kisi', ''),
            email=data.get('email', ''),
            odeme_sekli=data.get('odeme_sekli', ''),
            aciklama=data.get('aciklama', ''),
            teklif_tarihi=datetime.strptime(data.get('teklif_tarihi'), '%Y-%m-%d').date(),
            gecerlilik_tarihi=datetime.strptime(data.get('gecerlilik_tarihi'), '%Y-%m-%d').date() if data.get('gecerlilik_tarihi') else None,
            durum=data.get('durum', 'acik'),
            ara_toplam=Decimal(str(data.get('ara_toplam', 0))),
            kdv_tutari=Decimal(str(data.get('kdv_tutari', 0))),
            genel_toplam=Decimal(str(data.get('genel_toplam', 0))),
            rezerve=data.get('rezerve', False),
            proforma=data.get('proforma', False),
            ic_not=data.get('ic_not', ''),
            olusturan=request.user
        )
        
        # Ürünleri kaydet
        for idx, item in enumerate(data.get('items', [])):
            QuotationItem.objects.create(
                teklif=quotation,
                marka=item.get('marka', ''),
                urun_adi=item.get('urun_adi', ''),
                mevsim=item.get('mevsim', ''),
                miktar=Decimal(str(item.get('miktar', 1))),
                birim_fiyat=Decimal(str(item.get('birim_fiyat', 0))),
                toplam=Decimal(str(item.get('toplam', 0))),
                sira=idx
            )
        
        return JsonResponse({
            'success': True,
            'message': 'Teklif başarıyla kaydedildi!',
            'teklif_no': teklif_no
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)

@login_required
@require_POST
def delete_quotation(request, quotation_id):
    """Teklif silme API endpoint'i"""
    try:
        from .models import Quotation
        quotation = get_object_or_404(Quotation, id=quotation_id, olusturan=request.user)
        teklif_no = quotation.teklif_no
        quotation.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'{teklif_no} numaralı teklif silindi!'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_POST
def send_quotation_email(request, quotation_id):
    """Teklifi email ile gönderme API endpoint'i"""
    try:
        from .models import Quotation
        from django.core.mail import EmailMessage
        from django.template.loader import render_to_string
        
        data = json.loads(request.body)
        quotation = get_object_or_404(Quotation, id=quotation_id, olusturan=request.user)
        
        email_to = data.get('email_to')
        subject = data.get('subject')
        message = data.get('message')
        
        if not email_to or not subject:
            return JsonResponse({
                'success': False,
                'error': 'Email ve konu alanları gereklidir'
            }, status=400)
        
        # HTML email içeriği oluştur
        html_content = render_to_string('dashboard/quotation_email.html', {
            'quotation': quotation,
            'message': message
        })
        
        # Email gönder
        email = EmailMessage(
            subject=subject,
            body=html_content,
            from_email=None,  # DEFAULT_FROM_EMAIL kullanılacak
            to=[email_to],
        )
        email.content_subtype = 'html'
        email.send()
        
        return JsonResponse({
            'success': True,
            'message': f'Email {email_to} adresine gönderildi!'
        })
        
    except Exception as e:
        import traceback
        return JsonResponse({
            'success': False,
            'error': str(e),
            'traceback': traceback.format_exc()
        }, status=500)

@login_required
def get_events(request):
    """Etkinlikleri getiren API endpoint'i"""
    try:
        # Tarih filtreleri
        start_date = request.GET.get('start')
        end_date = request.GET.get('end')
        
        events = Event.objects.filter(created_by=request.user)
        
        if start_date:
            events = events.filter(date__gte=start_date)
        if end_date:
            events = events.filter(date__lte=end_date)
        
        events_data = []
        for event in events:
            events_data.append({
                'id': event.id,
                'title': event.title,
                'type': event.type,
                'date': event.date.strftime('%Y-%m-%d'),
                'time': event.time.strftime('%H:%M'),
                'timeStr': event.time.strftime('%H:%M'),
                'dateStr': event.date.strftime('%d %b'),
                'description': event.description,
                'location': event.location,
                'priority': event.priority,
                'duration': event.duration,
                'attendees': event.attendees,
                'recurring': event.recurring,
                'recurrence': event.recurrence,
                'dateObj': event.date.isoformat(),
                'read': True
            })
        
        return JsonResponse({
            'success': True,
            'events': events_data
        })

    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def kategoriler(request):
    """Kategoriler sayfası - Kategori yönetimi"""
    if request.method == 'POST':
        # Yeni kategori ekleme
        kategori_adi = request.POST.get('kategori_adi')
        parent_id = request.POST.get('parent_id')
        alt_kategoriler = request.POST.getlist('alt_kategoriler[]')
        
        if kategori_adi:
            parent = None
            if parent_id:
                try:
                    parent = TransactionCategory.objects.get(id=parent_id)
                except TransactionCategory.DoesNotExist:
                    pass
            
            # Ana kategoriyi oluştur
            ana_kategori = TransactionCategory.objects.create(
                name=kategori_adi,
                parent=parent,
                created_by=request.user
            )
            
            # Alt kategorileri oluştur (eğer varsa ve parent_id yoksa)
            if not parent_id and alt_kategoriler:
                eklenen_alt_kategoriler = 0
                for alt_kategori_adi in alt_kategoriler:
                    if alt_kategori_adi.strip():  # Boş değilse
                        TransactionCategory.objects.create(
                            name=alt_kategori_adi.strip(),
                            parent=ana_kategori,
                            created_by=request.user
                        )
                        eklenen_alt_kategoriler += 1
                
                if eklenen_alt_kategoriler > 0:
                    messages.success(request, f'Ana kategori "{kategori_adi}" ve {eklenen_alt_kategoriler} alt kategori başarıyla eklendi!')
                else:
                    messages.success(request, f'Ana kategori "{kategori_adi}" başarıyla eklendi!')
            elif parent:
                messages.success(request, f'Alt kategori "{kategori_adi}" başarıyla "{parent.name}" kategorisinin altına eklendi!')
            else:
                messages.success(request, f'Ana kategori "{kategori_adi}" başarıyla eklendi!')
        else:
            messages.error(request, 'Kategori adı boş olamaz!')
        
        return redirect('dashboard:kategoriler')
    
    # Kategorileri hiyerarşik olarak getir (sadece kullanıcının kategorileri)
    ana_kategoriler = TransactionCategory.objects.filter(parent=None, created_by=request.user).order_by('name')
    alt_kategoriler = TransactionCategory.objects.filter(parent__isnull=False, created_by=request.user).order_by('name')
    
    # Hiyerarşik sıralama: Ana kategori -> Alt kategorileri -> Sonraki ana kategori
    tum_kategoriler = []
    for ana_kategori in ana_kategoriler:
        # Ana kategoriyi ekle
        tum_kategoriler.append(ana_kategori)
        # Bu ana kategorinin alt kategorilerini ekle
        alt_kategoriler_bu_ana = TransactionCategory.objects.filter(parent=ana_kategori, created_by=request.user).order_by('name')
        tum_kategoriler.extend(alt_kategoriler_bu_ana)
    
    # Tüm kullanıcıları getir (dropdown için)
    tum_kullanicilar = User.objects.all().order_by('username')
    
    context = {
        'page_title': 'Kategoriler',
        'ana_kategoriler': ana_kategoriler,
        'tum_kategoriler': tum_kategoriler,
        'alt_kategoriler': alt_kategoriler,
        'tum_kullanicilar': tum_kullanicilar,
        'stats': {
            'toplam': len(tum_kategoriler),
            'ana': ana_kategoriler.count(),
            'alt': alt_kategoriler.count(),
            'aktif': len(tum_kategoriler),
        }
    }
    return render(request, 'dashboard/kategoriler.html', context)

@login_required
def settings(request):
    """Settings sayfası"""
    from django.contrib.auth import update_session_auth_hash
    
    success_message = None
    error_message = None

    if request.method == 'POST':
        action = request.POST.get('action', 'profile')

        if action == 'profile':
            email = request.POST.get('email', '').strip()
            first_name = request.POST.get('first_name', '').strip()
            last_name = request.POST.get('last_name', '').strip()
            request.user.email = email
            request.user.first_name = first_name
            request.user.last_name = last_name
            request.user.save()
            success_message = 'Profil bilgileri güncellendi.'

        elif action == 'password':
            current_password = request.POST.get('current_password', '')
            new_password = request.POST.get('new_password', '')
            confirm_password = request.POST.get('confirm_password', '')
            if not request.user.check_password(current_password):
                error_message = 'Mevcut şifre hatalı.'
            elif new_password != confirm_password:
                error_message = 'Yeni şifreler eşleşmiyor.'
            elif len(new_password) < 6:
                error_message = 'Yeni şifre en az 6 karakter olmalıdır.'
            else:
                request.user.set_password(new_password)
                request.user.save()
                update_session_auth_hash(request, request.user)
                success_message = 'Şifre başarıyla değiştirildi.'

    context = {
        'page_title': 'Ayarlar',
        'success_message': success_message,
        'error_message': error_message,
    }
    return render(request, 'dashboard/settings.html', context)

@login_required
def security(request):
    """Security sayfası"""
    from django.contrib.auth import update_session_auth_hash

    success_message = None
    error_message = None

    if request.method == 'POST':
        current_password = request.POST.get('current_password', '')
        new_password = request.POST.get('new_password', '')
        confirm_password = request.POST.get('confirm_password', '')

        if not request.user.check_password(current_password):
            error_message = 'Mevcut şifre hatalı.'
        elif new_password != confirm_password:
            error_message = 'Yeni şifreler eşleşmiyor.'
        elif len(new_password) < 8:
            error_message = 'Yeni şifre en az 8 karakter olmalıdır.'
        else:
            request.user.set_password(new_password)
            request.user.save()
            update_session_auth_hash(request, request.user)
            success_message = 'Şifre başarıyla güncellendi.'

    # Son giriş bilgisi
    last_login = request.user.last_login

    context = {
        'page_title': 'Güvenlik',
        'success_message': success_message,
        'error_message': error_message,
        'last_login': last_login,
    }
    return render(request, 'dashboard/security.html', context)

@login_required
def help(request):
    """Help sayfası"""
    context = {
        'page_title': 'Help & Support',
    }
    return render(request, 'dashboard/help.html', context)

@login_required
def export_excel(request):
    """Sipariş Envanterini Excel'e aktar"""
    # Filtreleme parametrelerini al
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    grup = request.GET.get('grup', '')
    durum = request.GET.get('durum', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Siparişleri filtrele (iptal edilenleri ve kontrol edilenleri hariç tut, sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(user=request.user).exclude(durum__in=['iptal', 'kontrol'])
    
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        # Türkçe karakter varyantları oluştur ve hepsiyle ara
        search_variants = create_turkish_search_variants(marka)
        q_objects = Q()
        for variant in search_variants:
            q_objects |= Q(marka__icontains=variant)
        siparisler = siparisler.filter(q_objects)
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if durum:
        siparisler = siparisler.filter(durum=durum)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Sipariş Envanteri"
    
    # Başlık satırı
    headers = [
        'CARI (FIRMA)', 'ÜRÜN', 'MARKA', 'GRUP', 'MEVSİM', 'ADET', 
        'BİRİM FİYAT', 'DURUM', 'AMBAR', 'AÇIKLAMA 1', 'TOPLAM FİYAT', 
        'ÖDEME', 'SMS', 'ÖNE ÇIKAR', 'OLUŞTURMA TARİHİ', 'GÜNCELLEME TARİHİ'
    ]
    
    # Başlık stilini ayarla
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Veri satırları
    for row, siparis in enumerate(siparisler, 2):
        ws.cell(row=row, column=1, value=siparis.cari_firma)
        ws.cell(row=row, column=2, value=siparis.urun)
        ws.cell(row=row, column=3, value=siparis.marka)
        ws.cell(row=row, column=4, value=siparis.get_grup_display())
        ws.cell(row=row, column=5, value=siparis.get_mevsim_display())
        ws.cell(row=row, column=6, value=siparis.adet)
        ws.cell(row=row, column=7, value=siparis.birim_fiyat)
        ws.cell(row=row, column=8, value=siparis.get_durum_display())
        ws.cell(row=row, column=9, value=siparis.get_ambar_display())
        ws.cell(row=row, column=10, value=siparis.aciklama)
        ws.cell(row=row, column=11, value=siparis.toplam_fiyat)
        ws.cell(row=row, column=12, value=siparis.get_odeme_display())
        ws.cell(row=row, column=13, value=siparis.get_sms_durum_display())
        ws.cell(row=row, column=14, value="Evet" if siparis.one_cikar else "Hayır")
        ws.cell(row=row, column=15, value=siparis.olusturma_tarihi.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=16, value=siparis.guncelleme_tarihi.strftime('%d.%m.%Y %H:%M'))
    
    # Sütun genişliklerini ayarla
    column_widths = [20, 25, 15, 12, 12, 8, 12, 15, 10, 30, 12, 12, 10, 12, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # HTTP response oluştur
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="siparis_envanteri.xlsx"'
    
    # Excel dosyasını response'a yaz
    wb.save(response)
    return response

@login_required
def export_cancelled_excel(request):
    """İptal Edilen Siparişleri Excel'e aktar"""
    # Filtreleme parametrelerini al
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    grup = request.GET.get('grup', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Sadece iptal edilen siparişleri filtrele (sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(durum='iptal', user=request.user)
    
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        # Türkçe karakter varyantları oluştur ve hepsiyle ara
        search_variants = create_turkish_search_variants(marka)
        q_objects = Q()
        for variant in search_variants:
            q_objects |= Q(marka__icontains=variant)
        siparisler = siparisler.filter(q_objects)
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "İptal Edilen Siparişler"
    
    # Başlık satırı
    headers = [
        'ID', 'CARI (FIRMA)', 'ÜRÜN', 'MARKA', 'GRUP', 'MEVSİM', 'ADET', 
        'BİRİM FİYAT', 'AMBAR', 'AÇIKLAMA', 'TOPLAM FİYAT', 
        'ÖDEME', 'SMS', 'ÖNE ÇIKAR', 'OLUŞTURMA TARİHİ', 'İPTAL TARİHİ'
    ]
    
    # Başlık stilini ayarla
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")  # Kırmızı renk
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Veri satırları
    for row, siparis in enumerate(siparisler, 2):
        ws.cell(row=row, column=1, value=siparis.id)
        ws.cell(row=row, column=2, value=siparis.cari_firma)
        ws.cell(row=row, column=3, value=siparis.urun)
        ws.cell(row=row, column=4, value=siparis.marka)
        ws.cell(row=row, column=5, value=siparis.get_grup_display())
        ws.cell(row=row, column=6, value=siparis.get_mevsim_display())
        ws.cell(row=row, column=7, value=siparis.adet)
        ws.cell(row=row, column=8, value=siparis.birim_fiyat)
        ws.cell(row=row, column=9, value=siparis.get_ambar_display())
        ws.cell(row=row, column=10, value=siparis.aciklama)
        ws.cell(row=row, column=11, value=siparis.toplam_fiyat)
        ws.cell(row=row, column=12, value=siparis.get_odeme_display())
        ws.cell(row=row, column=13, value=siparis.get_sms_durum_display())
        ws.cell(row=row, column=14, value="Evet" if siparis.one_cikar else "Hayır")
        ws.cell(row=row, column=15, value=siparis.olusturma_tarihi.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=16, value=siparis.guncelleme_tarihi.strftime('%d.%m.%Y %H:%M'))
    
    # Sütun genişliklerini ayarla
    column_widths = [8, 20, 25, 15, 12, 12, 8, 12, 10, 30, 12, 12, 10, 12, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # HTTP response oluştur
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="iptal_edilen_siparisler.xlsx"'
    
    # Excel dosyasını response'a yaz
    wb.save(response)
    return response

@login_required
def export_checked_excel(request):
    """Kontrol Edilen Siparişleri Excel'e aktar"""
    # Filtreleme parametrelerini al
    firma = request.GET.get('firma', '')
    marka = request.GET.get('marka', '')
    grup = request.GET.get('grup', '')
    mevsim = request.GET.get('mevsim', '')
    ambar = request.GET.get('ambar', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Sadece kontrol edilen siparişleri filtrele (sadece kullanıcının siparişleri)
    siparisler = Siparis.objects.filter(durum='kontrol', user=request.user)
    
    if firma:
        siparisler = siparisler.filter(cari_firma__icontains=firma)
    if marka:
        # Türkçe karakter varyantları oluştur ve hepsiyle ara
        search_variants = create_turkish_search_variants(marka)
        q_objects = Q()
        for variant in search_variants:
            q_objects |= Q(marka__icontains=variant)
        siparisler = siparisler.filter(q_objects)
    if grup:
        siparisler = siparisler.filter(grup=grup)
    if mevsim:
        siparisler = siparisler.filter(mevsim=mevsim)
    if ambar:
        siparisler = siparisler.filter(ambar=ambar)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    
    # Eğer hiçbir tarih filtresi yoksa, varsayılan olarak son 3 ayın verilerini getir
    if not tarih_filtre and not baslangic_tarihi and not bitis_tarihi:
        start_date = now - timedelta(days=90)
        siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        tarih_filtre = 'son-3-ay'  # Varsayılan filtreyi işaretle
    
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            start_date = now.replace(hour=0, minute=0, second=0, microsecond=0)
            end_date = now.replace(hour=23, minute=59, second=59, microsecond=999999)
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            start_date = start_date.replace(hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d')
            start_date = timezone.make_aware(start_date.replace(hour=0, minute=0, second=0))
            siparisler = siparisler.filter(olusturma_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d')
            end_date = timezone.make_aware(end_date.replace(hour=23, minute=59, second=59))
            siparisler = siparisler.filter(olusturma_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Kontrol Edilen Siparişler"
    
    # Başlık satırı
    headers = [
        'CARI (FIRMA)', 'ÜRÜN', 'MARKA', 'GRUP', 'MEVSİM', 'ADET', 
        'BİRİM FİYAT', 'AMBAR', 'AÇIKLAMA', 'TOPLAM FİYAT', 
        'ÖDEME', 'SMS', 'ÖNE ÇIKAR', 'OLUŞTURMA TARİHİ', 'KONTROL TARİHİ'
    ]
    
    # Başlık stilini ayarla
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")  # Sarı renk
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Veri satırları
    for row, siparis in enumerate(siparisler, 2):
        ws.cell(row=row, column=1, value=siparis.cari_firma)
        ws.cell(row=row, column=2, value=siparis.urun)
        ws.cell(row=row, column=3, value=siparis.marka)
        ws.cell(row=row, column=4, value=siparis.get_grup_display())
        ws.cell(row=row, column=5, value=siparis.get_mevsim_display())
        ws.cell(row=row, column=6, value=siparis.adet)
        ws.cell(row=row, column=7, value=siparis.birim_fiyat)
        ws.cell(row=row, column=8, value=siparis.get_ambar_display())
        ws.cell(row=row, column=9, value=siparis.aciklama)
        ws.cell(row=row, column=10, value=siparis.toplam_fiyat)
        ws.cell(row=row, column=11, value=siparis.get_odeme_display())
        ws.cell(row=row, column=12, value=siparis.get_sms_durum_display())
        ws.cell(row=row, column=13, value="Evet" if siparis.one_cikar else "Hayır")
        ws.cell(row=row, column=14, value=siparis.olusturma_tarihi.strftime('%d.%m.%Y %H:%M'))
        ws.cell(row=row, column=15, value=siparis.guncelleme_tarihi.strftime('%d.%m.%Y %H:%M'))
    
    # Sütun genişliklerini ayarla
    column_widths = [20, 25, 15, 12, 12, 8, 12, 10, 30, 12, 12, 10, 12, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # HTTP response oluştur
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="kontrol_edilen_siparisler.xlsx"'
    
    # Excel dosyasını response'a yaz
    wb.save(response)
    return response

def login_view(request):
    """Kullanıcı giriş sayfası"""
    if request.user.is_authenticated:
        return redirect('dashboard:index')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        if username and password:
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                messages.success(request, f'Hoş geldiniz, {user.first_name or user.username}!')
                next_url = request.GET.get('next', 'dashboard:index')
                return redirect(next_url)
            else:
                messages.error(request, 'Kullanıcı adı veya şifre hatalı!')
        else:
            messages.error(request, 'Lütfen tüm alanları doldurun!')
    
    context = {
        'page_title': 'Giriş Yap',
    }
    return render(request, 'dashboard/login.html', context)

def logout_view(request):
    """Kullanıcı çıkış"""
    if request.user.is_authenticated:
        username = request.user.first_name or request.user.username
        logout(request)
        messages.success(request, f'Güle güle, {username}!')
    return redirect('dashboard:login')

# Bildirim API'leri

@login_required
def get_notifications(request):
    """Kullanıcının bildirimlerini getiren API endpoint'i"""
    try:
        # Kullanıcının bildirimlerini getir
        notifications = Notification.objects.filter(user=request.user).order_by('-scheduled_time')[:20]
        
        notifications_data = []
        for notification in notifications:
            # Ertelenmiş bildirimleri kontrol et
            is_snoozed = notification.is_snoozed()
            
            notifications_data.append({
                'id': notification.id,
                'title': notification.title,
                'message': notification.message,
                'type': notification.type,
                'status': notification.status,
                'icon': notification.get_type_icon(),
                'color': notification.get_type_color(),
                'scheduled_time': notification.scheduled_time.isoformat(),
                'sent_time': notification.sent_time.isoformat() if notification.sent_time else None,
                'read_time': notification.read_time.isoformat() if notification.read_time else None,
                'snoozed_until': notification.snoozed_until.isoformat() if notification.snoozed_until else None,
                'snooze_count': notification.snooze_count,
                'is_snoozed': is_snoozed,
                'event_id': None,  # Event geçici olarak devre dışı
                'is_overdue': notification.is_overdue(),
                'extra_data': notification.get_extra_data_dict()
            })
        
        # Okunmamış bildirim sayısı (ertelenmiş olanları dahil etme)
        unread_count = Notification.objects.filter(
            user=request.user, 
            status__in=['pending', 'sent']
        ).count()
        
        return JsonResponse({
            'success': True,
            'notifications': notifications_data,
            'unread_count': unread_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def mark_notification_read(request, notification_id):
    """Bildirimi okundu olarak işaretle"""
    try:
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.mark_as_read()
        
        return JsonResponse({
            'success': True,
            'message': 'Bildirim okundu olarak işaretlendi'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def mark_all_notifications_read(request):
    """Tüm bildirimleri okundu olarak işaretle"""
    try:
        notifications = Notification.objects.filter(
            user=request.user, 
            status__in=['pending', 'sent']
        )
        
        for notification in notifications:
            notification.mark_as_read()
        
        return JsonResponse({
            'success': True,
            'message': f'{notifications.count()} bildirim okundu olarak işaretlendi'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def dismiss_notification(request, notification_id):
    """Bildirimi kapat"""
    try:
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.status = 'dismissed'
        notification.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Bildirim kapatıldı'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def mark_notification_sent(request, notification_id):
    """Bildirimi gönderildi olarak işaretle"""
    try:
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        if notification.status == 'pending':
            notification.status = 'sent'
            notification.sent_time = timezone.now()
            notification.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Bildirim gönderildi olarak işaretlendi'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
@require_POST
def snooze_notification(request, notification_id):
    """Bildirimi ertele"""
    try:
        data = json.loads(request.body)
        minutes = int(data.get('minutes', 5))
        
        notification = get_object_or_404(Notification, id=notification_id, user=request.user)
        notification.snooze(minutes)
        
        return JsonResponse({
            'success': True,
            'message': f'Bildirim {minutes} dakika ertelendi',
            'snoozed_until': notification.snoozed_until.isoformat() if notification.snoozed_until else None
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

def create_event_notifications(event):
    """Etkinlik için bildirimler oluştur"""
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    try:
        # Hatırlatıcı zamanlarını parse et
        reminders = json.loads(event.reminders) if event.reminders else ['15']
        
        # Etkinlik tarih ve saatini birleştir
        naive_datetime = datetime.combine(event.date, event.time)
        
        # Timezone aware yap
        if timezone.is_naive(naive_datetime):
            event_datetime = timezone.make_aware(naive_datetime)
        else:
            event_datetime = naive_datetime
        
        now = timezone.now()
        
        # Her hatırlatıcı için bildirim oluştur
        for reminder_minutes in reminders:
            try:
                minutes = int(reminder_minutes)
                reminder_time = event_datetime - timedelta(minutes=minutes)
                
                # Geçmiş tarih kontrolü
                if reminder_time > now:
                    # Hatırlatıcı mesajını oluştur
                    if minutes == 0:
                        reminder_msg = f"'{event.title}' etkinliği şimdi başlıyor!"
                    elif minutes < 60:
                        reminder_msg = f"'{event.title}' etkinliği {minutes} dakika sonra başlayacak."
                    elif minutes < 1440:
                        hours = minutes // 60
                        reminder_msg = f"'{event.title}' etkinliği {hours} saat sonra başlayacak."
                    else:
                        days = minutes // 1440
                        reminder_msg = f"'{event.title}' etkinliği {days} gün sonra başlayacak."
                    
                    if event.location:
                        reminder_msg += f" Konum: {event.location}"
                    
                    # Hatırlatıcı bildirimi oluştur
                    Notification.objects.create(
                        title=f"Etkinlik Hatırlatıcısı: {event.title}",
                        message=reminder_msg,
                        type='event_reminder',
                        user=event.created_by,
                        scheduled_time=reminder_time,
                        status='pending',
                        extra_data=json.dumps({
                            'event_id': event.id,
                            'reminder_minutes': minutes,
                            'event_type': event.type,
                            'event_priority': event.priority
                        })
                    )
            except (ValueError, TypeError):
                continue
            
    except Exception as e:
        # Hata logla
        print(f"Bildirim oluşturma hatası: {e}")
        pass

# Etkinlik oluşturma view'ını güncelle (devre dışı sürüm kullanılmıyor)
# Aşağıdaki sürüm kaldırıldı; aktif olan basit `create_event` yukarıda tanımlıdır.

@login_required
@misafir_forbidden
def finance(request):
    """Gelir/Gider İşlemleri sayfası"""
    form = TransactionForm(request.POST or None, user=request.user)
    
    if request.method == 'POST':
        if form.is_valid():
            transaction = form.save(commit=False)
            transaction.created_by = request.user
            transaction.save()
            
            # Pafgo işlemi varsa özel mesaj
            if transaction.pafgo and float(transaction.pafgo) > 0:
                messages.success(request, f'Pafgo işlemi başarıyla kaydedildi! Tutar: {transaction.pafgo} TL - Gelir/Gider Raporunda görüntüleyebilirsiniz.')
            else:
                messages.success(request, 'İşlem başarıyla kaydedildi!')
            
            # next parametresi varsa oraya yönlendir
            next_url = request.POST.get('next') or request.GET.get('next')
            if next_url and next_url.startswith('/'):
                return redirect(next_url)
            return redirect('dashboard:products')
        else:
            messages.error(request, 'Form hatası! Lütfen alanları kontrol edin.')
            print("Form Errors:", form.errors)  # Debug için
    
    # Ana kategorileri context'e ekle
    ana_kategoriler = TransactionCategory.objects.filter(
        parent=None, 
        created_by=request.user
    ).order_by('name')
    
    # Son işlemleri getir (kategorileri ve parent kategorileri de yükle)
    last_transactions = Transaction.objects.filter(
        created_by=request.user
    ).select_related(
        'kategori1', 'kategori1__parent', 'kategori2', 'kategori2__parent', 'kategori3', 'kategori3__parent'
    ).order_by('-tarih', '-id')[:10]
    
    # Debug için
    print(f"Ana Kategoriler Sayısı: {ana_kategoriler.count()}")
    for kat in ana_kategoriler:
        print(f"  - {kat.name} (ID: {kat.id})")
    
    context = {
        'page_title': 'Gelir/Gider İşlemleri',
        'form': form,
        'ana_kategoriler': ana_kategoriler,
        'transactions': last_transactions,
    }
    return render(request, 'dashboard/finance.html', context)

@login_required
def transaction_duzenle(request, transaction_id):
    """Transaction düzenleme sayfası"""
    transaction = get_object_or_404(Transaction, id=transaction_id, created_by=request.user)
    
    if request.method == 'POST':
        form = TransactionForm(request.POST, instance=transaction, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, 'İşlem başarıyla güncellendi!')
            
            # next parametresi varsa oraya yönlendir
            next_url = request.POST.get('next') or request.GET.get('next')
            if next_url and next_url.startswith('/'):
                return redirect(next_url)
            return redirect('dashboard:products')
        else:
            messages.error(request, 'Form hatası! Lütfen alanları kontrol edin.')
    else:
        form = TransactionForm(instance=transaction, user=request.user)
    
    # Ana kategorileri context'e ekle
    ana_kategoriler = TransactionCategory.objects.filter(
        parent=None, 
        created_by=request.user
    ).order_by('name')
    
    context = {
        'page_title': 'İşlem Düzenle',
        'form': form,
        'transaction': transaction,
        'ana_kategoriler': ana_kategoriler,
    }
    return render(request, 'dashboard/transaction_duzenle.html', context)

@login_required
def kategori_sil(request, kategori_id):
    """Kategori silme"""
    try:
        kategori = get_object_or_404(TransactionCategory, id=kategori_id, created_by=request.user)
        kategori_adi = kategori.name
        kategori.delete()
        messages.success(request, f'Kategori "{kategori_adi}" başarıyla silindi!')
    except Exception as e:
        messages.error(request, f'Kategori silinirken hata oluştu: {str(e)}')
    
    return redirect('dashboard:kategoriler')

@login_required
def kategori_duzenle(request, kategori_id):
    """Kategori düzenleme"""
    kategori = get_object_or_404(TransactionCategory, id=kategori_id, created_by=request.user)
    
    if request.method == 'POST':
        kategori_adi = request.POST.get('kategori_adi')
        parent_id = request.POST.get('parent_id')
        
        if kategori_adi:
            # Üst kategori kontrolü
            parent = None
            if parent_id:
                try:
                    parent = TransactionCategory.objects.get(id=parent_id, created_by=request.user)
                    # Kendi kendisinin alt kategorisi olamaz
                    if parent.id == kategori.id:
                        messages.error(request, 'Kategori kendi kendisinin alt kategorisi olamaz!')
                        return redirect('dashboard:kategoriler')
                except TransactionCategory.DoesNotExist:
                    pass
            
            # Kategoriyi güncelle (created_by değiştirilmez)
            eski_ad = kategori.name
            kategori.name = kategori_adi
            kategori.parent = parent
            kategori.save()
            
            messages.success(request, f'Kategori "{eski_ad}" → "{kategori_adi}" olarak güncellendi!')
        else:
            messages.error(request, 'Kategori adı boş olamaz!')
    
    return redirect('dashboard:kategoriler')

@login_required
def get_alt_kategoriler(request, ana_kategori_id):
    """Ana kategoriye ait alt kategorileri JSON olarak döndür"""
    try:
        # Ana kategoriyi kontrol et
        ana_kategori = get_object_or_404(
            TransactionCategory, 
            id=ana_kategori_id, 
            parent=None, 
            created_by=request.user
        )
        
        # Alt kategorileri getir
        alt_kategoriler = TransactionCategory.objects.filter(
            parent=ana_kategori,
            created_by=request.user
        ).order_by('order', 'name')
        
        # JSON formatında döndür
        data = {
            'success': True,
            'ana_kategori': {
                'id': ana_kategori.id,
                'name': ana_kategori.name
            },
            'alt_kategoriler': [
                {
                    'id': alt_kat.id,
                    'name': alt_kat.name
                }
                for alt_kat in alt_kategoriler
            ]
        }
        
        return JsonResponse(data)
    except TransactionCategory.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': 'Ana kategori bulunamadı'
        }, status=404)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)

@login_required
def brand_distribution_api(request):
    """Lastik marka dağılımı API endpoint'i - Canlı veri"""
    try:
        # Kullanıcının siparişlerinden marka dağılımını hesapla
        # İptal edilen siparişleri hariç tut
        brand_distribution = Siparis.objects.filter(
            user=request.user
        ).exclude(
            durum='iptal'
        ).values('marka').annotate(
            total_adet=Sum('adet')
        ).order_by('-total_adet')
        
        # Veri hazırla
        brands = []
        colors = [
            '#3b82f6', '#ef4444', '#10b981', '#f59e0b', 
            '#8b5cf6', '#ec4899', '#06b6d4', '#84cc16',
            '#f97316', '#06b6d4', '#8b5cf6', '#ef4444'
        ]
        
        for i, brand in enumerate(brand_distribution):
            if brand['marka'] and brand['total_adet']:  # Boş marka adlarını filtrele
                brands.append({
                    'name': brand['marka'],
                    'count': brand['total_adet'],
                    'color': colors[i % len(colors)]
                })
        
        return JsonResponse({
            'success': True,
            'brands': brands,
            'total_brands': len(brands),
            'total_count': sum(brand['count'] for brand in brands)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)

@login_required
def malzeme_excel_upload(request):
    if request.method == 'POST':
        form = MalzemeExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES['file']
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_file)
                ws = wb.active
                
                # Create MalzemeDosya first
                dosya = MalzemeDosya.objects.create(
                    dosya_adi=excel_file.name,
                    kullanici=request.user
                )
                
                # Get headers from first row
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                
                eklenen = 0
                for row_num in range(2, ws.max_row + 1):
                    try:
                        # Create row dict
                        row_data = {}
                        for col_num, header in enumerate(headers, 1):
                            if col_num <= len(headers):
                                cell_value = ws.cell(row=row_num, column=col_num).value
                                row_data[header] = cell_value
                        
                        tarih = row_data.get('TARİH') or row_data.get('TARIH') or ''
                        # Handle different date formats
                        if isinstance(tarih, datetime.datetime):
                            tarih = tarih.date()
                        elif isinstance(tarih, str) and tarih and '.' in tarih:
                            try:
                                tarih = datetime.datetime.strptime(tarih, '%d.%m.%Y').date()
                            except:
                                tarih = date.today()
                        elif isinstance(tarih, str) and tarih and '-' in tarih:
                            try:
                                tarih = datetime.datetime.strptime(tarih, '%Y-%m-%d').date()
                            except:
                                tarih = date.today()
                        else:
                            tarih = date.today()

                        tutar_raw = row_data.get('TUTAR') or '0'
                        tutar = parse_decimal_value(tutar_raw)
                            
                        hareket = MalzemeHareketi(
                            dosya=dosya,
                            tarih=tarih,
                            faturano=str(row_data.get('FATURA NO') or row_data.get('FATURANO') or '')[:100],
                            musteri=str(row_data.get('MÜŞTERİ') or row_data.get('MÜŞTERI') or '')[:255],
                            kategori=str(row_data.get('KATEGORİ') or row_data.get('KATEGORI') or '')[:255],
                            urun=str(row_data.get('ÜRÜN') or row_data.get('URUN') or '')[:255],
                            tutar=tutar,
                            odeme_sekli=str(row_data.get('ÖDEME ŞEKLİ') or row_data.get('ÖDEME PLANI') or '')[:100],
                            ref=str(row_data.get('REF') or '')[:100],
                            kullanici=request.user,
                        )
                        hareket.save()
                        eklenen += 1
                    except Exception as e:
                        print(f"Row error: {e}")
                        continue
                        
                messages.success(request, f"Başarıyla {eklenen} satır kaydedildi!")
                return redirect('dashboard:products')
            except Exception as e:
                print(f"Excel upload error: {e}")
                messages.error(request, f"Excel yükleme hatası: {str(e)}")
                return redirect('dashboard:finance')
                messages.error(request, f"Dosya okunamadı: {str(e)}")
        else:
            messages.error(request, "Lütfen geçerli bir dosya seçin.")
    else:
        form = MalzemeExcelUploadForm()
    return render(request, 'dashboard/malzeme_excel_upload.html', {'form': form})

def _safe_print(*args, **kwargs):
    """Windows charmap encoding hatasını önlemek için güvenli print (₺ vb. Unicode karakterler için)"""
    import sys
    s = ' '.join(str(a) for a in args)
    try:
        print(*args, **kwargs)
    except UnicodeEncodeError:
        sys.stdout.buffer.write((s + '\n').encode('utf-8', errors='replace'))


@csrf_exempt
@login_required
def malzeme_excel_kaydet(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Sadece POST!'}, status=400)
    try:
        # JSON parsing
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError as e:
            return JsonResponse({'success': False, 'error': f'JSON parsing hatası: {str(e)}'}, status=400)
        
        filename = data.get('filename')
        rows = data.get('rows', [])
        
        # Debug: Gelen veriyi kontrol et (Unicode güvenli)
        _safe_print(f"Gelen dosya: {filename}")
        _safe_print(f"Gelen satır sayısı: {len(rows)}")
        _safe_print(f"Request body boyutu: {len(request.body)} bytes")
        if rows:
            _safe_print(f"İlk satır örneği: {rows[0]}")
            _safe_print(f"İlk satır keys: {list(rows[0].keys()) if isinstance(rows[0], dict) else 'Not dict'}")
        
        if not filename:
            return JsonResponse({'success': False, 'error': 'Dosya adı eksik.'}, status=400)
        if not isinstance(rows, list):
            return JsonResponse({'success': False, 'error': f'Rows veri tipi yanlış: {type(rows)}'}, status=400)
        if len(rows) == 0:
            return JsonResponse({'success': False, 'error': 'Excel dosyasında veri bulunamadı.'}, status=400)
            
        dosya = MalzemeDosya.objects.create(dosya_adi=filename, kullanici=request.user)
        _safe_print(f"Dosya oluşturuldu: ID={dosya.id}")
        
        eklenen = 0
        hatalar = []
        
        for i, row in enumerate(rows):
            try:
                # Debug: Satır verilerini kontrol et (Unicode güvenli)
                _safe_print(f"Satır {i+1}: {row}")
                
                if not row or not isinstance(row, dict):
                    _safe_print(f"Satır {i+1} geçersiz: {row}")
                    continue
                
                # Tarih işleme - daha esnek
                tarih_raw = None
                for key in row.keys():
                    if any(x in key.upper() for x in ['TARİH', 'TARIH', 'DATE']):
                        tarih_raw = row[key]
                        break
                
                tarih = date.today()  # Default tarih
                
                if tarih_raw:
                    try:
                        if isinstance(tarih_raw, str) and tarih_raw.strip():
                            tarih_str = tarih_raw.strip()
                            if '.' in tarih_str:
                                # DD.MM.YYYY veya DD.MM.YY formatı
                                parts = tarih_str.split('.')
                                if len(parts) == 3:
                                    day, month, year = parts
                                    if len(year) == 2:
                                        year = '20' + year if int(year) < 50 else '19' + year
                                    tarih = date(int(year), int(month), int(day))
                            elif '-' in tarih_str:
                                # YYYY-MM-DD formatı
                                tarih = datetime.datetime.strptime(tarih_str, '%Y-%m-%d').date()
                        elif isinstance(tarih_raw, (int, float)) and tarih_raw > 0:
                            # Excel serial date
                            tarih = datetime.fromordinal(date(1900,1,1).toordinal() + int(tarih_raw) - 2).date()
                    except Exception as e:
                        _safe_print(f"Tarih parse hatası: {e}, raw: {tarih_raw}")
                        tarih = date.today()
                
                # Tüm sütun adlarını kontrol et ve esnek eşleştirme yap
                row_keys = list(row.keys())
                _safe_print(f"Satır {i+1} sütunları: {row_keys}")
                
                # Esnek sütun eşleştirmesi - daha geniş arama
                faturano = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['FATURA', 'INVOICE', 'NO', 'BELGE']):
                        val = row[key]
                        faturano = str(val).strip() if val is not None else ''
                        break
                
                musteri = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['MÜŞTERİ', 'MÜŞTERI', 'MUSTERI', 'CUSTOMER', 'CLIENT', 'CARİ', 'CARI']):
                        val = row[key]
                        musteri = str(val).strip() if val is not None else ''
                        break
                
                urun = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['ÜRÜN', 'URUN', 'PRODUCT', 'ITEM', 'MALZEME', 'HİZMET', 'HIZMET']):
                        val = row[key]
                        urun = str(val).strip() if val is not None else ''
                        break
                
                kategori = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['KATEGORİ', 'KATEGORI', 'CATEGORY', 'CAT', 'TİP', 'TIP', 'TYPE']):
                        val = row[key]
                        kategori = str(val).strip() if val is not None else ''
                        break
                
                marka = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['MARKA', 'BRAND', 'MAKE']):
                        val = row[key]
                        marka = str(val).strip() if val is not None else ''
                        break
                
                tutar_raw = 0
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['TUTAR', 'AMOUNT', 'PRICE', 'FİYAT', 'FIYAT', 'MIKTAR', 'TOPLAM']):
                        tutar_raw = row[key]
                        break
                
                odeme_sekli = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if any(x in key_upper for x in ['ÖDEME', 'ODEME', 'PAYMENT', 'PAY', 'PLAN']):
                        val = row[key]
                        odeme_sekli = str(val).strip() if val is not None else ''
                        break
                
                ref = ''
                for key in row_keys:
                    key_upper = key.upper().strip()
                    if key_upper == 'REF' or 'REF' in key_upper:
                        val = row[key]
                        ref = str(val).strip() if val is not None else ''
                        break
                
                tutar = parse_decimal_value(tutar_raw)
                
                _safe_print(f"İşlenmiş veri: TARİH={tarih}, FATURANO='{faturano}', MÜŞTERİ='{musteri}', ÜRÜN='{urun}', TUTAR={tutar}, ÖDEME='{odeme_sekli}'")
                
                # Daha esnek kontrol - en az bir anlamlı veri olsun
                has_meaningful_data = any([
                    faturano and len(faturano.strip()) > 0,
                    musteri and len(musteri.strip()) > 0,
                    urun and len(urun.strip()) > 0,
                    tutar > 0,
                    odeme_sekli and len(odeme_sekli.strip()) > 0
                ])
                
                if not has_meaningful_data:
                    _safe_print(f"Satır {i+1} anlamlı veri yok, atlanıyor")
                    continue
                
                # Kayıt oluştur - boş alanları varsayılan değerlerle doldur
                try:
                    hareket = MalzemeHareketi.objects.create(
                        dosya=dosya,
                        tarih=tarih,
                        faturano=(faturano or f'AUTO-{i+1}')[:100],  # Boşsa otomatik numara
                        musteri=(musteri or 'Belirtilmemiş')[:255],
                        kategori=(kategori or '')[:255],
                        marka=(marka or '')[:255],
                        urun=(urun or 'Belirtilmemiş')[:255],
                        tutar=tutar,
                        odeme_sekli=(odeme_sekli or 'Belirtilmemiş')[:100],
                        ref=(ref or '')[:100],
                        kullanici=request.user,
                    )
                    _safe_print(f"Kayıt başarılı: ID={hareket.id}")
                    eklenen += 1
                except Exception as db_error:
                    error_msg = f"Satır {i+1} DB hatası: {db_error}"
                    _safe_print(error_msg)
                    hatalar.append(error_msg)
                    continue
                
            except Exception as e:
                error_msg = f"Satır {i+1} işlem hatası: {e}"
                _safe_print(error_msg)
                hatalar.append(error_msg)
                continue
        _safe_print(f"Toplam eklenen kayıt: {eklenen}")
        _safe_print(f"Toplam hata sayısı: {len(hatalar)}")
        
        if eklenen == 0:
            error_detail = f"Hiçbir kayıt eklenemedi. Toplam {len(rows)} satır işlendi."
            if hatalar:
                error_detail += f" İlk 3 hata: {'; '.join(hatalar[:3])}"
            return JsonResponse({'success': False, 'error': error_detail}, status=400)
        
        # Kısmi başarı durumu
        result = {'success': True, 'count': eklenen}
        if hatalar:
            result['warnings'] = f"{len(hatalar)} satırda hata oluştu"
            result['errors'] = hatalar[:5]  # İlk 5 hatayı göster
        
        return JsonResponse(result)
    except Exception as e:
        _safe_print(f"Genel hata: {e}")
        import traceback
        import sys
        try:
            traceback.print_exc()
        except UnicodeEncodeError:
            sys.stderr.buffer.write(traceback.format_exc().encode('utf-8', errors='replace'))
        return JsonResponse({'success': False, 'error': f'Sunucu hatası: {str(e)}'}, status=500)

def health_check(request):
    """Health check endpoint"""
    return JsonResponse({'status': 'healthy', 'timestamp': datetime.now().isoformat()})

@login_required
def cikma_lastikler(request):
    """Çıkma Lastikler sayfası"""
    from .models import CikmaLastik
    from django.core.paginator import Paginator
    
    # Kullanıcı profili kontrolü
    try:
        user_profile = request.user.userprofile
        is_misafir = user_profile.is_misafir()
    except:
        is_misafir = False
    
    # Misafir kullanıcılar için POST işlemleri yasak
    if request.method == 'POST' and is_misafir:
        messages.error(request, 'Bu işlemi yapmaya yetkiniz yok.')
        return redirect('dashboard:cikma_lastikler')
    
    # Yeni kayıt ekleme - POST işlemi önce kontrol edilir
    if request.method == 'POST':
        print(f"DEBUG: POST request geldi, kullanıcı: {request.user.username}")
        print(f"DEBUG: POST verileri: {dict(request.POST)}")
        try:
            # Form verilerini al
            marka = request.POST.get('marka')
            model = request.POST.get('model', '')
            ebat = request.POST.get('ebat')
            mevsim = request.POST.get('mevsim')
            adet = int(request.POST.get('adet', 1))
            durum = request.POST.get('durum', 'depolandi')  # Default 'depolandi' olmalı
            kalite_notu = request.POST.get('kalite_notu', '')
            depo_konumu = request.POST.get('depo_konumu', '')
            aciklama = request.POST.get('aciklama', '')
            
            # Satış bilgileri - sadece durum "satildi" ise
            satis_fiyati = None
            satis_tarihi = None
            mehmet_havale = False
            canta = False
            cari = False
            
            if durum == 'satildi':
                satis_fiyati_str = request.POST.get('satis_fiyati', '')
                if satis_fiyati_str:
                    try:
                        satis_fiyati = Decimal(satis_fiyati_str.replace(',', '.'))
                    except:
                        pass
                
                satis_tarihi_str = request.POST.get('satis_tarihi', '')
                if satis_tarihi_str:
                    try:
                        from datetime import datetime
                        satis_tarihi = datetime.strptime(satis_tarihi_str, '%Y-%m-%d').date()
                    except:
                        satis_tarihi = timezone.now().date()
                else:
                    satis_tarihi = timezone.now().date()
                
                # Ödeme seçeneklerini al
                mehmet_havale = request.POST.get('mehmet_havale') == 'on'
                canta = request.POST.get('canta') == 'on'
                cari = request.POST.get('cari') == 'on'
            
            # Yeni kayıt oluştur
            yeni_kayit = CikmaLastik.objects.create(
                user=request.user,
                musteri_adi="Çıkma Lastik",  # Sabit müşteri adı
                musteri_telefon="",  # Boş
                musteri_plaka="",    # Boş
                marka=marka,
                model=model,
                ebat=ebat,
                mevsim=mevsim,
                arac_tipi='binek',   # Sabit araç tipi
                adet=adet,
                durum=durum,
                kalite_notu=kalite_notu,
                satis_fiyati=satis_fiyati,
                satis_tarihi=satis_tarihi,
                mehmet_havale=mehmet_havale,
                canta=canta,
                cari=cari,
                depo_konumu=depo_konumu,
                aciklama=aciklama
            )
            
            # Eğer durum "satıldı" ise ve ödeme seçenekleri varsa Transaction kaydı oluştur
            if durum == 'satildi' and satis_fiyati and (mehmet_havale or canta or cari):
                from .models import Transaction
                
                # Açıklama oluştur
                transaction_aciklama = f"Çıkma Lastik Satışı - {marka} {ebat} ({adet} adet)"
                if aciklama:
                    transaction_aciklama += f" - {aciklama}"
                
                # Ödeme seçeneklerine göre transaction oluştur
                if mehmet_havale:
                    Transaction.objects.create(
                        hareket_tipi='gelir',
                        tarih=satis_tarihi,
                        kasa_adi='servis',
                        mehmet_havale=satis_fiyati,
                        aciklama=transaction_aciklama,
                        created_by=request.user
                    )
                
                if canta:
                    Transaction.objects.create(
                        hareket_tipi='gelir',
                        tarih=satis_tarihi,
                        kasa_adi='servis',
                        nakit=satis_fiyati,
                        aciklama=transaction_aciklama + " (Çanta)",
                        created_by=request.user
                    )
            
            print(f"DEBUG: Yeni kayıt oluşturuldu - ID: {yeni_kayit.id}, Kullanıcı: {request.user.username}")
            
            messages.success(request, 'Çıkma lastik kaydı başarıyla eklendi!')
            # Filtreleri temizlemek için HttpResponseRedirect kullan
            return HttpResponseRedirect(reverse('dashboard:cikma_lastikler'))
            
        except Exception as e:
            print(f"DEBUG: Hata oluştu - {str(e)}")
            import traceback
            traceback.print_exc()
            messages.error(request, f'Kayıt eklenirken hata oluştu: {str(e)}')
            # Hata durumunda da sayfayı yeniden yükle
            return redirect('dashboard:cikma_lastikler')
    
    # Filtreleme parametreleri - sadece GET request'te
    marka = request.GET.get('marka', '')
    ebat = request.GET.get('ebat', '')
    mevsim = request.GET.get('mevsim', '')
    arac_tipi = request.GET.get('arac_tipi', '')
    depo_konumu = request.GET.get('depo_konumu', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Çıkma lastikleri getir
    # Misafir kullanıcılar tüm aktif verileri görür, diğer kullanıcılar sadece kendi kayıtlarını görür
    if is_misafir:
        # Misafir kullanıcılar için tüm aktif veriler (satılmayanlar)
        cikma_lastikler = CikmaLastik.objects.exclude(durum='satildi')
    else:
        # Diğer kullanıcılar için sadece kendi kayıtları (satılmayanlar)
        cikma_lastikler = CikmaLastik.objects.filter(user=request.user).exclude(durum='satildi')
    
    print(f"DEBUG: Kullanıcı {request.user.username} için toplam kayıt sayısı: {cikma_lastikler.count()}")
    if cikma_lastikler.exists():
        print(f"DEBUG: İlk 3 kayıt: {list(cikma_lastikler.values('id', 'marka', 'ebat', 'user__username')[:3])}")
    else:
        print("DEBUG: Hiç kayıt bulunamadı")
    
    # Filtreleme uygula
    if marka:
        # Türkçe karakter varyantları oluştur ve hepsiyle ara
        search_variants = create_turkish_search_variants(marka)
        q_objects = Q()
        for variant in search_variants:
            q_objects |= Q(marka__icontains=variant)
        cikma_lastikler = cikma_lastikler.filter(q_objects)
    if ebat:
        # Ebat formatını otomatik düzenle (2055516 -> 205/55R16)
        formatted_ebat = format_tire_size(ebat)
        cikma_lastikler = cikma_lastikler.filter(ebat__icontains=formatted_ebat)
    if mevsim:
        cikma_lastikler = cikma_lastikler.filter(mevsim=mevsim)
    if arac_tipi:
        cikma_lastikler = cikma_lastikler.filter(arac_tipi=arac_tipi)
    if depo_konumu:
        cikma_lastikler = cikma_lastikler.filter(depo_konumu__icontains=depo_konumu)
    
    # Tarih filtreleme uygula
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            cikma_lastikler = cikma_lastikler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            cikma_lastikler = cikma_lastikler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            cikma_lastikler = cikma_lastikler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bugun':
            cikma_lastikler = cikma_lastikler.filter(olusturma_tarihi__date=now.date())
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            cikma_lastikler = cikma_lastikler.filter(olusturma_tarihi__gte=start_date)
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1)
            cikma_lastikler = cikma_lastikler.filter(olusturma_tarihi__gte=start_date)
    
    # Özel tarih aralığı filtreleme
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date()
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d').date()
            cikma_lastikler = cikma_lastikler.filter(olusturma_tarihi__date__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date()
            cikma_lastikler = cikma_lastikler.filter(olusturma_tarihi__date__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d').date()
            cikma_lastikler = cikma_lastikler.filter(olusturma_tarihi__date__lte=end_date)
        except ValueError:
            pass
    
    # Sıralama - en yeni kayıtlar üstte
    cikma_lastikler = cikma_lastikler.order_by('-olusturma_tarihi')
    
    # Sayfalama
    paginator = Paginator(cikma_lastikler, 50)  # Sayfa başına 50 kayıt
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # İstatistikler (satılanlar hariç)
    toplam_kayit = cikma_lastikler.count()
    toplam_adet = cikma_lastikler.aggregate(total=Sum('adet'))['total'] or 0
    depolanan_adet = cikma_lastikler.filter(durum='depolandi').aggregate(total=Sum('adet'))['total'] or 0
    cikti_adet = cikma_lastikler.filter(durum='cikti').aggregate(total=Sum('adet'))['total'] or 0
    
    # Durum dağılımı
    durum_dagilimi = cikma_lastikler.values('durum').annotate(
        kayit_sayisi=Count('id'),
        adet_toplam=Sum('adet')
    ).order_by('durum')
    
    # Filtreleme parametrelerini URL query string olarak hazırla
    filter_params = []
    if marka:
        filter_params.append(f'marka={marka}')
    if ebat:
        filter_params.append(f'ebat={ebat}')
    if mevsim:
        filter_params.append(f'mevsim={mevsim}')
    if arac_tipi:
        filter_params.append(f'arac_tipi={arac_tipi}')
    if depo_konumu:
        filter_params.append(f'depo_konumu={depo_konumu}')
    if tarih_filtre:
        filter_params.append(f'tarih={tarih_filtre}')
    if baslangic_tarihi:
        filter_params.append(f'baslangic_tarihi={baslangic_tarihi}')
    if bitis_tarihi:
        filter_params.append(f'bitis_tarihi={bitis_tarihi}')
    
    filter_query_string = '&'.join(filter_params)
    
    context = {
        'page_title': 'Çıkma Lastikler',
        'cikma_lastikler': page_obj,
        'filters': {
            'marka': marka,
            'ebat': ebat,
            'mevsim': mevsim,
            'arac_tipi': arac_tipi,
            'depo_konumu': depo_konumu,
            'tarih': tarih_filtre,
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
        },
        'filter_query_string': filter_query_string,
        'stats': {
            'toplam_kayit': toplam_kayit,
            'toplam_adet': toplam_adet,
            'depolanan_adet': depolanan_adet,
            'cikti_adet': cikti_adet,
        },
        'durum_dagilimi': durum_dagilimi,
        'durum_choices': CikmaLastik.DURUM_CHOICES,
        'mevsim_choices': CikmaLastik.MEVSIM_CHOICES,
        'arac_tipi_choices': CikmaLastik.ARAC_TIPI_CHOICES,
        'kalite_choices': CikmaLastik.KALITE_CHOICES,
        'is_misafir': is_misafir,
    }
    return render(request, 'dashboard/cikma_lastikler.html', context)


@login_required
def cikma_lastik_duzenle(request, lastik_id):
    """Çıkma lastik düzenleme sayfası"""
    from .models import CikmaLastik
    
    # Kullanıcı profili kontrolü
    try:
        user_profile = request.user.userprofile
        if user_profile.is_misafir():
            messages.error(request, 'Bu işlemi yapmaya yetkiniz yok.')
            return redirect('dashboard:cikma_lastikler')
    except:
        pass
    
    # Kaydı getir (sadece kullanıcının kayıtları)
    lastik = get_object_or_404(CikmaLastik, id=lastik_id, user=request.user)
    
    # Eğer durum "cikti" ise otomatik olarak "depolandi" yap
    if lastik.durum == 'cikti':
        lastik.durum = 'depolandi'
    
    if request.method == 'POST':
        try:
            # Eski ödeme seçeneklerini sakla (Transaction oluşturmak için)
            eski_cari = lastik.cari
            eski_mehmet_havale = lastik.mehmet_havale
            eski_canta = lastik.canta
            eski_satis_fiyati = lastik.satis_fiyati
            eski_durum = lastik.durum
            
            # Form verilerini al ve güncelle
            lastik.marka = request.POST.get('marka', lastik.marka)
            lastik.model = request.POST.get('model', lastik.model)
            lastik.ebat = request.POST.get('ebat', lastik.ebat)
            lastik.mevsim = request.POST.get('mevsim', lastik.mevsim)
            lastik.adet = int(request.POST.get('adet', lastik.adet))
            
            # Durum sadece depolandi veya satildi olabilir
            durum = request.POST.get('durum', 'depolandi')
            if durum in ['depolandi', 'satildi']:
                lastik.durum = durum
            else:
                lastik.durum = 'depolandi'  # Varsayılan
            
            lastik.diş_derinligi = request.POST.get('dis_derinligi', lastik.diş_derinligi)
            lastik.kalite_notu = request.POST.get('kalite_notu', lastik.kalite_notu)
            lastik.hasar_durumu = request.POST.get('hasar_durumu', lastik.hasar_durumu)
            lastik.depo_konumu = request.POST.get('depo_konumu', lastik.depo_konumu)
            lastik.aciklama = request.POST.get('aciklama', lastik.aciklama)
            
            # Tahmini değeri güncelle
            tahmini_deger = request.POST.get('tahmini_deger', '')
            if tahmini_deger:
                try:
                    lastik.tahmini_deger = Decimal(tahmini_deger.replace(',', '.'))
                except:
                    pass
            
            # Satış fiyatı ve tarihi - sadece durum "satildi" ise
            if lastik.durum == 'satildi':
                satis_fiyati = request.POST.get('satis_fiyati', '')
                if satis_fiyati:
                    try:
                        lastik.satis_fiyati = Decimal(satis_fiyati.replace(',', '.'))
                    except:
                        pass
                
                satis_tarihi = request.POST.get('satis_tarihi', '')
                if satis_tarihi:
                    try:
                        from datetime import datetime
                        lastik.satis_tarihi = datetime.strptime(satis_tarihi, '%Y-%m-%d').date()
                    except:
                        lastik.satis_tarihi = timezone.now().date()
                else:
                    lastik.satis_tarihi = timezone.now().date()
                
                # Ödeme seçeneklerini güncelle
                lastik.mehmet_havale = request.POST.get('mehmet_havale') == 'on'
                lastik.canta = request.POST.get('canta') == 'on'
                lastik.cari = request.POST.get('cari') == 'on'
                
                # Eğer Cari kapatılıp Mehmet Havale veya Çanta seçildiyse Transaction oluştur
                if eski_durum == 'satildi' and eski_cari and not lastik.cari:
                    if (lastik.mehmet_havale or lastik.canta) and lastik.satis_fiyati:
                        # Transaction oluştur
                        from .models import Transaction
                        
                        # Açıklama oluştur
                        aciklama = f"Çıkma Lastik Satışı - {lastik.marka}"
                        if lastik.model:
                            aciklama += f" {lastik.model}"
                        aciklama += f" {lastik.ebat} ({lastik.adet} adet)"
                        if lastik.musteri_adi:
                            aciklama += f" - {lastik.musteri_adi}"
                        
                        # Transaction oluştur
                        transaction = Transaction(
                            created_by=request.user,
                            tarih=lastik.satis_tarihi or timezone.now().date(),
                            hareket_tipi='gelir',
                            kasa_adi='cikma-lastik',
                            aciklama=aciklama,
                            mehmet_havale=lastik.satis_fiyati if lastik.mehmet_havale else Decimal('0'),
                            canta_cikis=lastik.satis_fiyati if lastik.canta else Decimal('0'),
                            nakit=Decimal('0'),
                            kredi_karti=Decimal('0'),
                            cari=Decimal('0'),
                            sanal_pos=Decimal('0'),
                            banka_havale=Decimal('0'),
                        )
                        transaction.save()
                        messages.success(request, 'Ödeme kaydı oluşturuldu ve işlem sayfalarında görünecek!')
            else:
                # Depolandı durumunda satış bilgilerini temizle
                lastik.satis_fiyati = None
                lastik.satis_tarihi = None
                lastik.mehmet_havale = False
                lastik.canta = False
                lastik.cari = False
            
            # Çıkış tarihini güncelle
            cikis_tarihi = request.POST.get('cikis_tarihi', '')
            if cikis_tarihi:
                try:
                    lastik.cikis_tarihi = datetime.strptime(cikis_tarihi, '%Y-%m-%d').date()
                except:
                    pass
            
            lastik.save()
            
            messages.success(request, 'Çıkma lastik kaydı başarıyla güncellendi!')
            return HttpResponseRedirect(reverse('dashboard:cikma_lastikler'))
            
        except Exception as e:
            messages.error(request, f'Kayıt güncellenirken hata oluştu: {str(e)}')
    
    context = {
        'page_title': 'Çıkma Lastik Düzenle',
        'lastik': lastik,
        'mevsim_choices': CikmaLastik.MEVSIM_CHOICES,
        'kalite_choices': CikmaLastik.KALITE_CHOICES,
    }
    return render(request, 'dashboard/cikma_lastik_duzenle.html', context)


@login_required
def cikma_lastik_sil(request, lastik_id):
    """Çıkma lastik silme"""
    from .models import CikmaLastik
    
    # Kullanıcı profili kontrolü
    try:
        user_profile = request.user.userprofile
        if user_profile.is_misafir():
            messages.error(request, 'Bu işlemi yapmaya yetkiniz yok.')
            return redirect('dashboard:cikma_lastikler')
    except:
        pass
    
    # Kaydı getir (sadece kullanıcının kayıtları)
    lastik = get_object_or_404(CikmaLastik, id=lastik_id, user=request.user)
    
    # Silme işleminden sonra hangi sayfaya dönüleceğini belirle
    next_url = request.GET.get('next', 'dashboard:cikma_lastikler')
    if 'satilan' in request.META.get('HTTP_REFERER', ''):
        next_url = 'dashboard:satilan_cikma_lastikler'
    
    if request.method == 'POST':
        try:
            lastik_info = f"{lastik.marka} {lastik.ebat} ({lastik.adet} adet)"
            lastik.delete()
            messages.success(request, f'Çıkma lastik kaydı silindi: {lastik_info}')
        except Exception as e:
            messages.error(request, f'Kayıt silinirken hata oluştu: {str(e)}')
    else:
        # GET request ile silme işlemi güvenli değil
        messages.warning(request, 'Silme işlemi için POST metodu kullanılmalıdır.')
    
    return redirect(next_url)


@login_required
def cikma_lastik_sat(request, lastik_id):
    """Çıkma lastik satış işlemi"""
    from .models import CikmaLastik
    from datetime import datetime
    
    # Kaydı getir (sadece kullanıcının kayıtları)
    lastik = get_object_or_404(CikmaLastik, id=lastik_id, user=request.user)
    
    if request.method == 'POST':
        try:
            # Form verilerini al
            satis_adet = int(request.POST.get('satis_adet', lastik.adet))
            satis_fiyati = request.POST.get('satis_fiyati', '')
            satis_tarihi = request.POST.get('satis_tarihi', '')
            satis_aciklama = request.POST.get('satis_aciklama', '').strip()
            
            # Ödeme seçeneklerini al
            mehmet_havale = request.POST.get('mehmet_havale') == 'on'
            canta = request.POST.get('canta') == 'on'
            cari = request.POST.get('cari') == 'on'
            
            # Satış fiyatını decimal'e çevir
            if satis_fiyati:
                try:
                    satis_fiyati = Decimal(satis_fiyati.replace(',', '.'))
                except:
                    messages.error(request, 'Geçersiz satış fiyatı!')
                    return redirect('dashboard:cikma_lastikler')
            else:
                messages.error(request, 'Satış fiyatı gereklidir!')
                return redirect('dashboard:cikma_lastikler')
            
            # Satış tarihini date'e çevir
            if satis_tarihi:
                try:
                    satis_tarihi = datetime.strptime(satis_tarihi, '%Y-%m-%d').date()
                except:
                    satis_tarihi = timezone.now().date()
            else:
                satis_tarihi = timezone.now().date()
            
            # Adet kontrolü
            if satis_adet > lastik.adet:
                messages.error(request, f'Satılacak adet ({satis_adet}) mevcut adetten ({lastik.adet}) fazla olamaz!')
                return redirect('dashboard:cikma_lastikler')
            
            # Eğer tüm adet satılıyorsa, mevcut kaydı güncelle
            if satis_adet == lastik.adet:
                lastik.durum = 'satildi'
                lastik.satis_fiyati = satis_fiyati
                lastik.satis_tarihi = satis_tarihi
                lastik.mehmet_havale = mehmet_havale
                lastik.canta = canta
                lastik.cari = cari
                # Satış açıklamasını kaydet
                mevcut_aciklama = (lastik.aciklama or '').strip()
                if satis_aciklama:
                    if mevcut_aciklama:
                        lastik.aciklama = mevcut_aciklama + '\nSatış Notu: ' + satis_aciklama
                    else:
                        lastik.aciklama = 'Satış Notu: ' + satis_aciklama
                lastik.save()
                
                # Transaction kaydı oluştur
                from .models import Transaction
                
                # Açıklama oluştur
                transaction_aciklama = f"Çıkma Lastik Satışı - {lastik.marka} {lastik.ebat} ({satis_adet} adet)"
                if satis_aciklama:
                    transaction_aciklama += f" - {satis_aciklama}"
                
                # Ödeme seçeneklerine göre transaction oluştur
                if mehmet_havale:
                    Transaction.objects.create(
                        hareket_tipi='gelir',
                        tarih=satis_tarihi,
                        kasa_adi='servis',
                        mehmet_havale=satis_fiyati,
                        aciklama=transaction_aciklama,
                        created_by=request.user
                    )
                
                if canta:
                    Transaction.objects.create(
                        hareket_tipi='gelir',
                        tarih=satis_tarihi,
                        kasa_adi='servis',
                        nakit=satis_fiyati,
                        aciklama=transaction_aciklama + " (Çanta)",
                        created_by=request.user
                    )
                
                messages.success(request, f'{lastik.marka} {lastik.ebat} ({satis_adet} adet) başarıyla satıldı!')
            
            # Eğer kısmi satış yapılıyorsa, yeni kayıt oluştur ve mevcut kaydın adetini azalt
            else:
                # Satılan kısım için yeni kayıt oluştur
                mevcut_aciklama_kismi = (lastik.aciklama or '').strip()
                CikmaLastik.objects.create(
                    user=request.user,
                    musteri_adi=lastik.musteri_adi,
                    musteri_telefon=lastik.musteri_telefon,
                    musteri_plaka=lastik.musteri_plaka,
                    marka=lastik.marka,
                    model=lastik.model,
                    ebat=lastik.ebat,
                    mevsim=lastik.mevsim,
                    arac_tipi=lastik.arac_tipi,
                    adet=satis_adet,
                    durum='satildi',
                    kalite_notu=lastik.kalite_notu,
                    satis_fiyati=satis_fiyati,
                    satis_tarihi=satis_tarihi,
                    mehmet_havale=mehmet_havale,
                    canta=canta,
                    cari=cari,
                    depo_konumu=lastik.depo_konumu,
                    aciklama=(mevcut_aciklama_kismi + '\nSatış Notu: ' + satis_aciklama).strip() if satis_aciklama else mevcut_aciklama_kismi or None
                )
                
                # Transaction kaydı oluştur
                from .models import Transaction
                
                # Açıklama oluştur
                transaction_aciklama = f"Çıkma Lastik Satışı - {lastik.marka} {lastik.ebat} ({satis_adet} adet)"
                if satis_aciklama:
                    transaction_aciklama += f" - {satis_aciklama}"
                
                # Ödeme seçeneklerine göre transaction oluştur
                if mehmet_havale:
                    Transaction.objects.create(
                        hareket_tipi='gelir',
                        tarih=satis_tarihi,
                        kasa_adi='servis',
                        mehmet_havale=satis_fiyati,
                        aciklama=transaction_aciklama,
                        created_by=request.user
                    )
                
                if canta:
                    Transaction.objects.create(
                        hareket_tipi='gelir',
                        tarih=satis_tarihi,
                        kasa_adi='servis',
                        nakit=satis_fiyati,
                        aciklama=transaction_aciklama + " (Çanta)",
                        created_by=request.user
                    )
                
                # Mevcut kayıttan satılan adedi çıkar
                lastik.adet -= satis_adet
                lastik.save()
                
                messages.success(request, f'{lastik.marka} {lastik.ebat} ({satis_adet} adet) başarıyla satıldı! Kalan: {lastik.adet} adet')
            
            return redirect('dashboard:cikma_lastikler')
            
        except Exception as e:
            messages.error(request, f'Satış işlemi sırasında hata oluştu: {str(e)}')
    
    return redirect('dashboard:cikma_lastikler')


@login_required
def export_cikma_lastikler_excel(request):
    """Çıkma Lastikler listesini Excel'e aktar"""
    from .models import CikmaLastik

    # Kullanıcı profili kontrolü
    try:
        user_profile = request.user.userprofile
        is_misafir = user_profile.is_misafir()
    except Exception:
        is_misafir = False

    # Filtreleme parametrelerini al
    marka = request.GET.get('marka', '')
    ebat = request.GET.get('ebat', '')
    mevsim = request.GET.get('mevsim', '')
    arac_tipi = request.GET.get('arac_tipi', '')
    depo_konumu = request.GET.get('depo_konumu', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')

    # Queryset
    if is_misafir:
        lastikler = CikmaLastik.objects.exclude(durum='satildi')
    else:
        lastikler = CikmaLastik.objects.filter(user=request.user).exclude(durum='satildi')

    # Filtreler
    if marka:
        search_variants = create_turkish_search_variants(marka)
        q_objects = Q()
        for variant in search_variants:
            q_objects |= Q(marka__icontains=variant)
        lastikler = lastikler.filter(q_objects)
    if ebat:
        formatted_ebat = format_tire_size(ebat)
        lastikler = lastikler.filter(ebat__icontains=formatted_ebat)
    if mevsim:
        lastikler = lastikler.filter(mevsim=mevsim)
    if arac_tipi:
        lastikler = lastikler.filter(arac_tipi=arac_tipi)
    if depo_konumu:
        lastikler = lastikler.filter(depo_konumu__icontains=depo_konumu)

    # Tarih filtreleme
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            lastikler = lastikler.filter(olusturma_tarihi__gte=now - timedelta(days=30))
        elif tarih_filtre == 'son-3-ay':
            lastikler = lastikler.filter(olusturma_tarihi__gte=now - timedelta(days=90))
        elif tarih_filtre == 'son-6-ay':
            lastikler = lastikler.filter(olusturma_tarihi__gte=now - timedelta(days=180))
        elif tarih_filtre == 'bugun':
            lastikler = lastikler.filter(olusturma_tarihi__date=now.date())
        elif tarih_filtre == 'bu-hafta':
            lastikler = lastikler.filter(olusturma_tarihi__gte=now - timedelta(days=now.weekday()))
        elif tarih_filtre == 'bu-ay':
            lastikler = lastikler.filter(olusturma_tarihi__gte=now.replace(day=1))

    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date()
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d').date()
            lastikler = lastikler.filter(olusturma_tarihi__date__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date()
            lastikler = lastikler.filter(olusturma_tarihi__date__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d').date()
            lastikler = lastikler.filter(olusturma_tarihi__date__lte=end_date)
        except ValueError:
            pass

    lastikler = lastikler.order_by('-olusturma_tarihi')

    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Çıkma Lastikler"

    # Başlık stili
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        'EBAT', 'MARKA', 'MODEL', 'MEVSİM', 'ADET', 'AÇIKLAMA',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Veri satırları
    for row, lastik in enumerate(lastikler, 2):
        ws.cell(row=row, column=1, value=lastik.ebat)
        ws.cell(row=row, column=2, value=lastik.marka)
        ws.cell(row=row, column=3, value=lastik.model or '')
        ws.cell(row=row, column=4, value=lastik.get_mevsim_display())
        ws.cell(row=row, column=5, value=lastik.adet)
        ws.cell(row=row, column=6, value=lastik.aciklama or '')

    # Sütun genişlikleri
    column_widths = [14, 18, 18, 12, 7, 40]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Dosya adı
    filename = f"cikma_lastikler_{now.strftime('%Y%m%d_%H%M')}.xlsx"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def satilan_cikma_lastikler(request):
    """Satılan Çıkma Lastikler sayfası"""
    from .models import CikmaLastik
    
    # Kullanıcı profili kontrolü - Misafir kullanıcılar bu sayfayı göremez
    try:
        user_profile = request.user.userprofile
        if user_profile.is_misafir():
            messages.error(request, 'Bu sayfayı görüntüleme yetkiniz yok.')
            return redirect('dashboard:cikma_lastikler')
    except:
        pass
    from django.core.paginator import Paginator
    
    # Filtreleme parametreleri
    marka = request.GET.get('marka', '')
    ebat = request.GET.get('ebat', '')
    mevsim = request.GET.get('mevsim', '')
    tarih_filtre = request.GET.get('tarih', '')
    baslangic_tarihi = request.GET.get('baslangic_tarihi', '')
    bitis_tarihi = request.GET.get('bitis_tarihi', '')
    
    # Sadece satılan çıkma lastikleri getir (kullanıcının kayıtları)
    satilan_lastikler = CikmaLastik.objects.filter(user=request.user, durum='satildi')
    
    # Filtreleme uygula
    if marka:
        # Türkçe karakter varyantları oluştur ve hepsiyle ara
        search_variants = create_turkish_search_variants(marka)
        q_objects = Q()
        for variant in search_variants:
            q_objects |= Q(marka__icontains=variant)
        satilan_lastikler = satilan_lastikler.filter(q_objects)
    if ebat:
        satilan_lastikler = satilan_lastikler.filter(ebat__icontains=ebat)
    if mevsim:
        satilan_lastikler = satilan_lastikler.filter(mevsim=mevsim)
    
    # Tarih filtreleme uygula (satış tarihine göre)
    now = timezone.now()
    if tarih_filtre:
        if tarih_filtre == 'son-1-ay':
            start_date = now - timedelta(days=30)
            satilan_lastikler = satilan_lastikler.filter(satis_tarihi__gte=start_date.date())
        elif tarih_filtre == 'son-3-ay':
            start_date = now - timedelta(days=90)
            satilan_lastikler = satilan_lastikler.filter(satis_tarihi__gte=start_date.date())
        elif tarih_filtre == 'son-6-ay':
            start_date = now - timedelta(days=180)
            satilan_lastikler = satilan_lastikler.filter(satis_tarihi__gte=start_date.date())
        elif tarih_filtre == 'bugun':
            satilan_lastikler = satilan_lastikler.filter(satis_tarihi=now.date())
        elif tarih_filtre == 'bu-hafta':
            start_date = now - timedelta(days=now.weekday())
            satilan_lastikler = satilan_lastikler.filter(satis_tarihi__gte=start_date.date())
        elif tarih_filtre == 'bu-ay':
            start_date = now.replace(day=1)
            satilan_lastikler = satilan_lastikler.filter(satis_tarihi__gte=start_date.date())
    
    # Özel tarih aralığı filtreleme (satış tarihine göre)
    if baslangic_tarihi and bitis_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date()
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d').date()
            satilan_lastikler = satilan_lastikler.filter(satis_tarihi__range=[start_date, end_date])
        except ValueError:
            pass
    elif baslangic_tarihi:
        try:
            start_date = datetime.strptime(baslangic_tarihi, '%Y-%m-%d').date()
            satilan_lastikler = satilan_lastikler.filter(satis_tarihi__gte=start_date)
        except ValueError:
            pass
    elif bitis_tarihi:
        try:
            end_date = datetime.strptime(bitis_tarihi, '%Y-%m-%d').date()
            satilan_lastikler = satilan_lastikler.filter(satis_tarihi__lte=end_date)
        except ValueError:
            pass
    
    # Sıralama - önce cari olanlar (kırmızılar), sonra en yeni satışlar
    satilan_lastikler = satilan_lastikler.order_by('-cari', '-satis_tarihi', '-guncelleme_tarihi')
    
    # Sayfalama
    paginator = Paginator(satilan_lastikler, 50)  # Sayfa başına 50 kayıt
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # İstatistikler
    toplam_kayit = satilan_lastikler.count()
    toplam_adet = satilan_lastikler.aggregate(total=Sum('adet'))['total'] or 0
    
    # Sadece satış fiyatı olan kayıtlardan gelir hesapla
    toplam_gelir = satilan_lastikler.filter(satis_fiyati__isnull=False).aggregate(total=Sum('satis_fiyati'))['total'] or 0
    
    # Aylık satış verileri (son 6 ay) - Sadece satış tarihi olan kayıtlardan hesapla
    tum_satilan_lastikler = CikmaLastik.objects.filter(user=request.user, durum='satildi')
    aylik_satis = []
    aylik_labels = []
    for i in range(5, -1, -1):
        start_date = now - timedelta(days=30*i)
        end_date = start_date + timedelta(days=30)
        
        # Sadece satış tarihi olan kayıtlardan hesapla
        ay_satis = tum_satilan_lastikler.filter(
            satis_tarihi__isnull=False,
            satis_tarihi__gte=start_date.date(),
            satis_tarihi__lt=end_date.date()
        ).aggregate(total=Sum('adet'))['total'] or 0
        
        aylik_satis.append(ay_satis)
        aylik_labels.append(start_date.strftime('%b'))
    
    # Bu ay satış verisi (hızlı istatistikler için)
    bu_ay_baslangic = now.replace(day=1).date()
    bu_ay_satis = tum_satilan_lastikler.filter(
        satis_tarihi__gte=bu_ay_baslangic
    ).aggregate(total=Sum('adet'))['total'] or 0
    
    # Filtreleme parametrelerini URL query string olarak hazırla
    filter_params = []
    if marka:
        filter_params.append(f'marka={marka}')
    if ebat:
        filter_params.append(f'ebat={ebat}')
    if mevsim:
        filter_params.append(f'mevsim={mevsim}')
    if tarih_filtre:
        filter_params.append(f'tarih={tarih_filtre}')
    if baslangic_tarihi:
        filter_params.append(f'baslangic_tarihi={baslangic_tarihi}')
    if bitis_tarihi:
        filter_params.append(f'bitis_tarihi={bitis_tarihi}')
    
    filter_query_string = '&'.join(filter_params)
    
    context = {
        'page_title': 'Satılan Çıkma Lastikler',
        'satilan_lastikler': page_obj,
        'filters': {
            'marka': marka,
            'ebat': ebat,
            'mevsim': mevsim,
            'tarih': tarih_filtre,
            'baslangic_tarihi': baslangic_tarihi,
            'bitis_tarihi': bitis_tarihi,
        },
        'filter_query_string': filter_query_string,
        'stats': {
            'toplam_kayit': toplam_kayit,
            'toplam_adet': toplam_adet,
            'toplam_gelir': toplam_gelir,
        },
        'mevsim_choices': CikmaLastik.MEVSIM_CHOICES,
        'aylik_satis': json.dumps(aylik_satis),
        'aylik_labels': json.dumps(aylik_labels),
    }
    return render(request, 'dashboard/satilan_cikma_lastikler.html', context)


def get_filtered_transactions(user, **filters):
    """Filtrelenmiş işlemleri getir
    
    Args:
        user: Kullanıcı
        **filters: Filtreler (baslangic_tarih, bitis_tarih, hareket_tipi, kasa_adi, kategori_id)
    """
    # Temel sorgu oluştur
    islemler = Transaction.objects.filter(created_by=user).select_related(
        'kategori1', 'kategori1__parent', 'kategori2', 'kategori2__parent', 'kategori3', 'kategori3__parent'
    )
    
    # Temel filtreler
    # Merkez-satis kasasındaki gelir işlemlerini hariç tut, ANCAK:
    # - Nakit, Pafgo veya Mehmet Havale tutarı varsa dahil et
    # - Kullanıcı açıkça kasa_adi=merkez-satis filtresi seçmişse hariç tutma
    if filters.get('kasa_adi') != 'merkez-satis':
        islemler = islemler.exclude(
            Q(kasa_adi='merkez-satis') &
            Q(hareket_tipi='gelir') &
            Q(nakit=0) &
            Q(pafgo=0) &
            Q(mehmet_havale=0)
        )
    islemler = islemler.exclude(
        kasa_adi='virman'
    )
    
    # Tarih filtreleri
    if filters.get('baslangic_tarih'):
        try:
            baslangic = datetime.strptime(filters['baslangic_tarih'], '%Y-%m-%d').date()
            islemler = islemler.filter(tarih__gte=baslangic)
        except ValueError:
            pass
    
    if filters.get('bitis_tarih'):
        try:
            bitis = datetime.strptime(filters['bitis_tarih'], '%Y-%m-%d').date()
            islemler = islemler.filter(tarih__lte=bitis)
        except ValueError:
            pass
    
    # Diğer filtreler
    if filters.get('hareket_tipi'):
        islemler = islemler.filter(hareket_tipi=filters['hareket_tipi'])
    
    if filters.get('kasa_adi'):
        islemler = islemler.filter(kasa_adi=filters['kasa_adi'])
    
    if filters.get('kategori_id'):
        # Seçilen ana kategoriye ait alt kategorileri bul
        alt_kategoriler = TransactionCategory.objects.filter(
            parent_id=filters['kategori_id'],
            created_by=user
        ).values_list('id', flat=True)
        
        # Ana kategori veya alt kategorilerden biri ile eşleşenleri filtrele
        islemler = islemler.filter(
            Q(kategori1_id=filters['kategori_id']) |  # Doğrudan ana kategori
            Q(kategori1_id__in=alt_kategoriler)  # Alt kategorilerden biri
        )
    
    return islemler.order_by('-tarih', '-created_at')


def get_excel_hizmet_transactions(user, **filters):
    """Excel'den yüklenen hizmet kategorisindeki işlemleri Transaction formatında döndür"""
    # Excel'den hizmet kategorisindeki hareketleri al
    # Türkçe karakter sorunu için hem büyük hem küçük harf kontrol et
    excel_hareketler = MalzemeHareketi.objects.filter(
        kullanici=user
    ).filter(
        Q(kategori__icontains='hizmet') | 
        Q(kategori__icontains='HİZMET') |
        Q(kategori__iexact='hizmet') |
        Q(kategori__iexact='HİZMET')
    )
    
    # Tarih filtreleri uygula
    if filters.get('baslangic_tarih'):
        try:
            baslangic = datetime.strptime(filters['baslangic_tarih'], '%Y-%m-%d').date()
            excel_hareketler = excel_hareketler.filter(tarih__gte=baslangic)
        except ValueError:
            pass
    
    if filters.get('bitis_tarih'):
        try:
            bitis = datetime.strptime(filters['bitis_tarih'], '%Y-%m-%d').date()
            excel_hareketler = excel_hareketler.filter(tarih__lte=bitis)
        except ValueError:
            pass
    
    # Hareket tipi filtresi - Excel verileri sadece gelir olarak kabul edilir
    if filters.get('hareket_tipi') and filters['hareket_tipi'] != 'gelir':
        return []
    
    # Kasa adı filtresi - Excel hizmet verileri servis kasasına ait
    if filters.get('kasa_adi') and filters['kasa_adi'] != 'servis':
        return []
    
    # Excel verilerini Transaction benzeri objeler olarak döndür
    excel_transactions = []
    for hareket in excel_hareketler:
        # Ödeme şeklini Transaction alanlarına dönüştür
        nakit = kredi_karti = cari = sanal_pos = mehmet_havale = banka_havale = Decimal('0')
        
        odeme_sekli = (hareket.odeme_sekli or '').lower().strip()
        
        if 'nakit' in odeme_sekli:
            nakit = hareket.tutar
        elif any(x in odeme_sekli for x in ['kart', 'pos', 'kuveyttürk', 'kuveyt']) and 'sanal' not in odeme_sekli:
            kredi_karti = hareket.tutar
        elif any(x in odeme_sekli for x in ['cari', 'carı']) or 'gün cari' in odeme_sekli:
            cari = hareket.tutar
        elif 'sanal' in odeme_sekli:
            sanal_pos = hareket.tutar
        elif 'havale' in odeme_sekli:
            # Garanti/Banka Havale (M.Havale'den önce kontrol edilmeli)
            if any(x in odeme_sekli for x in ['garanti', 'garantı', 'GARANTİ', 'banka', 'b.havale', 'b havale', 'toplam havale']):
                banka_havale = hareket.tutar
            # M.HAVALE, M HAVALE, MEHMET HAVALE
            elif any(x in odeme_sekli for x in ['m.havale', 'm havale', 'mhavale', 'mehmet havale']):
                mehmet_havale = hareket.tutar
            else:
                banka_havale = hareket.tutar
        else:
            # Belirtilmemiş ödeme şekilleri nakit olarak kabul et
            nakit = hareket.tutar
        
        # Transaction benzeri obje oluştur
        class ExcelTransaction:
            def __init__(self, hareket, nakit, kredi_karti, cari, sanal_pos, mehmet_havale, banka_havale):
                self.id = f"excel_{hareket.id}"
                self.tarih = hareket.tarih
                self.hareket_tipi = 'gelir'
                self.kasa_adi = 'servis'
                self.nakit = nakit
                self.kredi_karti = kredi_karti
                self.cari = cari
                self.sanal_pos = sanal_pos
                self.mehmet_havale = mehmet_havale
                self.banka_havale = banka_havale
                self.pafgo = Decimal('0')
                self.aciklama = f"Excel: {hareket.musteri} - {hareket.urun}"
                self.kategori1 = None
                self.kategori2 = None
                self.kategori3 = None
                self.created_at = hareket.eklenme_zamani
                
            def get_kasa_adi_display(self):
                return 'Servis'
                
            def get_hareket_tipi_display(self):
                return 'Gelir'
                
            @property
            def toplam(self):
                return (self.nakit + self.kredi_karti + self.cari + 
                       self.sanal_pos + self.mehmet_havale + self.banka_havale + self.pafgo)
        
        excel_transactions.append(ExcelTransaction(
            hareket, nakit, kredi_karti, cari, sanal_pos, mehmet_havale, banka_havale
        ))
    
    return excel_transactions


def income_expense_report(request):
    """Gelir/Gider Raporu sayfası"""
    from datetime import date, timedelta
    
    # Filtreleme parametreleri
    today = date.today()
    
    # Varsayılan olarak bu ayın verilerini göster
    if not request.GET.get('baslangic_tarih') and not request.GET.get('bitis_tarih'):
        # Bu ayın ilk günü
        default_start = today.replace(day=1).strftime('%Y-%m-%d')
        # Bu ayın son günü
        if today.month == 12:
            next_month = today.replace(year=today.year + 1, month=1, day=1)
        else:
            next_month = today.replace(month=today.month + 1, day=1)
        default_end = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
    else:
        default_start = ''
        default_end = ''
    
    baslangic_tarih = request.GET.get('baslangic_tarih', default_start)
    bitis_tarih = request.GET.get('bitis_tarih', default_end)
    hareket_tipi = request.GET.get('hareket_tipi', '')
    kasa_adi = request.GET.get('kasa_adi', '')
    kategori_id = request.GET.get('kategori', '')
    
    # Filtrelenmiş işlemleri al (tüm filtreler get_filtered_transactions'da uygulanıyor)
    islemler = get_filtered_transactions(
        user=request.user,
        baslangic_tarih=baslangic_tarih,
        bitis_tarih=bitis_tarih,
        hareket_tipi=hareket_tipi,
        kasa_adi=kasa_adi,
        kategori_id=kategori_id
    )
    
    # Excel'den gelen hizmet işlemlerini de al
    excel_hizmet_islemleri = get_excel_hizmet_transactions(
        user=request.user,
        baslangic_tarih=baslangic_tarih,
        bitis_tarih=bitis_tarih,
        hareket_tipi=hareket_tipi,
        kasa_adi=kasa_adi,
        kategori_id=kategori_id
    )
    
    # İki listeyi birleştir
    tum_islemler = list(islemler) + excel_hizmet_islemleri
    
    # Tarihe göre sırala
    tum_islemler.sort(key=lambda x: (x.tarih, x.created_at), reverse=True)
    
    # Özet hesaplamaları ve detaylı işlem listeleri
    summary = {
        'nakit': 0,
        'kredi_karti': 0,  # Net tutar (toplam için)
        'kredi_karti_brut': 0,  # Brüt tutar (kart için)
        'cari': 0,  # Net tutar (toplam için)
        'cari_brut': 0,  # Brüt tutar (kart için)
        'sanal_pos': 0,  # Net tutar (toplam için)
        'sanal_pos_brut': 0,  # Brüt tutar (kart için)
        'mehmet_havale': 0,
        'banka_havale': 0,  # Net tutar (toplam için)
        'banka_havale_brut': 0,  # Brüt tutar (kart için)
        'pafgo': 0,
        'toplam': 0,
        'komisyon': 0  # %20 Devlet Payı/Gideri
    }
    
    # Her ödeme yöntemi için detaylı işlem listeleri
    odeme_detaylari = {
        'nakit': [],
        'kredi_karti': [],
        'cari': [],
        'sanal_pos': [],
        'mehmet_havale': [],
        'banka_havale': [],
        'pafgo': [],
        'toplam': []
    }
    
    def get_kategori_bilgisi(islem):
        """İşlemden ana ve alt kategori bilgilerini çıkar"""
        ana_kategori = ''
        alt_kategori = ''
        
        # Excel işlemleri için özel durum
        if hasattr(islem, 'id') and str(islem.id).startswith('excel_'):
            ana_kategori = 'Hizmet'
            alt_kategori = 'Excel Verisi'
        elif islem.kategori1:
            if islem.kategori1.parent:
                ana_kategori = islem.kategori1.parent.name
                alt_kategori = islem.kategori1.name
            else:
                ana_kategori = islem.kategori1.name
        
        return ana_kategori, alt_kategori
    
    for islem in tum_islemler:
        multiplier = 1 if islem.hareket_tipi == 'gelir' else -1
        ana_kategori, alt_kategori = get_kategori_bilgisi(islem)
        
        # Nakit işlemleri
        if islem.nakit and float(islem.nakit) > 0:
            amount = float(islem.nakit) * multiplier
            summary['nakit'] += amount
            odeme_detaylari['nakit'].append({
                'tarih': islem.tarih.strftime('%d.%m.%Y'),
                'kasa_adi': islem.get_kasa_adi_display() if islem.kasa_adi else '-',
                'hareket': islem.get_hareket_tipi_display(),
                'ana_kategori': ana_kategori,
                'alt_kategori': alt_kategori,
                'aciklama': islem.aciklama or '-',
                'amount': amount,
            })
        
        # Çanta Çıkış işlemleri - Nakit işlemleri modalında göster (sadece Transaction objelerinde)
        if hasattr(islem, 'canta_cikis') and islem.canta_cikis and float(islem.canta_cikis) > 0:
            amount = float(islem.canta_cikis) * multiplier
            summary['nakit'] += amount
            odeme_detaylari['nakit'].append({
                'tarih': islem.tarih.strftime('%d.%m.%Y'),
                'kasa_adi': islem.get_kasa_adi_display() if islem.kasa_adi else '-',
                'hareket': islem.get_hareket_tipi_display(),
                'ana_kategori': ana_kategori,
                'alt_kategori': alt_kategori,
                'aciklama': (islem.aciklama or '-') + ' [Çanta Çıkış]',
                'amount': amount,
            })
        
        # Kredi Kartı işlemleri - Gelir ise 1.20'ye böl (120 TL → 100 TL net, 20 TL komisyon)
        if islem.kredi_karti and float(islem.kredi_karti) > 0:
            brut_amount = float(islem.kredi_karti)
            komisyon = 0
            
            # Sadece gelir işlemlerinde komisyon hesapla
            if islem.hareket_tipi == 'gelir':
                net_amount = brut_amount / 1.20  # 120 ÷ 1.20 = 100
                komisyon = brut_amount - net_amount  # 120 - 100 = 20
                amount = net_amount * multiplier
                summary['komisyon'] += komisyon
                summary['kredi_karti_brut'] += brut_amount * multiplier  # Brüt tutar (kart için)
            else:
                amount = brut_amount * multiplier
                summary['kredi_karti_brut'] += brut_amount * multiplier
            
            summary['kredi_karti'] += amount  # Net tutar (toplam için)
            odeme_detaylari['kredi_karti'].append({
                'tarih': islem.tarih.strftime('%d.%m.%Y'),
                'kasa_adi': islem.get_kasa_adi_display() if islem.kasa_adi else '-',
                'hareket': islem.get_hareket_tipi_display(),
                'ana_kategori': ana_kategori,
                'alt_kategori': alt_kategori,
                'aciklama': islem.aciklama or '-',
                'amount': brut_amount * multiplier,  # Modalda brüt göster
                'komisyon': komisyon,
            })
        
        # Cari işlemleri - Gelir ise 1.20'ye böl (120 TL → 100 TL net, 20 TL komisyon)
        if islem.cari and float(islem.cari) > 0:
            brut_amount = float(islem.cari)
            komisyon = 0
            
            # Sadece gelir işlemlerinde komisyon hesapla
            if islem.hareket_tipi == 'gelir':
                net_amount = brut_amount / 1.20  # 120 ÷ 1.20 = 100
                komisyon = brut_amount - net_amount  # 120 - 100 = 20
                amount = net_amount * multiplier
                summary['komisyon'] += komisyon
                summary['cari_brut'] += brut_amount * multiplier  # Brüt tutar (kart için)
            else:
                amount = brut_amount * multiplier
                summary['cari_brut'] += brut_amount * multiplier
            
            summary['cari'] += amount  # Net tutar (toplam için)
            odeme_detaylari['cari'].append({
                'tarih': islem.tarih.strftime('%d.%m.%Y'),
                'kasa_adi': islem.get_kasa_adi_display() if islem.kasa_adi else '-',
                'hareket': islem.get_hareket_tipi_display(),
                'ana_kategori': ana_kategori,
                'alt_kategori': alt_kategori,
                'aciklama': islem.aciklama or '-',
                'amount': brut_amount * multiplier,  # Modalda brüt göster
                'komisyon': komisyon,
            })
        
        # Sanal Pos işlemleri - Gelir ise 1.20'ye böl (120 TL → 100 TL net, 20 TL komisyon)
        if islem.sanal_pos and float(islem.sanal_pos) > 0:
            brut_amount = float(islem.sanal_pos)
            komisyon = 0
            
            # Sadece gelir işlemlerinde komisyon hesapla
            if islem.hareket_tipi == 'gelir':
                net_amount = brut_amount / 1.20  # 120 ÷ 1.20 = 100
                komisyon = brut_amount - net_amount  # 120 - 100 = 20
                amount = net_amount * multiplier
                summary['komisyon'] += komisyon
                summary['sanal_pos_brut'] += brut_amount * multiplier  # Brüt tutar (kart için)
            else:
                amount = brut_amount * multiplier
                summary['sanal_pos_brut'] += brut_amount * multiplier
            
            summary['sanal_pos'] += amount  # Net tutar (toplam için)
            odeme_detaylari['sanal_pos'].append({
                'tarih': islem.tarih.strftime('%d.%m.%Y'),
                'kasa_adi': islem.get_kasa_adi_display() if islem.kasa_adi else '-',
                'hareket': islem.get_hareket_tipi_display(),
                'ana_kategori': ana_kategori,
                'alt_kategori': alt_kategori,
                'aciklama': islem.aciklama or '-',
                'amount': brut_amount * multiplier,  # Modalda brüt göster
                'komisyon': komisyon,
            })
        
        # Mehmet Havale işlemleri
        if islem.mehmet_havale and float(islem.mehmet_havale) > 0:
            amount = float(islem.mehmet_havale) * multiplier
            summary['mehmet_havale'] += amount
            odeme_detaylari['mehmet_havale'].append({
                'tarih': islem.tarih.strftime('%d.%m.%Y'),
                'kasa_adi': islem.get_kasa_adi_display() if islem.kasa_adi else '-',
                'hareket': islem.get_hareket_tipi_display(),
                'ana_kategori': ana_kategori,
                'alt_kategori': alt_kategori,
                'aciklama': islem.aciklama or '-',
                'amount': amount,
            })
        
        # Banka Havale işlemleri - Gelir ise 1.20'ye böl (120 TL → 100 TL net, 20 TL komisyon)
        if islem.banka_havale and float(islem.banka_havale) > 0:
            brut_amount = float(islem.banka_havale)
            komisyon = 0
            
            # Sadece gelir işlemlerinde komisyon hesapla
            if islem.hareket_tipi == 'gelir':
                net_amount = brut_amount / 1.20  # 120 ÷ 1.20 = 100
                komisyon = brut_amount - net_amount  # 120 - 100 = 20
                amount = net_amount * multiplier
                summary['komisyon'] += komisyon
                summary['banka_havale_brut'] += brut_amount * multiplier  # Brüt tutar (kart için)
            else:
                amount = brut_amount * multiplier
                summary['banka_havale_brut'] += brut_amount * multiplier
            
            summary['banka_havale'] += amount  # Net tutar (toplam için)
            odeme_detaylari['banka_havale'].append({
                'tarih': islem.tarih.strftime('%d.%m.%Y'),
                'kasa_adi': islem.get_kasa_adi_display() if islem.kasa_adi else '-',
                'hareket': islem.get_hareket_tipi_display(),
                'ana_kategori': ana_kategori,
                'alt_kategori': alt_kategori,
                'aciklama': islem.aciklama or '-',
                'amount': brut_amount * multiplier,  # Modalda brüt göster
                'komisyon': komisyon,
            })
        
        # Pafgo işlemleri
        if islem.pafgo and float(islem.pafgo) > 0:
            amount = float(islem.pafgo) * multiplier
            summary['pafgo'] += amount
            odeme_detaylari['pafgo'].append({
                'tarih': islem.tarih.strftime('%d.%m.%Y'),
                'kasa_adi': islem.get_kasa_adi_display() if islem.kasa_adi else '-',
                'hareket': islem.get_hareket_tipi_display(),
                'ana_kategori': ana_kategori,
                'alt_kategori': alt_kategori,
                'aciklama': islem.aciklama or '-',
                'amount': amount,
            })
        
        # Toplam için tüm işlemleri ekle
        total_amount = float(islem.toplam or 0) * multiplier
        
        odeme_detaylari['toplam'].append({
            'tarih': islem.tarih.strftime('%d.%m.%Y'),
            'kasa_adi': islem.get_kasa_adi_display() if islem.kasa_adi else '-',
            'hareket': islem.get_hareket_tipi_display(),
            'ana_kategori': ana_kategori,
            'alt_kategori': alt_kategori,
            'aciklama': islem.aciklama or '-',
            'amount': total_amount,
        })
    
    # Toplam hesapla (komisyon düşülmeden, sadece bilgi amaçlı gösterilir)
    summary['toplam'] = (
        summary['nakit'] + 
        summary['kredi_karti'] + 
        summary['cari'] + 
        summary['sanal_pos'] + 
        summary['mehmet_havale'] + 
        summary['banka_havale'] + 
        summary['pafgo']
    )
    
    # Debug: Summary değerlerini yazdır
    print(f"DEBUG: Summary değerleri:")
    print(f"  - Nakit: {summary['nakit']}")
    print(f"  - Kredi Kartı: {summary['kredi_karti']}")
    print(f"  - Cari: {summary['cari']}")
    print(f"  - Sanal Pos: {summary['sanal_pos']}")
    print(f"  - Mehmet Havale: {summary['mehmet_havale']}")
    print(f"  - Banka Havale: {summary['banka_havale']}")
    print(f"  - Pafgo: {summary['pafgo']}")
    print(f"  - Toplam: {summary['toplam']}")
    print(f"DEBUG: Pafgo işlem sayısı: {len(odeme_detaylari['pafgo'])}")
    print(f"DEBUG: Excel hizmet işlem sayısı: {len(excel_hizmet_islemleri)}")
    
    # Her ödeme yöntemi için komisyon toplamlarını hesapla
    komisyon_kredi_karti = sum(entry.get('komisyon', 0) for entry in odeme_detaylari['kredi_karti'])
    komisyon_cari = sum(entry.get('komisyon', 0) for entry in odeme_detaylari['cari'])
    komisyon_sanal_pos = sum(entry.get('komisyon', 0) for entry in odeme_detaylari['sanal_pos'])
    komisyon_banka_havale = sum(entry.get('komisyon', 0) for entry in odeme_detaylari['banka_havale'])
    
    # Debug: Pafgo işlemlerini kontrol et
    pafgo_islemler = Transaction.objects.filter(
        created_by=request.user,
        pafgo__gt=0
    ).values('id', 'tarih', 'pafgo', 'hareket_tipi')
    print(f"DEBUG: Toplam Pafgo işlem sayısı: {pafgo_islemler.count()}")
    for islem in pafgo_islemler:
        print(f"  - ID: {islem['id']}, Tarih: {islem['tarih']}, Pafgo: {islem['pafgo']}, Tip: {islem['hareket_tipi']}")
    
    # Ana kategorileri al (parent'ı olmayan)
    kategoriler = TransactionCategory.objects.filter(
        created_by=request.user,
        parent__isnull=True
    ).order_by('order', 'name')
    
    # Kasa seçeneklerini al
    kasa_choices = Transaction.KASA_CHOICES
    
    # Tarih aralığı etiketi
    if baslangic_tarih and bitis_tarih:
        try:
            baslangic_obj = datetime.strptime(baslangic_tarih, '%Y-%m-%d').date()
            bitis_obj = datetime.strptime(bitis_tarih, '%Y-%m-%d').date()
            date_range_label = f"{baslangic_obj.strftime('%d.%m.%Y')} - {bitis_obj.strftime('%d.%m.%Y')}"
        except:
            date_range_label = "Tüm zamanlar"
    elif baslangic_tarih:
        try:
            baslangic_obj = datetime.strptime(baslangic_tarih, '%Y-%m-%d').date()
            date_range_label = f"{baslangic_obj.strftime('%d.%m.%Y')} sonrası"
        except:
            date_range_label = "Tüm zamanlar"
    elif bitis_tarih:
        try:
            bitis_obj = datetime.strptime(bitis_tarih, '%Y-%m-%d').date()
            date_range_label = f"{bitis_obj.strftime('%d.%m.%Y')} öncesi"
        except:
            date_range_label = "Tüm zamanlar"
    else:
        date_range_label = "Tüm zamanlar"
    
    context = {
        'page_title': 'Gelir/Gider Raporu',
        'islemler': tum_islemler,
        'summary': summary,
        'kategoriler': kategoriler,
        'kasa_choices': kasa_choices,
        'filters': {
            'baslangic_tarih': baslangic_tarih,
            'bitis_tarih': bitis_tarih,
            'hareket_tipi': hareket_tipi,
            'kasa_adi': kasa_adi,
            'kategori': kategori_id,
        },
        'odeme_detaylari': odeme_detaylari,
        'date_range_label': date_range_label,
        'komisyon_kredi_karti': komisyon_kredi_karti,
        'komisyon_cari': komisyon_cari,
        'komisyon_sanal_pos': komisyon_sanal_pos,
        'komisyon_banka_havale': komisyon_banka_havale,
        'current_month_default': not request.GET.get('baslangic_tarih') and not request.GET.get('bitis_tarih'),  # Bu ay varsayılan mı?
    }
    
    return render(request, 'dashboard/income_expense_report.html', context)


@login_required
def export_income_expense_excel(request):
    """Gelir/Gider Raporu Excel'e aktar"""
    from datetime import date, timedelta
    
    # Filtreleme parametreleri
    today = date.today()
    
    # Varsayılan olarak bu ayın verilerini göster
    if not request.GET.get('baslangic_tarih') and not request.GET.get('bitis_tarih'):
        # Bu ayın ilk günü
        default_start = today.replace(day=1).strftime('%Y-%m-%d')
        # Bu ayın son günü
        if today.month == 12:
            next_month = today.replace(year=today.year + 1, month=1, day=1)
        else:
            next_month = today.replace(month=today.month + 1, day=1)
        default_end = (next_month - timedelta(days=1)).strftime('%Y-%m-%d')
    else:
        default_start = ''
        default_end = ''
    
    baslangic_tarih = request.GET.get('baslangic_tarih', default_start)
    bitis_tarih = request.GET.get('bitis_tarih', default_end)
    hareket_tipi = request.GET.get('hareket_tipi', '')
    kasa_adi = request.GET.get('kasa_adi', '')
    kategori_id = request.GET.get('kategori', '')
    
    # Filtrelenmiş işlemleri al (normal Transaction verileri)
    islemler = get_filtered_transactions(
        user=request.user,
        baslangic_tarih=baslangic_tarih,
        bitis_tarih=bitis_tarih,
        hareket_tipi=hareket_tipi,
        kasa_adi=kasa_adi,
        kategori_id=kategori_id
    )
    
    # Excel'den gelen hizmet işlemlerini de al
    excel_hizmet_islemleri = get_excel_hizmet_transactions(
        user=request.user,
        baslangic_tarih=baslangic_tarih,
        bitis_tarih=bitis_tarih,
        hareket_tipi=hareket_tipi,
        kasa_adi=kasa_adi,
        kategori_id=kategori_id
    )
    
    # İki listeyi birleştir
    tum_islemler = list(islemler) + excel_hizmet_islemleri
    
    # Tarihe göre sırala
    tum_islemler.sort(key=lambda x: (x.tarih, x.created_at), reverse=True)
    
    # Excel dosyası oluştur
    wb = Workbook()
    ws = wb.active
    ws.title = "Gelir Gider Raporu"
    
    # Başlık satırı
    headers = [
        'TARİH', 'KASA', 'ANA KATEGORİ', 'ALT KATEGORİ', 'NAKİT', 
        'KREDİ KARTI', 'CARİ', 'SANAL POS', 'M.HAVALE', 'B.HAVALE',
        'ÇANTA ÇIKIŞ', 'KOMİSYON (%20)', 'TOPLAM', 'AÇIKLAMA', 'İŞLEM TİPİ'
    ]
    
    # Başlık stilini ayarla
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Veri satırları
    for row, islem in enumerate(tum_islemler, 2):
        # Ana kategori
        ana_kategori = ''
        alt_kategori = ''
        
        # Excel işlemleri için özel durum
        if hasattr(islem, 'id') and str(islem.id).startswith('excel_'):
            ana_kategori = 'Hizmet'
            alt_kategori = 'Excel Verisi'
        else:
            if islem.kategori1 and islem.kategori1.parent:
                ana_kategori = islem.kategori1.parent.name
                alt_kategori = islem.kategori1.name
            elif islem.kategori1:
                ana_kategori = islem.kategori1.name
            elif islem.kategori2:
                alt_kategori = islem.kategori2.name
            elif islem.kategori3:
                alt_kategori = islem.kategori3.name
        
        # Tüm ödeme yöntemlerinin brüt değerlerini al
        kredi_karti_brut = float(islem.kredi_karti or 0)
        cari_brut = float(islem.cari or 0)
        sanal_pos_brut = float(islem.sanal_pos or 0)
        banka_havale_brut = float(islem.banka_havale or 0)
        
        # Komisyon hesapla - Sadece Gelir işlemlerinde 1.20'ye böl
        komisyon = 0
        if islem.hareket_tipi == 'gelir':
            # Her ödeme yöntemi için net değer hesapla
            kredi_karti_net = kredi_karti_brut / 1.20 if kredi_karti_brut > 0 else 0
            cari_net = cari_brut / 1.20 if cari_brut > 0 else 0
            sanal_pos_net = sanal_pos_brut / 1.20 if sanal_pos_brut > 0 else 0
            banka_havale_net = banka_havale_brut / 1.20 if banka_havale_brut > 0 else 0
            
            # Toplam komisyon
            komisyon = (
                (kredi_karti_brut - kredi_karti_net) + 
                (cari_brut - cari_net) + 
                (sanal_pos_brut - sanal_pos_net) + 
                (banka_havale_brut - banka_havale_net)
            )
            
            kredi_karti_value = kredi_karti_net
            cari_value = cari_net
            sanal_pos_value = sanal_pos_net
            banka_havale_value = banka_havale_net
        else:
            kredi_karti_value = kredi_karti_brut
            cari_value = cari_brut
            sanal_pos_value = sanal_pos_brut
            banka_havale_value = banka_havale_brut
        
        # Toplam hesapla
        toplam = (
            float(islem.nakit or 0) + 
            kredi_karti_value + 
            cari_value + 
            sanal_pos_value + 
            float(islem.mehmet_havale or 0) + 
            banka_havale_value +
            float(islem.canta_cikis or 0)
        )
        
        ws.cell(row=row, column=1, value=islem.tarih.strftime('%d.%m.%Y'))
        ws.cell(row=row, column=2, value=islem.get_kasa_adi_display())
        ws.cell(row=row, column=3, value=ana_kategori)
        ws.cell(row=row, column=4, value=alt_kategori)
        ws.cell(row=row, column=5, value=float(islem.nakit or 0))
        ws.cell(row=row, column=6, value=kredi_karti_value)
        ws.cell(row=row, column=7, value=cari_value)
        ws.cell(row=row, column=8, value=sanal_pos_value)
        ws.cell(row=row, column=9, value=float(islem.mehmet_havale or 0))
        ws.cell(row=row, column=10, value=banka_havale_value)
        ws.cell(row=row, column=11, value=float(islem.canta_cikis or 0))
        ws.cell(row=row, column=12, value=komisyon)
        ws.cell(row=row, column=13, value=toplam)
        ws.cell(row=row, column=14, value=islem.aciklama or '-')
        ws.cell(row=row, column=15, value=islem.get_hareket_tipi_display())
    
    # Sütun genişliklerini ayarla
    column_widths = [12, 15, 20, 20, 12, 12, 12, 12, 12, 12, 12, 15, 12, 30, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
    
    # HTTP response oluştur
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Dosya adını tarih aralığına göre oluştur
    if baslangic_tarih and bitis_tarih:
        filename = f"gelir_gider_raporu_{baslangic_tarih}_{bitis_tarih}.xlsx"
    elif baslangic_tarih:
        filename = f"gelir_gider_raporu_{baslangic_tarih}_sonrasi.xlsx"
    elif bitis_tarih:
        filename = f"gelir_gider_raporu_{bitis_tarih}_oncesi.xlsx"
    else:
        filename = f"gelir_gider_raporu_{today.strftime('%Y-%m-%d')}.xlsx"
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Excel dosyasını response'a yaz
    wb.save(response)
    return response






@login_required
@misafir_forbidden
def joker_satis(request):
    """Joker Satış Toplamları sayfası"""
    # Joker satış verilerini getir - admin tüm verileri, diğerleri sadece kendi verilerini görür
    if request.user.is_superuser:
        joker_hareketleri = JokerSatisHareketi.objects.all()
    else:
        joker_hareketleri = JokerSatisHareketi.objects.filter(kullanici=request.user)
    
    joker_hareketleri = joker_hareketleri.select_related('dosya')
    
    # Filtreleme parametreleri
    tarih_baslangic = request.GET.get('tarih_baslangic')
    tarih_bitis = request.GET.get('tarih_bitis')
    cari = request.GET.get('cari')
    kategori = request.GET.get('kategori')
    marka = request.GET.get('marka')
    urun_kodu = request.GET.get('urun_kodu')
    urun = request.GET.get('urun')
    
    # Filtreleri uygula
    if tarih_baslangic:
        joker_hareketleri = joker_hareketleri.filter(tarih__gte=tarih_baslangic)
    if tarih_bitis:
        joker_hareketleri = joker_hareketleri.filter(tarih__lte=tarih_bitis)
    if cari:
        joker_hareketleri = joker_hareketleri.filter(cari__icontains=cari)
    if kategori:
        joker_hareketleri = joker_hareketleri.filter(kategori__icontains=kategori)
    if marka:
        joker_hareketleri = joker_hareketleri.filter(marka__icontains=marka)
    if urun_kodu:
        joker_hareketleri = joker_hareketleri.filter(urun_kodu__icontains=urun_kodu)
    if urun:
        joker_hareketleri = joker_hareketleri.filter(urun__icontains=urun)
    
    # Sıralama
    joker_hareketleri = joker_hareketleri.order_by('-tarih', '-eklenme_zamani')
    
    # Toplam hesaplamalar - sadece kar
    toplam_kar = joker_hareketleri.aggregate(
        toplam=Sum('kar_tutari')
    )['toplam'] or 0
    
    # Aylık toplamlar - sadece kar
    aylik_toplamlar = joker_hareketleri.annotate(
        ay=TruncMonth('tarih')
    ).values('ay').annotate(
        kar_toplam=Sum('kar_tutari')
    ).order_by('-ay')[:12]
    
    # Dosya bazında toplamlar - sadece kar
    if request.user.is_superuser:
        dosya_toplamlar = JokerSatisDosya.objects.all()
    else:
        dosya_toplamlar = JokerSatisDosya.objects.filter(
            satirlar__kullanici=request.user
        ).distinct()
    
    dosya_toplamlar = dosya_toplamlar.annotate(
        toplam_kar=Sum('satirlar__kar_tutari'),
        satir_sayisi=Count('satirlar')
    ).order_by('-yukleme_tarihi')
    
    context = {
        'page_title': 'Joker Satış Toplamları',
        'joker_hareketleri': joker_hareketleri[:50],  # Son 50 kayıt
        'toplam_kar': toplam_kar,
        'aylik_toplamlar': aylik_toplamlar,
        'dosya_toplamlar': dosya_toplamlar,
    }
    
    return render(request, 'dashboard/joker_satis.html', context)

@login_required
@misafir_forbidden
def joker_satis_excel_upload(request):
    """Joker Satış Excel dosyası yükleme"""
    if request.method == 'POST':
        form = MalzemeExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            excel_file = request.FILES.get('file')
            
            if not excel_file:
                messages.error(request, "Lütfen bir dosya seçin.")
                return redirect('dashboard:joker_satis_excel_upload')
            
            # Dosya uzantısını kontrol et
            if not (excel_file.name.endswith('.xlsx') or excel_file.name.endswith('.xls')):
                messages.error(request, "Sadece .xlsx veya .xls dosyaları yükleyebilirsiniz.")
                return redirect('dashboard:joker_satis_excel_upload')
            
            try:
                from openpyxl import load_workbook
                wb = load_workbook(excel_file)
                ws = wb.active
                
                # JokerSatisDosya oluştur
                dosya = JokerSatisDosya.objects.create(
                    dosya_adi=excel_file.name,
                    kullanici=request.user,
                )
                
                # İlk satırdan başlıkları al
                headers = []
                for cell in ws[1]:
                    if cell.value:
                        headers.append(str(cell.value).strip())
                
                if not headers:
                    messages.error(request, "Excel dosyasında başlık satırı bulunamadı.")
                    dosya.delete()
                    return redirect('dashboard:joker_satis_excel_upload')
                
                eklenen = 0
                toplam_kar = 0
                for row_num in range(2, ws.max_row + 1):
                    try:
                        # Satır verilerini oluştur
                        row_data = {}
                        for col_num, header in enumerate(headers, 1):
                            if col_num <= len(headers):
                                cell_value = ws.cell(row=row_num, column=col_num).value
                                row_data[header] = cell_value
                        
                        print(f"Satır {row_num} verisi: {row_data}")  # Debug
                        
                        # Tarih işleme
                        tarih = row_data.get('TARİH') or row_data.get('TARIH') or ''
                        if isinstance(tarih, datetime):
                            tarih = tarih.date()
                        elif isinstance(tarih, str) and tarih and '.' in tarih:
                            try:
                                tarih = datetime.strptime(tarih, '%d.%m.%Y').date()
                            except:
                                tarih = date.today()
                        else:
                            tarih = date.today()

                        # Yeni alanları parse et - daha esnek eşleştirme
                        miktar = 1
                        for key in row_data.keys():
                            if any(x in key.upper() for x in ['MİKTAR', 'MIKTAR', 'QUANTITY', 'QTY']):
                                miktar = parse_decimal_value(row_data[key] or '1')
                                break
                        
                        alis_fiyati = 0
                        for key in row_data.keys():
                            if any(x in key.upper() for x in ['ALIŞ', 'ALIS', 'PURCHASE', 'BUY']):
                                alis_fiyati = parse_decimal_value(row_data[key] or '0')
                                break
                        
                        satis_fiyati = 0
                        for key in row_data.keys():
                            if any(x in key.upper() for x in ['SATIŞ', 'SATIS', 'SALE', 'SELL']):
                                satis_fiyati = parse_decimal_value(row_data[key] or '0')
                                break
                        
                        print(f"Parse edilen değerler - Miktar: {miktar}, Alış: {alis_fiyati}, Satış: {satis_fiyati}")  # Debug
                        
                        # Yeni kar hesaplama: (Satış × 1.20 - Alış) × Miktar
                        kar_per_unit = (satis_fiyati * Decimal('1.20')) - alis_fiyati
                        kar_tutari = kar_per_unit * miktar
                        # Cari alanı - daha spesifik arama
                        cari = ''
                        for key in row_data.keys():
                            key_clean = key.upper().strip()
                            if key_clean in ['CARİ', 'CARI'] or key_clean == 'MÜŞTERİ' or key_clean == 'MÜŞTERI':
                                cari = str(row_data[key] or '')[:255]
                                print(f"Cari bulundu - Sütun: '{key}', Değer: '{cari}'")  # Debug
                                break
                        
                        # Kategori alanı
                        kategori = ''
                        for key in row_data.keys():
                            if any(x in key.upper() for x in ['KATEGORİ', 'KATEGORI', 'CATEGORY', 'CAT']):
                                kategori = str(row_data[key] or '')[:255]
                                break
                        
                        # Marka alanı
                        marka = ''
                        for key in row_data.keys():
                            if any(x in key.upper() for x in ['MARKA', 'BRAND', 'MARK']):
                                marka = str(row_data[key] or '')[:255]
                                break
                        
                        # Ürün kodu alanı
                        urun_kodu = ''
                        for key in row_data.keys():
                            if any(x in key.upper() for x in ['ÜRÜN KODU', 'URUN_KODU', 'PRODUCT_CODE', 'CODE', 'KOD']):
                                urun_kodu = str(row_data[key] or '')[:255]
                                break
                        
                        # Ürün alanı - daha spesifik arama
                        urun = ''
                        for key in row_data.keys():
                            key_clean = key.upper().strip()
                            if key_clean in ['ÜRÜN', 'URUN'] or key_clean == 'PRODUCT':
                                urun = str(row_data[key] or '')[:255]
                                print(f"Ürün bulundu - Sütun: '{key}', Değer: '{urun}'")  # Debug
                                break
                        
                        # KZ alanı - daha geniş arama kriterleri
                        kz = ''
                        for key in row_data.keys():
                            key_upper = key.upper().strip()
                            if any(x in key_upper for x in ['KZ', 'K.Z', 'K Z', 'K-Z']):
                                kz = str(row_data[key] or '')[:100]
                                print(f"KZ bulundu - Sütun: '{key}', Değer: '{kz}'")  # Debug
                                break
                        
                        # Fatura No alanı
                        faturano = ''
                        for key in row_data.keys():
                            if any(x in key.upper() for x in ['FATURA', 'INVOICE', 'NO']):
                                faturano = str(row_data[key] or '')[:100]
                                break
                        
                        print(f"Satır {row_num} - Cari: '{cari}', Kategori: '{kategori}', Marka: '{marka}', Ürün Kodu: '{urun_kodu}', Ürün: '{urun}'")  # Debug
                            
                        hareket = JokerSatisHareketi(
                            dosya=dosya,
                            tarih=tarih,
                            cari=cari,
                            kategori=kategori,
                            marka=marka,
                            urun_kodu=urun_kodu,
                            urun=urun,
                            miktar=miktar,
                            alis_fiyati=alis_fiyati,
                            satis_fiyati=satis_fiyati,
                            kar_tutari=kar_tutari,
                            kullanici=request.user,
                        )
                        hareket.save()
                        eklenen += 1
                        toplam_kar += kar_tutari
                    except Exception as e:
                        print(f"Row error: {e}")
                        continue
                
                if eklenen == 0:
                    messages.warning(request, "Excel dosyasından hiç veri eklenemedi. Lütfen dosya formatını kontrol edin.")
                    dosya.delete()
                else:
                    messages.success(request, f"Başarıyla {eklenen} joker satış kaydı eklendi! Toplam kar: {toplam_kar:.2f} TL")
                return redirect('dashboard:joker_satis')
            except Exception as e:
                print(f"Excel upload error: {e}")
                import traceback
                traceback.print_exc()
                messages.error(request, f"Excel yükleme hatası: {str(e)}")
                return redirect('dashboard:joker_satis_excel_upload')
        else:
            messages.error(request, "Form geçersiz. Lütfen dosya seçin.")
            print(f"Form errors: {form.errors}")
    else:
        form = MalzemeExcelUploadForm()
    
    context = {
        'page_title': 'Joker Satış Excel Yükle',
        'form': form,
    }
    return render(request, 'dashboard/joker_satis_excel_upload.html', context)
@login_required
def debug_excel_headers(request):
    """Excel dosyasının başlıklarını kontrol etmek için debug view"""
    if request.method == 'POST' and request.FILES.get('file'):
        excel_file = request.FILES['file']
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_file)
            ws = wb.active
            
            # İlk satırdan başlıkları al
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value).strip())
            
            # İlk birkaç satırın verilerini al
            sample_data = []
            for row_num in range(2, min(5, ws.max_row + 1)):
                row_data = {}
                for col_num, header in enumerate(headers, 1):
                    if col_num <= len(headers):
                        cell_value = ws.cell(row=row_num, column=col_num).value
                        row_data[header] = cell_value
                sample_data.append(row_data)
            
            # KZ sütunu kontrolü
            kz_columns = []
            for header in headers:
                if any(x in header.upper() for x in ['KZ', 'K.Z', 'K Z', 'K-Z', 'KOD', 'CODE']):
                    kz_columns.append(header)
            
            return JsonResponse({
                'success': True,
                'headers': headers,
                'sample_data': sample_data,
                'kz_columns_found': kz_columns,
                'total_rows': ws.max_row - 1
            })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({'success': False, 'error': 'Dosya yüklenemedi'})

@login_required
def lastik_karsilastirma(request):
    """Lastik Karşılaştırma sayfası"""
    context = {
        'page_title': 'Lastik Karşılaştırma',
    }
    return render(request, 'dashboard/lastik_karsilastirma.html', context)

@login_required
def get_oe_tire_data(request):
    """Wheel-Size API'den OE lastik verilerini getir"""
    from django.http import JsonResponse
    import json
    import re
    
    if request.method != 'GET':
        return JsonResponse({'error': 'Sadece GET metodu desteklenir'}, status=405)
    
    make = request.GET.get('make', '').strip().lower()
    model = request.GET.get('model', '').strip().lower()
    year = request.GET.get('year', '').strip()
    generation = request.GET.get('generation', '').strip()
    
    if not make or not model or not year:
        return JsonResponse({'error': 'Marka, model ve yıl bilgileri gereklidir'}, status=400)
    
    # Kapsamlı mock data - gerçek Wheel-Size verilerine dayalı
    mock_data = {
        # BMW
        'bmw': {
            '1 series': {'2020': '205/55R16', '2019': '205/55R16', '2021': '205/55R16', '2018': '205/60R16'},
            '2 series': {'2020': '205/60R16', '2019': '205/60R16', '2021': '225/45R17', '2018': '205/60R16'},
            '3 series': {'2020': '225/45R17', '2019': '225/50R17', '2021': '225/45R18', '2018': '225/50R17', '2017': '225/50R17', '2016': '225/55R16'},
            '4 series': {'2020': '225/50R17', '2019': '225/50R17', '2021': '225/45R18', '2018': '225/50R17'},
            '5 series': {'2020': '245/45R18', '2019': '245/50R18', '2021': '245/45R19', '2018': '245/50R18', '2017': '245/50R18'},
            '6 series': {'2020': '245/45R18', '2019': '245/45R18', '2018': '245/40R19'},
            '7 series': {'2020': '245/50R18', '2019': '245/50R18', '2021': '245/45R19', '2018': '245/50R18'},
            'x1': {'2020': '225/50R17', '2019': '225/55R17', '2021': '225/50R18', '2018': '225/55R17'},
            'x3': {'2020': '245/50R18', '2019': '245/50R18', '2021': '245/45R19', '2018': '245/50R18'},
            'x5': {'2020': '255/50R19', '2019': '255/55R18', '2021': '275/45R20', '2018': '255/55R18'},
        },
        # Mercedes-Benz
        'mercedes': {
            'a-class': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'b-class': {'2020': '205/55R16', '2019': '205/55R16', '2018': '205/55R16'},
            'c-class': {'2020': '225/50R17', '2019': '225/55R16', '2021': '225/45R18', '2018': '225/55R16', '2017': '225/55R16'},
            'e-class': {'2020': '245/45R18', '2019': '245/45R18', '2021': '245/40R19', '2018': '245/45R18', '2017': '245/45R18'},
            's-class': {'2020': '245/50R18', '2019': '245/50R18', '2021': '275/40R20', '2018': '245/50R18'},
            'cla': {'2020': '225/45R17', '2019': '225/45R17', '2021': '225/40R18', '2018': '225/45R17'},
            'glc': {'2020': '235/55R18', '2019': '235/60R17', '2021': '255/45R20', '2018': '235/60R17'},
            'gle': {'2020': '255/50R19', '2019': '265/50R19', '2021': '275/45R21', '2018': '265/50R19'},
        },
        # Audi
        'audi': {
            'a1': {'2020': '195/55R16', '2019': '195/55R16', '2021': '215/45R17', '2018': '195/55R16'},
            'a3': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16', '2017': '205/55R16'},
            'a4': {'2020': '225/50R17', '2019': '225/55R16', '2021': '245/40R18', '2018': '225/55R16', '2017': '225/55R16'},
            'a5': {'2020': '225/50R17', '2019': '225/50R17', '2021': '245/40R18', '2018': '225/50R17'},
            'a6': {'2020': '245/45R18', '2019': '245/45R18', '2021': '255/40R19', '2018': '245/45R18', '2017': '245/45R18'},
            'a7': {'2020': '255/40R19', '2019': '255/40R19', '2021': '275/35R21', '2018': '255/40R19'},
            'a8': {'2020': '255/45R18', '2019': '255/45R18', '2021': '275/40R20', '2018': '255/45R18'},
            'q3': {'2020': '215/65R16', '2019': '215/65R16', '2021': '235/55R17', '2018': '215/65R16'},
            'q5': {'2020': '235/60R17', '2019': '235/60R17', '2021': '255/45R19', '2018': '235/60R17'},
            'q7': {'2020': '255/55R18', '2019': '255/55R18', '2021': '285/45R20', '2018': '255/55R18'},
        },
        # Volkswagen
        'volkswagen': {
            'polo': {'2020': '185/65R15', '2019': '185/65R15', '2021': '195/55R16', '2018': '185/65R15'},
            'golf': {'2020': '205/55R16', '2019': '205/60R15', '2021': '225/45R17', '2018': '205/60R15', '2017': '205/55R16'},
            'jetta': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'passat': {'2020': '215/60R16', '2019': '215/60R16', '2021': '235/45R17', '2018': '215/60R16', '2017': '215/60R16'},
            'arteon': {'2020': '235/45R17', '2019': '235/45R17', '2021': '245/40R18', '2018': '235/45R17'},
            'tiguan': {'2020': '215/65R16', '2019': '215/65R16', '2021': '235/55R17', '2018': '215/65R16'},
            'touareg': {'2020': '255/55R18', '2019': '255/55R18', '2021': '285/45R20', '2018': '255/55R18'},
        },
        # Ford
        'ford': {
            'fiesta': {'2020': '185/60R15', '2019': '185/60R15', '2021': '195/55R16', '2018': '185/60R15'},
            'focus': {'2020': '205/55R16', '2019': '195/65R15', '2021': '215/50R17', '2018': '195/65R15', '2017': '205/55R16'},
            'mondeo': {'2020': '215/55R17', '2019': '215/55R17', '2021': '235/45R18', '2018': '215/55R17', '2017': '215/55R17'},
            'mustang': {'2020': '235/55R17', '2019': '235/55R17', '2021': '255/40R19', '2018': '235/55R17'},
            'kuga': {'2020': '235/60R16', '2019': '235/60R16', '2021': '245/50R18', '2018': '235/60R16'},
            'edge': {'2020': '245/60R18', '2019': '245/60R18', '2018': '245/60R18'},
        },
        # Toyota
        'toyota': {
            'yaris': {'2020': '185/60R15', '2019': '185/60R15', '2021': '195/50R16', '2018': '185/60R15'},
            'corolla': {'2020': '205/55R16', '2019': '195/65R15', '2021': '215/45R17', '2018': '195/65R15', '2017': '205/55R16'},
            'camry': {'2020': '215/55R17', '2019': '215/55R17', '2021': '235/45R18', '2018': '215/55R17', '2017': '215/55R17'},
            'avalon': {'2020': '215/55R17', '2019': '215/55R17', '2021': '235/45R18', '2018': '215/55R17'},
            'prius': {'2020': '215/45R17', '2019': '215/45R17', '2021': '215/45R17', '2018': '215/45R17'},
            'rav4': {'2020': '225/65R17', '2019': '225/65R17', '2021': '235/55R19', '2018': '225/65R17'},
            'highlander': {'2020': '245/60R18', '2019': '245/60R18', '2021': '245/55R19', '2018': '245/60R18'},
        },
        # Honda
        'honda': {
            'civic': {'2020': '215/55R16', '2019': '205/55R16', '2021': '235/40R18', '2018': '205/55R16', '2017': '215/55R16'},
            'accord': {'2020': '225/50R17', '2019': '225/50R17', '2021': '235/45R18', '2018': '225/50R17', '2017': '225/50R17'},
            'cr-v': {'2020': '235/60R17', '2019': '235/65R16', '2021': '235/55R18', '2018': '235/65R16', '2017': '235/60R17'},
            'pilot': {'2020': '245/50R20', '2019': '245/50R20', '2021': '265/45R20', '2018': '245/50R20'},
            'hr-v': {'2020': '215/60R16', '2019': '215/60R16', '2021': '215/55R17', '2018': '215/60R16'},
        },
        # Nissan
        'nissan': {
            'micra': {'2020': '185/65R15', '2019': '185/65R15', '2021': '195/55R16', '2018': '185/65R15'},
            'sentra': {'2020': '205/55R16', '2019': '205/55R16', '2021': '215/50R17', '2018': '205/55R16'},
            'altima': {'2020': '215/60R16', '2019': '215/60R16', '2021': '235/40R19', '2018': '215/60R16'},
            'maxima': {'2020': '245/45R18', '2019': '245/45R18', '2021': '245/40R19', '2018': '245/45R18'},
            'qashqai': {'2020': '215/60R17', '2019': '215/65R16', '2021': '215/55R18', '2018': '215/65R16'},
            'x-trail': {'2020': '225/65R17', '2019': '225/65R17', '2021': '235/55R18', '2018': '225/65R17'},
        },
        # Hyundai
        'hyundai': {
            'i10': {'2020': '165/70R14', '2019': '165/70R14', '2021': '185/55R15', '2018': '165/70R14'},
            'i20': {'2020': '185/65R15', '2019': '185/65R15', '2021': '195/55R16', '2018': '185/65R15'},
            'i30': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'elantra': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'sonata': {'2020': '215/60R16', '2019': '215/60R16', '2021': '235/45R18', '2018': '215/60R16'},
            'tucson': {'2020': '225/60R17', '2019': '225/60R17', '2021': '235/55R18', '2018': '225/60R17'},
            'santa fe': {'2020': '235/60R18', '2019': '235/65R17', '2021': '245/50R20', '2018': '235/65R17'},
        },
        # Kia
        'kia': {
            'picanto': {'2020': '165/70R14', '2019': '165/70R14', '2021': '175/60R15', '2018': '165/70R14'},
            'rio': {'2020': '185/65R15', '2019': '185/65R15', '2021': '195/55R16', '2018': '185/65R15'},
            'ceed': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'cerato': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'optima': {'2020': '215/60R16', '2019': '215/60R16', '2021': '235/45R18', '2018': '215/60R16'},
            'sportage': {'2020': '225/60R17', '2019': '225/60R17', '2021': '235/55R18', '2018': '225/60R17'},
            'sorento': {'2020': '235/60R18', '2019': '235/65R17', '2021': '245/50R20', '2018': '235/65R17'},
        },
        # Peugeot
        'peugeot': {
            '108': {'2020': '165/65R15', '2019': '165/65R15', '2018': '165/65R15'},
            '208': {'2020': '185/65R15', '2019': '185/65R15', '2021': '195/55R16', '2018': '185/65R15'},
            '308': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            '508': {'2020': '215/60R16', '2019': '215/60R16', '2021': '235/45R18', '2018': '215/60R16'},
            '2008': {'2020': '205/60R16', '2019': '205/60R16', '2021': '215/55R17', '2018': '205/60R16'},
            '3008': {'2020': '215/65R16', '2019': '215/65R16', '2021': '235/50R18', '2018': '215/65R16'},
            '5008': {'2020': '215/65R16', '2019': '215/65R16', '2021': '235/50R18', '2018': '215/65R16'},
        },
        # Renault
        'renault': {
            'clio': {'2020': '185/65R15', '2019': '185/65R15', '2021': '195/55R16', '2018': '185/65R15'},
            'megane': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'talisman': {'2020': '215/60R16', '2019': '215/60R16', '2021': '235/45R18', '2018': '215/60R16'},
            'captur': {'2020': '205/60R16', '2019': '205/60R16', '2021': '215/55R17', '2018': '205/60R16'},
            'kadjar': {'2020': '215/65R16', '2019': '215/65R16', '2021': '235/55R17', '2018': '215/65R16'},
            'koleos': {'2020': '225/65R17', '2019': '225/65R17', '2021': '235/55R18', '2018': '225/65R17'},
        },
        # Opel
        'opel': {
            'corsa': {'2020': '185/65R15', '2019': '185/65R15', '2021': '195/55R16', '2018': '185/65R15'},
            'astra': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'insignia': {'2020': '215/60R16', '2019': '215/60R16', '2021': '235/45R18', '2018': '215/60R16'},
            'crossland': {'2020': '205/60R16', '2019': '205/60R16', '2021': '215/55R17', '2018': '205/60R16'},
            'grandland': {'2020': '215/65R16', '2019': '215/65R16', '2021': '235/50R18', '2018': '215/65R16'},
        },
        # Skoda
        'skoda': {
            'fabia': {'2020': '185/60R15', '2019': '185/60R15', '2021': '195/55R16', '2018': '185/60R15'},
            'scala': {'2020': '195/55R16', '2019': '195/55R16', '2021': '215/45R17', '2018': '195/55R16'},
            'octavia': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'superb': {'2020': '215/60R16', '2019': '215/60R16', '2021': '235/45R18', '2018': '215/60R16'},
            'kamiq': {'2020': '205/60R16', '2019': '205/60R16', '2021': '215/55R17', '2018': '205/60R16'},
            'karoq': {'2020': '215/60R17', '2019': '215/60R17', '2021': '235/50R18', '2018': '215/60R17'},
            'kodiaq': {'2020': '235/60R17', '2019': '235/60R17', '2021': '255/45R19', '2018': '235/60R17'},
        },
        # Seat
        'seat': {
            'ibiza': {'2020': '185/60R15', '2019': '185/60R15', '2021': '195/55R16', '2018': '185/60R15'},
            'leon': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'toledo': {'2020': '205/55R16', '2019': '205/55R16', '2018': '205/55R16'},
            'arona': {'2020': '205/60R16', '2019': '205/60R16', '2021': '215/55R17', '2018': '205/60R16'},
            'ateca': {'2020': '215/60R17', '2019': '215/60R17', '2021': '235/50R18', '2018': '215/60R17'},
            'tarraco': {'2020': '235/60R17', '2019': '235/60R17', '2021': '255/45R19', '2018': '235/60R17'},
        },
        # Fiat
        'fiat': {
            'panda': {'2020': '165/70R14', '2019': '165/70R14', '2021': '175/65R15', '2018': '165/70R14'},
            'punto': {'2020': '185/65R15', '2019': '185/65R15', '2018': '185/65R15'},
            'tipo': {'2020': '195/65R15', '2019': '195/65R15', '2021': '205/55R16', '2018': '195/65R15'},
            '500': {'2020': '185/55R15', '2019': '185/55R15', '2021': '195/45R16', '2018': '185/55R15'},
            '500x': {'2020': '205/60R16', '2019': '205/60R16', '2021': '215/55R17', '2018': '205/60R16'},
        },
        # Alfa Romeo
        'alfa romeo': {
            'mito': {'2020': '195/55R16', '2019': '195/55R16', '2018': '195/55R16'},
            'giulietta': {'2020': '205/55R16', '2019': '205/55R16', '2021': '225/45R17', '2018': '205/55R16'},
            'giulia': {'2020': '225/50R17', '2019': '225/50R17', '2021': '245/40R18', '2018': '225/50R17'},
            'stelvio': {'2020': '235/60R17', '2019': '235/60R17', '2021': '255/45R19', '2018': '235/60R17'},
        },
        # Mazda
        'mazda': {
            '2': {'2020': '185/65R15', '2019': '185/65R15', '2021': '195/55R16', '2018': '185/65R15'},
            '3': {'2020': '205/60R16', '2019': '205/60R16', '2021': '215/45R18', '2018': '205/60R16'},
            '6': {'2020': '215/55R17', '2019': '215/55R17', '2021': '225/45R19', '2018': '215/55R17'},
            'cx-3': {'2020': '215/60R16', '2019': '215/60R16', '2021': '215/55R17', '2018': '215/60R16'},
            'cx-5': {'2020': '225/65R17', '2019': '225/65R17', '2021': '235/55R19', '2018': '225/65R17'},
            'cx-30': {'2020': '215/60R16', '2019': '215/60R16', '2021': '215/55R17'},
        },
        # Subaru
        'subaru': {
            'impreza': {'2020': '205/55R16', '2019': '205/55R16', '2021': '215/50R17', '2018': '205/55R16'},
            'legacy': {'2020': '215/60R16', '2019': '215/60R16', '2021': '225/50R17', '2018': '215/60R16'},
            'outback': {'2020': '225/60R17', '2019': '225/60R17', '2021': '225/55R18', '2018': '225/60R17'},
            'forester': {'2020': '225/60R17', '2019': '225/60R17', '2021': '225/55R18', '2018': '225/60R17'},
            'xv': {'2020': '215/60R16', '2019': '215/60R16', '2021': '225/55R17', '2018': '215/60R16'},
        },
        # Mitsubishi
        'mitsubishi': {
            'mirage': {'2020': '165/65R14', '2019': '165/65R14', '2021': '175/55R15', '2018': '165/65R14'},
            'lancer': {'2020': '205/60R16', '2019': '205/60R16', '2018': '205/60R16'},
            'eclipse cross': {'2020': '225/55R18', '2019': '225/55R18', '2021': '225/50R19', '2018': '225/55R18'},
            'outlander': {'2020': '225/55R18', '2019': '225/55R18', '2021': '235/50R19', '2018': '225/55R18'},
            'asx': {'2020': '215/70R16', '2019': '215/70R16', '2021': '225/55R18', '2018': '215/70R16'},
        }
    }
    
    try:
        # Önce gerçek API'yi dene
        try:
            import requests
            
            # Wheel-Size API çağrısı
            api_key = 'demo'  # Gerçek API key buraya
            
            if generation:
                api_url = f"https://api.wheel-size.com/v2/search/?user_key={api_key}&make={make}&model={model}&year={year}&modification={generation}"
            else:
                api_url = f"https://api.wheel-size.com/v2/search/?user_key={api_key}&make={make}&model={model}&year={year}"
            
            headers = {
                'Accept': 'application/json',
                'User-Agent': 'MesTakip-TireComparison/1.0'
            }
            
            response = requests.get(api_url, headers=headers, timeout=5)
            
            if response.status_code == 200:
                data = response.json()
                
                if data.get('data') and len(data['data']) > 0:
                    # İlk sonucu al
                    vehicle = data['data'][0]
                    
                    # Stock lastik ara
                    stock_wheel = None
                    for wheel in vehicle.get('wheels', []):
                        if wheel.get('is_stock') == True and wheel.get('front', {}).get('tire'):
                            stock_wheel = wheel
                            break
                    
                    if stock_wheel:
                        front_tire = stock_wheel['front']['tire']
                        
                        # Lastik ölçüsünü parse et
                        tire_match = re.match(r'(\d+)/(\d+)R(\d+)', front_tire)
                        if tire_match:
                            width, profile, rim = map(int, tire_match.groups())
                            
                            # Çap hesaplama
                            sidewall_height = (width * profile) / 100
                            rim_diameter_mm = rim * 25.4
                            total_diameter = rim_diameter_mm + (2 * sidewall_height)
                            
                            # Araç bilgilerini hazırla
                            vehicle_info = {
                                'make': vehicle.get('make', {}).get('name', make.title()),
                                'model': vehicle.get('model', {}).get('name', model.title()),
                                'year': year,
                                'generation': vehicle.get('generation', {}).get('name', ''),
                                'trim': vehicle.get('trim', ''),
                                'engine': ''
                            }
                            
                            if vehicle.get('engine'):
                                engine = vehicle['engine']
                                capacity = engine.get('capacity', '')
                                fuel = engine.get('fuel', '')
                                if capacity and fuel:
                                    vehicle_info['engine'] = f"{capacity}L {fuel}"
                            
                            return JsonResponse({
                                'success': True,
                                'data': {
                                    'tire': front_tire,
                                    'width': width,
                                    'profile': profile,
                                    'rim': rim,
                                    'diameter': round(total_diameter, 1),
                                    'vehicle': vehicle_info
                                },
                                'source': 'wheel_size_api'
                            })
        
        except Exception as api_error:
            print(f"API Error: {api_error}")
            # API hatası durumunda mock data'ya geç
            pass
        
        # Mock data kullan
        model_clean = model.replace(' ', '').replace('-', '').lower()
        
        # Marka kontrolü
        if make not in mock_data:
            available_brands = ', '.join([brand.title() for brand in mock_data.keys()])
            return JsonResponse({
                'error': 'Marka bulunamadı',
                'message': f'"{make.title()}" markası için veri bulunamadı.',
                'available_brands': available_brands
            }, status=404)
        
        # Model kontrolü
        brand_data = mock_data[make]
        model_found = None
        for model_key in brand_data.keys():
            if model_key.replace(' ', '').replace('-', '').lower() == model_clean:
                model_found = model_key
                break
        
        if not model_found:
            available_models = ', '.join([m.title() for m in brand_data.keys()])
            return JsonResponse({
                'error': 'Model bulunamadı',
                'message': f'"{model.title()}" modeli için veri bulunamadı.',
                'available_models': available_models
            }, status=404)
        
        # Yıl kontrolü
        model_data = brand_data[model_found]
        if year not in model_data:
            available_years = ', '.join(sorted(model_data.keys(), reverse=True))
            return JsonResponse({
                'error': 'Yıl bulunamadı',
                'message': f'{year} yılı için veri bulunamadı.',
                'available_years': available_years
            }, status=404)
        
        # Veriyi al
        tire_size = model_data[year]
        
        # Lastik ölçüsünü parse et
        tire_match = re.match(r'(\d+)/(\d+)R(\d+)', tire_size)
        if not tire_match:
            return JsonResponse({
                'error': 'Lastik formatı hatası',
                'message': f'Lastik ölçüsü formatı tanınamadı: {tire_size}'
            }, status=400)
        
        width, profile, rim = map(int, tire_match.groups())
        
        # Çap hesaplama
        sidewall_height = (width * profile) / 100
        rim_diameter_mm = rim * 25.4
        total_diameter = rim_diameter_mm + (2 * sidewall_height)
        
        # Araç bilgilerini hazırla
        vehicle_info = {
            'make': make.title(),
            'model': model.title(),
            'year': year,
            'generation': generation or '',
            'trim': '',
            'engine': ''
        }
        
        return JsonResponse({
            'success': True,
            'data': {
                'tire': tire_size,
                'width': width,
                'profile': profile,
                'rim': rim,
                'diameter': round(total_diameter, 1),
                'vehicle': vehicle_info
            },
            'source': 'mock_data'
        })
        
    except Exception as e:
        return JsonResponse({
            'error': 'Beklenmeyen hata',
            'message': str(e)
        }, status=500)

@login_required
@misafir_forbidden
def excel_dosya_sil(request, dosya_id):
    """Excel dosyasını ve tüm kayıtlarını sil"""
    if request.method == 'POST':
        try:
            # MalzemeDosya modelini kullanarak dosyayı bul
            dosya = get_object_or_404(MalzemeDosya, id=dosya_id, kullanici=request.user)
            
            # Dosya adını mesaj için sakla
            dosya_adi = dosya.dosya_adi
            
            # İlişkili tüm satırları sil (CASCADE ile otomatik silinir)
            satirlar_count = dosya.satirlar.count()
            
            # Dosyayı sil
            dosya.delete()
            
            messages.success(
                request, 
                f'"{dosya_adi}" dosyası ve {satirlar_count} kayıt başarıyla silindi.'
            )
            
        except Exception as e:
            messages.error(request, f'Dosya silinirken hata oluştu: {str(e)}')
    
    # Geri yönlendirme
    next_url = request.POST.get('next', reverse('dashboard:products'))
    return redirect(next_url)


@login_required
@misafir_forbidden
def garanti_belgeleri(request):
    """Garanti Belgeleri Listesi"""
    from .models import GarantiBelgesi
    from datetime import date
    
    belgeler = GarantiBelgesi.objects.filter(olusturan=request.user).order_by('-olusturma_tarihi')
    
    context = {
        'belgeler': belgeler,
        'today': date.today(),
    }
    
    return render(request, 'dashboard/garanti_belgeleri.html', context)


@login_required
@misafir_forbidden
def garanti_belgesi_view(request, belge_id):
    """Garanti Belgesi Görüntüleme/Yazdırma"""
    from .models import GarantiBelgesi
    
    belge = get_object_or_404(GarantiBelgesi, id=belge_id, olusturan=request.user)
    
    context = {
        'belge': belge,
    }
    
    return render(request, 'dashboard/garanti_belgesi.html', context)


@login_required
@misafir_forbidden
def garanti_belgesi_edit(request, belge_id):
    """Garanti Belgesi Düzenleme"""
    from .models import GarantiBelgesi
    
    belge = get_object_or_404(GarantiBelgesi, id=belge_id, olusturan=request.user)
    
    context = {
        'belge': belge,
    }
    
    return render(request, 'dashboard/garanti_belgesi_edit.html', context)


@login_required
@require_http_methods(["POST"])
def save_garanti_belgesi(request):
    """Yeni Garanti Belgesi Kaydet"""
    from .models import GarantiBelgesi, GarantiBelgesiLastik
    from datetime import datetime
    import json
    import traceback
    
    # Misafir kontrolü - JSON response döndür
    if hasattr(request.user, 'userprofile') and request.user.userprofile.is_misafir():
        return JsonResponse({
            'success': False,
            'error': 'Bu işlem için yetkiniz yok.'
        }, status=403)
    
    try:
        # Debug log
        print("save_garanti_belgesi called")
        print("Request method:", request.method)
        print("Request body:", request.body[:200] if request.body else "Empty")
        
        data = json.loads(request.body)
        
        # Belge numarası oluştur
        today = datetime.now()
        prefix = f"GB{today.strftime('%Y%m')}"
        
        # Son belge numarasını bul
        last_belge = GarantiBelgesi.objects.filter(
            belge_no__startswith=prefix
        ).order_by('-belge_no').first()
        
        if last_belge:
            last_number = int(last_belge.belge_no[-4:])
            new_number = last_number + 1
        else:
            new_number = 1
        
        belge_no = f"{prefix}{new_number:04d}"
        
        # Garanti belgesi oluştur
        belge = GarantiBelgesi.objects.create(
            belge_no=belge_no,
            musteri_adi=data['musteri_adi'],
            musteri_telefon=data['musteri_telefon'],
            arac_plaka=data['arac_plaka'],
            arac_marka_model=data['arac_marka_model'],
            arac_yil=data.get('arac_yil', ''),
            arac_km=data.get('arac_km', ''),
            montaj_tarihi=data['montaj_tarihi'],
            notlar=data.get('notlar', ''),
            olusturan=request.user
        )
        
        # Lastikleri ekle
        for idx, lastik_data in enumerate(data['lastikler']):
            GarantiBelgesiLastik.objects.create(
                belge=belge,
                marka=lastik_data['marka'],
                ebat=lastik_data['ebat'],
                adet=lastik_data['adet'],
                fiyat=lastik_data.get('fiyat', 0),
                sira=idx
            )
        
        return JsonResponse({
            'success': True,
            'belge_no': belge_no,
            'belge_id': belge.id,
            'view_url': reverse('dashboard:garanti_belgesi_view', args=[belge.id])
        })
        
    except Exception as e:
        print("Error in save_garanti_belgesi:", str(e))
        print("Traceback:", traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def update_garanti_belgesi(request, belge_id):
    """Garanti Belgesi Güncelle"""
    from .models import GarantiBelgesi, GarantiBelgesiLastik
    import json
    
    # Misafir kontrolü - JSON response döndür
    if hasattr(request.user, 'userprofile') and request.user.userprofile.is_misafir():
        return JsonResponse({
            'success': False,
            'error': 'Bu işlem için yetkiniz yok.'
        }, status=403)
    
    try:
        belge = get_object_or_404(GarantiBelgesi, id=belge_id, olusturan=request.user)
        data = json.loads(request.body)
        
        # Belge bilgilerini güncelle
        belge.musteri_adi = data['musteri_adi']
        belge.musteri_telefon = data['musteri_telefon']
        belge.arac_plaka = data['arac_plaka']
        belge.arac_marka_model = data['arac_marka_model']
        belge.arac_yil = data.get('arac_yil', '')
        belge.arac_km = data.get('arac_km', '')
        belge.montaj_tarihi = data['montaj_tarihi']
        belge.notlar = data.get('notlar', '')
        belge.save()
        
        # Mevcut lastikleri sil
        belge.lastikler.all().delete()
        
        # Yeni lastikleri ekle
        for idx, lastik_data in enumerate(data['lastikler']):
            GarantiBelgesiLastik.objects.create(
                belge=belge,
                marka=lastik_data['marka'],
                ebat=lastik_data['ebat'],
                adet=lastik_data['adet'],
                fiyat=lastik_data.get('fiyat', 0),
                sira=idx
            )
        
        return JsonResponse({
            'success': True,
            'belge_no': belge.belge_no
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_http_methods(["POST"])
def delete_garanti_belgesi(request, belge_id):
    """Garanti Belgesi Sil"""
    from .models import GarantiBelgesi
    from django.http import Http404
    
    # Misafir kontrolü - JSON response döndür
    if hasattr(request.user, 'userprofile') and request.user.userprofile.is_misafir():
        return JsonResponse({
            'success': False,
            'error': 'Bu işlem için yetkiniz yok.'
        }, status=403)
    
    try:
        # get_object_or_404 yerine try-except kullan
        try:
            belge = GarantiBelgesi.objects.get(id=belge_id, olusturan=request.user)
        except GarantiBelgesi.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Garanti belgesi bulunamadı veya bu belgeyi silme yetkiniz yok.'
            }, status=404)
        
        belge_no = belge.belge_no
        belge.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'{belge_no} numaralı garanti belgesi silindi.'
        })
        
    except Exception as e:
        import traceback
        print(f"Delete garanti belgesi error: {e}")
        print(traceback.format_exc())
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@require_POST
def lastik_mevsim_ai(request):
    """
    Lastik ürün adı + marka metninden mevsim ve grup tahmini yapar.
    Öncelik sırası:
      1. Veritabanı öğrenmesi (daha önce girilen aynı/benzer ürünler)
      2. Kural tabanlı hızlı tespit
      3. Groq AI
    Döner: { "mevsim": "yaz"|"kis"|"dort-mevsim"|"bilinmiyor", "kaynak": "veritabani"|"kural"|"groq",
             "grup": "binek"|"ticari"|"aku"|"jant"|"bilinmiyor" }
    """
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}

    urun = (payload.get("urun") or "").strip()
    marka = (payload.get("marka") or "").strip()

    metin = f"{marka} {urun}".strip()
    if not metin:
        return JsonResponse({"mevsim": "bilinmiyor", "kaynak": "kural", "grup": "bilinmiyor"})

    # --- 1. LastikModelBilgisi tablosundan öğrenilmiş modellere bak ---
    def model_tablosundan_bul(urun_text, marka_text):
        try:
            tum_modeller = LastikModelBilgisi.objects.all().values('model_adi', 'mevsim')
            metin_lower = f"{marka_text} {urun_text}".lower()
            for kayit in tum_modeller:
                if kayit['model_adi'].lower() in metin_lower:
                    return kayit['mevsim']
        except Exception:
            pass
        return None

    model_sonuc = model_tablosundan_bul(urun, marka)
    if model_sonuc:
        # Grup tespiti kural tabanlı yapılır (veritabanında yok)
        pass  # grup tespiti aşağıda yapılacak

    # --- Kural tabanlı hızlı tespit ---
    def kural_tabani(text):
        t = text.lower()
        for ch, rep in [("ı","i"),("ş","s"),("ğ","g"),("ü","u"),("ö","o"),("ç","c")]:
            t = t.replace(ch, rep)

        dort_mevsim = ["all season","allseason","4season","4 seasons","4 mevsim","4mevsim",
                       "four season","dort mevsim","crossclimate","dueler a/t","a/t","at "]
        kis = ["winter","kis","kar","snow","ice","blizzak","alpin","ultragrip",
               "sottozero","wintercontact","winter contact","m+s","3pmsf","stud",
               "nordicmaster","colddriving","winguard","snowresponse",
               "ts870","ts860","ts850","ug8","ug9","lm005","w330","w462"]
        yaz = ["summer","yaz","eagle sport","pilot sport","sportcontact","premiumcontact",
               "premium contact","primacy","asymmetric","efficientgrip","energy saver",
               "ecopia","turanza","cinturato","ventus","potenza","re50","re71","hp010",
               "evo3","f1","ps4","ps5","sc6","sc7","elt30","sport 3","sport maxx",
               "pc6","pc7","ec6","ec7","cpc5","cpc6","t005","t001","k125","k435"]

        for h in dort_mevsim:
            if h in t:
                return "dort-mevsim"
        for h in kis:
            if h in t:
                return "kis"
        for h in yaz:
            if h in t:
                return "yaz"
        return None

    def kural_grup(text):
        t = text.lower().strip()
        import re
        # Ebat sonunda C varsa ticari: 215/65R16C, 195/70R15C, vb.
        if re.search(r'\d+/\d+r\d+c\b', t):
            return "ticari"
        # Akü ve jant anahtar kelimeleri
        aku_kw = ["aku", "akumulator", "battery", "agm", "efb", "calcium", "start stop"]
        jant_kw = ["jant", "alaşim", "alasim", "aluminyum", "rim", "wheel"]
        for h in aku_kw:
            if h in t:
                return "aku"
        for h in jant_kw:
            if h in t:
                return "jant"
        # Standart ebat varsa (C yok) → binek
        if re.search(r'\d+/\d+r\d+\b', t):
            return "binek"
        return None

    kural_mevsim_sonuc = kural_tabani(metin)
    kural_grup_sonuc = kural_grup(metin)

    # Eğer veritabanında mevsim bulduysa, sadece grup için kural/AI kullan
    if model_sonuc:
        if kural_grup_sonuc:
            return JsonResponse({"mevsim": model_sonuc, "kaynak": "veritabani", "grup": kural_grup_sonuc})
        # Grup bilinmiyorsa Groq'a sor (aşağıda)

    # --- Groq ile tespit ---
    api_key = getattr(settings, "GROQ_API_KEY", None) or os.environ.get("GROQ_API_KEY", "")
    if not api_key:
        try:
            from dotenv import load_dotenv
            load_dotenv(getattr(settings, "BASE_DIR", None) / ".env")
        except Exception:
            pass
        api_key = os.environ.get("GROQ_API_KEY", "")

    if api_key:
        try:
            groq_yanit = groq_chat_completion(
                api_key=api_key,
                temperature=0.0,
                max_tokens=60,
                messages=[
                    {
                        "role": "system",
                        "content": (
                            "Sen bir lastik/akü/jant uzmanı AI'sın. Verilen ürün markası ve modeline bakarak mevsim ve grup türünü tahmin et.\n"
                            "SADECE ve KESİNLİKLE JSON formatında yanıt ver. Markdown (```) veya ek metin kullanma.\n"
                            "'mevsim' değerleri: 'yaz', 'kis', 'dort-mevsim', 'bilinmiyor'.\n"
                            "'grup' değerleri: 'binek', 'ticari', 'aku', 'jant', 'bilinmiyor'.\n"
                            "İpuçları:\n"
                            "- Sport, Primacy, Turanza, Eagle, PC6, PS4 -> mevsim:yaz, grup:binek\n"
                            "- Winter, Snow, Blizzak, TS870, LM005 -> mevsim:kis, grup:binek\n"
                            "- All Season, CrossClimate, A/T -> mevsim:dort-mevsim\n"
                            "- Van, Transporter, LT, C (ticari ebat: 195/70R15C gibi) -> grup:ticari\n"
                            "- Akü, Battery, AGM, EFB -> mevsim:bilinmiyor, grup:aku\n"
                            "- Jant, Alaşım, Rim -> mevsim:bilinmiyor, grup:jant\n\n"
                            "Örnek Yanıt:\n"
                            "{\"mevsim\": \"yaz\", \"grup\": \"binek\"}"
                        ),
                    },
                    {"role": "user", "content": metin},
                ],
            )
            groq_yanit = (groq_yanit or "").strip()

            # Markdown etiketlerini temizle
            if groq_yanit.startswith("```json"):
                groq_yanit = groq_yanit[7:]
            elif groq_yanit.startswith("```"):
                groq_yanit = groq_yanit[3:]
            if groq_yanit.endswith("```"):
                groq_yanit = groq_yanit[:-3]

            groq_yanit = groq_yanit.strip()

            try:
                parsed = json.loads(groq_yanit)
                mevsim = parsed.get("mevsim", "bilinmiyor")
                grup = parsed.get("grup", "bilinmiyor")
                if mevsim not in ("yaz", "kis", "dort-mevsim", "bilinmiyor"):
                    mevsim = "bilinmiyor"
                if grup not in ("binek", "ticari", "aku", "jant", "bilinmiyor"):
                    grup = "bilinmiyor"
                # Kural tabanlı sonuçlar varsa onları öncelik ver
                return JsonResponse({
                    "mevsim": kural_mevsim_sonuc or mevsim,
                    "grup": kural_grup_sonuc or ("binek" if (kural_mevsim_sonuc or mevsim) not in ("bilinmiyor",) else grup),
                    "kaynak": "groq"
                })
            except Exception:
                pass
        except Exception:
            pass

    # Groq başarısız ya da yoksa kural sonuçlarını kullan
    mevsim_sonuc = kural_mevsim_sonuc or (model_sonuc or "bilinmiyor")
    grup_sonuc = kural_grup_sonuc or ("binek" if mevsim_sonuc != "bilinmiyor" else "bilinmiyor")
    return JsonResponse({
        "mevsim": mevsim_sonuc,
        "grup": grup_sonuc,
        "kaynak": "kural" if not model_sonuc else "veritabani"
    })


